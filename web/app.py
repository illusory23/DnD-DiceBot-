"""尘封之卷 — Flask Web UI"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import os as _os
import io as _io
import builtins
import re as _re
from pathlib import Path as _Path

from flask import Flask, render_template, request, jsonify, send_file, abort, redirect, make_response as _make_response, session as _flask_session
from datetime import datetime
from core.database import db
from utils.logger import (
    get_logger, setup_request_logging,
    log_frontend_error,
    audit_log,
)
from core.dice_engine import roll, roll_ability_check
from core.dnd5e_rules import (
    ability_modifier, get_ability_for_skill, normalize_ability,
    normalize_skill, ABILITY_ORDER, SKILL_TO_ABILITY,
)
from core.character import (
    create_character, get_character, list_characters,
    set_ability, adjust_hp, use_spell_slot,
    long_rest,
    update_character, delete_character,
    resolve_portrait_path,
    list_character_groups, create_character_group,
    update_character_group, delete_character_group,
    set_character_group,
)
from utils.data_loader import search_spell, search_monster, load_conditions
from core.chm_search import (
    search_monster as chm_search_monster,
    get_monster_detail as chm_get_monster_detail,
    get_spell_detail as chm_get_spell_detail,
)
from utils.formatter import (
    format_dice_result, format_character_sheet,
)
from utils.excel_importer import import_character_from_excel, import_and_print_summary
from core.character import import_from_excel_data

# ━━━ 导入随机事件表命令函数 ━━━
import dnd_bot as _bot

# 事件命令映射: 命令名 → (标签, 骰子, 命令函数)
_EVENT_COMMANDS = {
    'qsj100': ('全随机事件', 'd100', _bot.cmd_qsj100),
    'zy100':  ('随机遭遇事件', 'd100', _bot.cmd_zy100),
    'ts100':  ('随机探索事件', 'd100', _bot.cmd_ts100),
    'dc100':  ('随机调查事件', 'd100', _bot.cmd_dc100),
    'hj6':    ('随机环境', 'd6', _bot.cmd_hj6),
    'sj6':    ('随机特殊事件', 'd6', _bot.cmd_sj6),
    'rl100':  ('随机人类', 'd100', _bot.cmd_rl100),
    'ys10':   ('随机野兽', 'd10', _bot.cmd_ys10),
    'zl12':   ('随机非敌对生物', 'd12', _bot.cmd_zl12),
    'sw100':  ('随机特殊生物', 'd100', _bot.cmd_sw100),
    'cl50':   ('随机材料', 'd50', _bot.cmd_cl50),
    'cl100':  ('随机特殊材料', 'd100', _bot.cmd_cl100),
    'kw100':  ('随机矿物', 'd100', _bot.cmd_kw100),
    'fx100':  ('随机特殊发现', 'd100', _bot.cmd_fx100),
    'yj8':    ('随机遗迹', 'd8', _bot.cmd_yj8),
    'wp':     ('物品表', '', _bot.cmd_wp),
}

# 事件分组（用于前端 UI 布局）
_EVENT_GROUPS = [
    ('🗺️ 主事件表', [
        ('qsj100', '全随机事件', 'd100', '综合所有事件类型，一骰定乾坤'),
        ('zy100', '随机遭遇事件', 'd100', '1-20环境 21-30人类 31-55野兽 56-60迷失 61-80友善 81-85恶霜 86-95特殊 96-100特殊生物'),
        ('ts100', '随机探索事件', 'd100', '1-20深入 21-40遭遇 41-60采集 61-85调查 86-95修整 96-100返回'),
        ('dc100', '随机调查事件', 'd100', '1-10特殊发现 11-35踪迹 36-60异常 61-85材料 86-100无获'),
    ]),
    ('🌍 环境与遭遇', [
        ('hj6', '随机环境', 'd6', '1雪坑 2暴风雪 3极光 4霜雾 5冻雨 6暖风'),
        ('sj6', '随机特殊事件', 'd6', '1雪崩 2冰层破裂 3冰层震动 4陷阱 5特殊标记 6寒鸦报信'),
        ('rl100', '随机人类', 'd100', '1-15商队 16-45公会 46-65士兵 66-100猎人'),
        ('ys10', '随机野兽', 'd10', '1-10 冰原狼/野猪/白熊/巨熊/巨鹰/寒脊蛇'),
    ]),
    ('👹 生物', [
        ('zl12', '随机非敌对生物', 'd12', '1-12 雪蹄兔/野鹿/雪鸮/寒鸦/雪狐/霜鼠/麝牛/银蛛/霜羽雉/绒蜂巢'),
        ('sw100', '随机特殊生物', 'd100', '1-5霜巨魔 6-45恶尸 46-60冬精灵 61-92霜灵 93-95瘦鹿 96-100水星守卫'),
    ]),
    ('⛏ 资源与材料', [
        ('cl50', '随机材料', 'd50', '1-6药草 7-15浆果 16-23松枝 24-30松脂 31-35兽骨 36-42矿粒 43-46钱袋 47-50特殊'),
        ('cl100', '随机特殊材料', 'd100', '1-25荧光苔 26-45松茸 46-65铁松松脂 66-72绒蜂蜜 73-85蛛网 86-95霜晶核 96-100寒铁髓'),
        ('kw100', '随机矿物', 'd100', '1-10岩盐 11-25煤矿 26-35铜矿 36-50铁矿 51-60寒铁矿 61-70银矿 71-75金矿 76-85冰水晶 86-92霜晶核 93-96化石 97-100宝石'),
    ]),
    ('🏛 探索与发现', [
        ('fx100', '随机特殊发现', 'd100', '1-5遗迹 6-15灵泉 16-40猎人小屋 41-55温泉 56-85岩洞 86-100矿坑'),
        ('yj8', '随机遗迹', 'd8', '1远古战场 2巨兽骨骸 3古老墓穴 4符文法阵 5废弃神殿 6残破石碑 7倒塌建筑 8神秘祭坛'),
    ]),
    ('📦 参考', [
        ('wp', '物品表 (1-29)', '', '查询或列出全部29项物品'),
    ]),
]

app = Flask(__name__)
# SECRET_KEY：优先环境变量，否则从 data/secret_key 读取（不存在则生成持久化随机值）
# 避免使用公开的硬编码默认值（否则攻击者可伪造 session 提权）
def _load_secret_key() -> str:
    env_key = os.environ.get('SECRET_KEY')
    if env_key:
        return env_key
    key_file = _Path(__file__).parent.parent / 'data' / 'secret_key'
    try:
        key_file.parent.mkdir(parents=True, exist_ok=True)
        if key_file.exists():
            key = key_file.read_text(encoding='utf-8').strip()
            if key:
                return key
        key = os.urandom(32).hex()
        key_file.write_text(key, encoding='utf-8')
        return key
    except Exception:
        return os.urandom(32).hex()

app.secret_key = _load_secret_key()
# 请求体大小上限（资源上传最大 50MB，留余量；防止恶意超大 JSON 耗尽内存）
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

# ━━━ 管理员后台 Blueprint ━━━
from web.admin import admin_bp
app.register_blueprint(admin_bp)

# ━━━ SQLAlchemy ORM 初始化 ━━━
from core.database import init_db as _init_orm_db
_init_orm_db(app)  # 创建所有数据库表（含 users 表）


def _ensure_user_columns():
    """旧库补列：users.is_admin / is_active / last_login_ip / avatar_url（幂等）。

    已有 users 表不会因 create_all 加列，这里手动 ALTER TABLE 补齐。
    is_admin 曾缺失导致任何 User 查询报错，必须最先补上。
    """
    try:
        from sqlalchemy import inspect as _sa_inspect
        cols = {c['name'] for c in _sa_inspect(db.engine).get_columns('users')}
        with db.engine.begin() as conn:
            if 'is_admin' not in cols:
                conn.exec_driver_sql(
                    "ALTER TABLE users ADD COLUMN is_admin BOOLEAN NOT NULL DEFAULT 0")
            if 'is_active' not in cols:
                conn.exec_driver_sql(
                    "ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1")
            if 'last_login_ip' not in cols:
                conn.exec_driver_sql(
                    "ALTER TABLE users ADD COLUMN last_login_ip VARCHAR(64) NOT NULL DEFAULT ''")
            if 'avatar_url' not in cols:
                conn.exec_driver_sql(
                    "ALTER TABLE users ADD COLUMN avatar_url TEXT NOT NULL DEFAULT ''")
    except Exception as _e:
        _app_logger.warning(f'补列 users 表失败: {_e}')


def _ensure_character_columns():
    """旧库补列：characters.is_public（DM 公开角色，幂等）。"""
    try:
        from sqlalchemy import inspect as _sa_inspect
        cols = {c['name'] for c in _sa_inspect(db.engine).get_columns('characters')}
        if 'is_public' not in cols:
            with db.engine.begin() as conn:
                conn.exec_driver_sql(
                    "ALTER TABLE characters ADD COLUMN is_public BOOLEAN NOT NULL DEFAULT FALSE")
            _app_logger.info('已补列 characters.is_public')
    except Exception as _e:
        _app_logger.warning(f'补列 characters.is_public 失败: {_e}')

# ━━━ 静态文件缓存 ━━━
from werkzeug.serving import make_server
# 本地工具场景优先保证代码改动立即可见：JS/CSS 不做长缓存，避免浏览器沿用旧版脚本
# 导致功能（如3D掷骰）异常；如后续需要性能可改为带版本号的资源引用。
@app.after_request
def add_cache_header(response):
    if response.content_type and ('text/css' in response.content_type or 'application/javascript' in response.content_type or 'text/html' in response.content_type):
        response.cache_control.max_age = 0
        response.cache_control.no_cache = True
        response.cache_control.must_revalidate = True
    return response

# ━━━ WebSocket 实时协作（flask-sock，与 HTTP 共用 5000 端口）━━━
from flask_sock import Sock
sock = Sock(app)

# 当前活跃角色 (简易会话，正式版用 Flask session)
active_char_id = None

# ━━━ 聊天系统 ━━━
import time as _time
import json as _json
import threading as _threading

_chat_messages: list[dict] = []  # [{name, text, time, is_dm, ip, color}]
MAX_CHAT_MSGS = 500  # 最多保留500条消息

# ━━━ 酒馆聊天系统（独立于主聊天室）━━━
_tavern_messages: list[dict] = []
MAX_TAVERN_CHAT_MSGS = 500

def _load_chat_log():
    """从 PostgreSQL 加载聊天记录（2026-08-15 迁移，原 chat_log.json）"""
    global _chat_messages
    try:
        from core.models import ChatMessage as _CM
        rows = _CM.query.filter_by(channel='chat').order_by(_CM.id).all()
        _chat_messages = [r.data if isinstance(r.data, dict) else {
            'name': r.name, 'text': r.text, 'time': r.time, 'is_dm': r.is_dm,
            'color': r.color, 'role': r.role, 'ip': r.ip, '_ts': r.ts,
        } for r in rows]
    except Exception:
        _chat_messages = []

def _save_chat_log():
    """保存聊天记录到 PostgreSQL（全删全插保留 id 语义）。

    异步线程调用时无 Flask context，函数内自建 app context。
    """
    try:
        with app.app_context():
            from core.models import ChatMessage as _CM
            msgs = _chat_messages[-MAX_CHAT_MSGS:]
            with db.engine.begin() as conn:
                conn.execute(_CM.__table__.delete().where(_CM.channel == 'chat'))
                for m in msgs:
                    conn.execute(_CM.__table__.insert().values(
                        channel='chat', name=m.get('name', ''), text=m.get('text', ''),
                        time=m.get('time', ''), is_dm=bool(m.get('is_dm')),
                        color=m.get('color', ''), role=m.get('role', ''), ip=m.get('ip', ''),
                        ts=m.get('_ts', 0) or 0, event_id=m.get('event_id'),
                        recalled=bool(m.get('_recalled')), data=m))
    except Exception as _e:
        _app_logger.warning(f'_save_chat_log 失败: {_e}')

def _archive_chat_messages(msgs_to_archive: list[dict]):
    """将超出限制的旧消息归档到 chat_archives 表（原 chat_archive/ 目录）。

    每次归档插入一行（filename 带时间戳），并清理30天前的归档行。
    """
    if not msgs_to_archive:
        return
    try:
        with app.app_context():
            from core.models import ChatArchive as _CA
            ts = _time.strftime('%Y%m%d_%H%M%S')
            created = _time.strftime('%Y-%m-%d %H:%M:%S')
            with db.engine.begin() as conn:
                conn.execute(_CA.__table__.insert().values(
                    channel='chat', filename=f'chat_{ts}.json',
                    messages=msgs_to_archive, created_at=created))
                # 清理30天前的归档行
                cutoff = (_time.time() - 30 * 86400)
                cutoff_str = _time.strftime('%Y-%m-%d %H:%M:%S', _time.localtime(cutoff))
                conn.execute(_CA.__table__.delete().where(_CA.channel == 'chat', _CA.created_at < cutoff_str))
    except Exception as _e:
        _app_logger.warning(f'_archive_chat_messages 失败: {_e}')

def _trim_and_archive_chat():
    """裁剪聊天消息：超过 MAX_CHAT_MSGS 时将旧消息归档。"""
    global _chat_messages
    if len(_chat_messages) > MAX_CHAT_MSGS:
        overflow = len(_chat_messages) - MAX_CHAT_MSGS
        to_archive = _chat_messages[:overflow]
        _chat_messages = _chat_messages[-MAX_CHAT_MSGS:]
        # 异步归档旧消息
        _threading.Thread(target=_archive_chat_messages, args=(to_archive,), daemon=True).start()

# ━━━ 酒馆聊天持久化 ━━━

def _load_tavern_chat_log():
    """从 PostgreSQL 加载酒馆聊天记录（原 tavern_chat_log.json）"""
    global _tavern_messages
    try:
        from core.models import ChatMessage as _CM
        rows = _CM.query.filter_by(channel='tavern').order_by(_CM.id).all()
        _tavern_messages = [r.data if isinstance(r.data, dict) else {
            'name': r.name, 'text': r.text, 'time': r.time,
            'color': r.color, 'role': r.role, '_ts': r.ts,
        } for r in rows]
    except Exception:
        _tavern_messages = []

def _save_tavern_chat_log():
    """保存酒馆聊天记录到 PostgreSQL（异步线程调用，函数内自建 app context）"""
    try:
        with app.app_context():
            from core.models import ChatMessage as _CM
            msgs = _tavern_messages[-MAX_TAVERN_CHAT_MSGS:]
            with db.engine.begin() as conn:
                conn.execute(_CM.__table__.delete().where(_CM.channel == 'tavern'))
                for m in msgs:
                    conn.execute(_CM.__table__.insert().values(
                        channel='tavern', name=m.get('name', ''), text=m.get('text', ''),
                        time=m.get('time', ''), is_dm=bool(m.get('is_dm')),
                        color=m.get('color', ''), role=m.get('role', ''), ip=m.get('ip', ''),
                        ts=m.get('_ts', 0) or 0, event_id=m.get('event_id'),
                        recalled=bool(m.get('_recalled')), data=m))
    except Exception as _e:
        _app_logger.warning(f'_save_tavern_chat_log 失败: {_e}')

def _archive_tavern_chat_messages(msgs_to_archive: list[dict]):
    """将超出限制的酒馆旧消息归档到 chat_archives 表（原 tavern_chat_archive/ 目录）。"""
    if not msgs_to_archive:
        return
    try:
        with app.app_context():
            from core.models import ChatArchive as _CA
            ts = _time.strftime('%Y%m%d_%H%M%S')
            created = _time.strftime('%Y-%m-%d %H:%M:%S')
            with db.engine.begin() as conn:
                conn.execute(_CA.__table__.insert().values(
                    channel='tavern', filename=f'tavern_{ts}.json',
                    messages=msgs_to_archive, created_at=created))
                cutoff = (_time.time() - 30 * 86400)
                cutoff_str = _time.strftime('%Y-%m-%d %H:%M:%S', _time.localtime(cutoff))
                conn.execute(_CA.__table__.delete().where(_CA.channel == 'tavern', _CA.created_at < cutoff_str))
    except Exception as _e:
        _app_logger.warning(f'_archive_tavern_chat_messages 失败: {_e}')

def _trim_and_archive_tavern_chat():
    """裁剪酒馆聊天消息：超过 MAX_TAVERN_CHAT_MSGS 时将旧消息归档。"""
    global _tavern_messages
    if len(_tavern_messages) > MAX_TAVERN_CHAT_MSGS:
        overflow = len(_tavern_messages) - MAX_TAVERN_CHAT_MSGS
        to_archive = _tavern_messages[:overflow]
        _tavern_messages = _tavern_messages[-MAX_TAVERN_CHAT_MSGS:]
        # 异步归档旧消息
        _threading.Thread(target=_archive_tavern_chat_messages, args=(to_archive,), daemon=True).start()

# 启动时加载（ORM 查询需要 Flask app context）
with app.app_context():
    _load_chat_log()
    _load_tavern_chat_log()

# ━━━ 日志系统初始化 ━━━
_app_logger = get_logger('dicebot')

# 旧库补列（is_active / last_login_ip），依赖 _app_logger，需在应用上下文内执行
with app.app_context():
    _ensure_user_columns()
    _ensure_character_columns()

# ━━━ DM/主机系统 ━━━
_dm_name: str | None = None      # 当前DM的显示名
_dm_ip: str | None = None        # DM的IP地址
_dm_user_id: int | None = None   # DM的官网用户ID（改名不丢失DM身份）

# ━━━ 在线用户/房间系统 ━━━
# {name: {'ip': str, 'color': str, 'role': str, 'last_heartbeat': float, 'joined_at': float}}
_online_users: dict[str, dict] = {}

# ━━━ 全站在线会话（打开任意页面即在线，关闭全部页面才下线）━━━
# {session_id: {'username': str, 'user_id': int, 'ip': str, 'page': str, 'last_seen': float}}
# session_id 由每个浏览器标签页生成，多标签 = 多会话；全部关闭即下线。
_online_sessions: dict[str, dict] = {}
_online_lock = _threading.Lock()
ONLINE_SESSION_TIMEOUT = 60  # 心跳超时秒数（异常关闭页面兜底清理）

# ━━━ @提及通知系统 ━━━
_mentions: list[dict] = []  # [{target_name, from_name, text, time, _ts}]
MAX_MENTIONS = 100

# ━━━ DM 事件系统（聊天室发布事件 / 全平台弹窗 / DM事件列表）━━━
_dm_event_list: list[dict] = []  # [{id, title, content, created_at}] 仅DM可见，持久化
_dm_published_events: list[dict] = []  # [{id, title, content, published_at, recalled_at, pinned}] 发布历史，持久化
_event_notifications: list[dict] = []  # [{event_id, title, content, time, _ts, recalled, pinned}] 发布时广播，供各页面轮询弹窗
MAX_EVENT_NOTIFICATIONS = 50

def _load_dm_event_list():
    """启动时从 PostgreSQL 加载DM事件列表（原 dm_event_list.json）"""
    global _dm_event_list
    try:
        from core.models import DmEvent as _DE
        rows = _DE.query.order_by(_DE.id).all()
        _dm_event_list = [{'id': r.id, 'title': r.title, 'content': r.content,
                           'created_at': r.created_at} for r in rows]
    except Exception:
        _dm_event_list = []

def _save_dm_event_list():
    """保存DM事件列表到 PostgreSQL（全删全插保留 id）"""
    try:
        from core.models import DmEvent as _DE
        with db.engine.begin() as conn:
            conn.execute(_DE.__table__.delete())
            for e in _dm_event_list:
                conn.execute(_DE.__table__.insert().values(
                    id=e.get('id'), title=e.get('title', ''),
                    content=e.get('content', ''), created_at=e.get('created_at', '')))
    except Exception:
        pass

with app.app_context():
    _load_dm_event_list()

def _load_dm_published_events():
    """启动时从 PostgreSQL 加载DM已发布事件历史（原 dm_published_events.json）"""
    global _dm_published_events
    try:
        from core.models import PublishedDmEvent as _PE
        rows = _PE.query.order_by(_PE.id).all()
        _dm_published_events = [{'id': r.id, 'title': r.title, 'content': r.content,
                                 'published_at': r.published_at, 'recalled_at': r.recalled_at,
                                 'pinned': bool(r.pinned)} for r in rows]
    except Exception:
        _dm_published_events = []

def _save_dm_published_events():
    """保存DM已发布事件历史到 PostgreSQL"""
    try:
        from core.models import PublishedDmEvent as _PE
        with db.engine.begin() as conn:
            conn.execute(_PE.__table__.delete())
            for e in _dm_published_events:
                conn.execute(_PE.__table__.insert().values(
                    id=e.get('id'), title=e.get('title', ''),
                    content=e.get('content', ''), published_at=e.get('published_at', ''),
                    recalled_at=e.get('recalled_at'), pinned=bool(e.get('pinned'))))
    except Exception:
        pass

with app.app_context():
    _load_dm_published_events()

# ━━━ 共享画布状态（WebSocket 实时为主 + HTTP 轮询降级）━━
from .shared_state import (
    get_shared_canvas, get_shared_canvas_ts, update_shared_canvas,
    append_shared_strokes, apply_incremental, remove_shared_strokes,
    clear_shared_canvas, get_state_snapshot, get_version, save_layer_image,
    apply_fog_incremental,
)

_ws_connected: set = set()
_ws_lock = _threading.Lock()
_canvas_meta: dict = {'width': 5000, 'height': 5000}  # 画布尺寸元数据（不持久化）


def _ws_broadcast(payload: str, exclude=None):
    """向所有 WebSocket 客户端广播；发送失败的死连接立即清理。"""
    with _ws_lock:
        clients = list(_ws_connected)
    for client in clients:
        if client is exclude:
            continue
        try:
            client.send(payload)
        except Exception:
            with _ws_lock:
                _ws_connected.discard(client)


def _ws_send_init(ws):
    """向单个客户端发送全量初始状态（含版本号）。"""
    state = get_state_snapshot()
    state['canvas'] = dict(_canvas_meta)
    ws.send(_json.dumps({'type': 'init', 'state': state, 'version': state.get('_ver', 0)}))


def _apply_canvas_message(data: dict) -> tuple[int | None, bool]:
    """将一条画布消息应用到权威状态。

    返回 (新版本号或None, 是否需要广播)。
    """
    global _canvas_meta
    action_type = data.get('type')

    if action_type == 'stroke':
        return append_shared_strokes([data.get('data', {})]), True
    if action_type == 'strokes_remove':
        return remove_shared_strokes(data.get('data', [])), True
    if action_type == 'strokes_clear':
        return update_shared_canvas('strokes', []), True
    if action_type == 'op':
        # 操作语义消息: {type:'op', key, upsert:[...], remove:[ids]}
        payload = data.get('data', {}) if isinstance(data.get('data'), dict) else data
        key = payload.get('key')
        if key in ('layers', 'tokens', 'texts', 'fog'):
            ver = apply_incremental(key, payload.get('upsert') or [],
                                    payload.get('remove') or [])
            return ver, True
        return None, False
    if action_type in ('layers_update', 'tokens_update', 'texts_update', 'fog_update'):
        key = action_type.split('_')[0]
        return update_shared_canvas(key, data.get('data', [])), True
    if action_type == 'canvas_update':
        _canvas_meta = data.get('data', {'width': 5000, 'height': 5000})
        return get_version(), True
    if action_type == 'clear_all':
        ver = clear_shared_canvas()
        _canvas_meta = {'width': 5000, 'height': 5000}
        return ver, True
    return None, False


# ━━━ 确定性骰子帧同步消息处理 ━━━
def _handle_dice_sync_message(data: dict, ws) -> bool:
    """处理骰子帧同步消息。返回 True 表示已处理（不应继续走画布逻辑）。"""
    msg_type = data.get('type', '')

    if msg_type == 'dice_roll_result':
        print(f'[dice-sync] result: {data.get("roller")} {data.get("notation")}={data.get("total")}')
        _ws_broadcast(_json.dumps(data), exclude=ws)
        return True

    if msg_type == 'dice_roll_animate' or msg_type == 'dice_roll_det':
        _ws_broadcast(_json.dumps(data), exclude=ws)
        return True

    if msg_type == 'dice_roll':
        # 投掷者发起掷骰 → 广播同步开始给所有其他客户端
        seed = data.get('seed')
        dice = data.get('dice', [])
        notation = data.get('notation', '')
        roller = data.get('roller', 'anonymous')

        sync_start = _json.dumps({
            'type': 'dice_sync_start',
            'seed': seed,
            'dice': dice,
            'notation': notation,
            'roller': roller,
            'frameStart': 0,
            'timestamp': _time.time(),
        })
        # 广播给所有客户端（包括投掷者自己，用于确认）
        _ws_broadcast(sync_start)
        print(f'[dice-sync] 掷骰开始: {notation} (种子:{seed}, 投掷者:{roller})')
        return True

    if msg_type == 'dice_result':
        # 投掷者本地模拟完成 → 广播权威结果
        results = data.get('results', [])
        total = data.get('total', 0)
        seed = data.get('seed')
        notation = data.get('notation', '')
        roller = data.get('roller', 'anonymous')

        sync_result = _json.dumps({
            'type': 'dice_sync_result',
            'results': results,
            'total': total,
            'seed': seed,
            'notation': notation,
            'roller': roller,
            'timestamp': _time.time(),
        })
        _ws_broadcast(sync_result)
        print(f'[dice-sync] 掷骰结果: {notation} = {total} (种子:{seed})')
        return True

    if msg_type == 'dice_sync_frame':
        # 周期性帧号广播（可选：用于锁步同步校准）
        frame = data.get('frame', 0)
        sync_frame = _json.dumps({
            'type': 'dice_sync_frame',
            'frame': frame,
            'checksum': data.get('checksum', ''),
        })
        _ws_broadcast(sync_frame, exclude=ws)
        return True

    return False


@sock.route('/ws')
def ws_canvas(ws):
    """WebSocket 画布同步 + 骰子帧同步，与 HTTP 共用 5000 端口。frp 单隧道兼容。"""
    with _ws_lock:
        _ws_connected.add(ws)
    try:
        _ws_send_init(ws)

        while True:
            raw = ws.receive()
            if raw is None:
                break
            try:
                data = _json.loads(raw)
            except (ValueError, TypeError):
                continue
            try:
                # 断线重连后的状态对账：版本不一致则单发全量
                if data.get('type') == 'sync':
                    if int(data.get('version', -1)) != get_version():
                        _ws_send_init(ws)
                    continue

                # ━━ 骰子帧同步消息 ━━
                if _handle_dice_sync_message(data, ws):
                    continue

                # ━━ 战斗状态同步 ━━
                if data.get('type') == 'combat_update':
                    state = data.get('state', {})
                    if isinstance(state, dict) and 'combatants' in state:
                        global _combat_state, _combat_state_ts
                        # 版本保护：客户端基于旧状态的推送不覆盖新状态（防网络延迟回弹）
                        base_ts = float(state.get('_ts', 0) or 0)
                        if not state.get('_force') and base_ts < _combat_state_ts - 0.001:
                            state['_ts'] = _combat_state_ts
                            state['_conflict'] = True
                            try:
                                ws.send(_json.dumps({
                                    'type': 'combat_update_conflict',
                                    'state': _combat_state,
                                    'timestamp': _combat_state_ts,
                                }))
                            except Exception:
                                pass
                            continue
                        _combat_state = state
                        _combat_state_ts = _time.time()
                        state['_ts'] = _combat_state_ts
                        _save_combat_state()
                        _ws_broadcast(_json.dumps(data), exclude=ws)
                        # 给发起者回确认（携带服务器时间戳），让本地时间戳与服务器同步
                        try:
                            ws.send(_json.dumps({
                                'type': 'combat_update_ack',
                                'timestamp': _combat_state_ts,
                                'state': state,
                                '_cid': data.get('_cid'),
                            }))
                        except Exception:
                            pass
                    continue

                # ━━ 迷雾增量同步（雾笔/擦除轨迹按 id 合并，避免全量重传）━━
                if data.get('type') == 'fog_incr':
                    d = data.get('data', {})
                    ver = apply_fog_incremental(
                        d.get('strokes') or [],
                        d.get('erasures') or [],
                        d.get('removeStrokes') or [],
                        d.get('removeErasures') or [],
                    )
                    if ver is not None:
                        data['_ver'] = ver
                    _ws_broadcast(_json.dumps(data), exclude=ws)
                    continue

                # ━━ 画布同步消息 ━━
                ver, should_broadcast = _apply_canvas_message(data)
                if should_broadcast:
                    if ver is not None:
                        data['_ver'] = ver
                    # 图层消息剥离 base64（图片应走 /api/shared-canvas/layer-image 上传）
                    if data.get('type') == 'layers_update':
                        for l in data.get('data', []):
                            if isinstance(l, dict) and l.get('url') and l.get('dataURL'):
                                l['dataURL'] = ''
                    _ws_broadcast(_json.dumps(data), exclude=ws)
            except Exception as e:
                print(f'[ws] 处理消息失败: {type(e).__name__}: {e}')
    finally:
        with _ws_lock:
            _ws_connected.discard(ws)


@app.route('/api/shared-canvas', methods=['GET'])
def api_get_shared_canvas():
    """获取共享画布状态。

    支持 ?since=<时间戳> 或 ?since_ver=<版本号>；未变化时只返回轻量应答，
    不携带全量状态（P1-2：降低轮询流量）。
    """
    ts = get_shared_canvas_ts()
    ver = get_version()

    since_ver = request.args.get('since_ver')
    if since_ver is not None:
        try:
            unchanged = int(since_ver) == ver
        except ValueError:
            unchanged = False
    else:
        try:
            since_ts = float(request.args.get('since', '0'))
        except ValueError:
            since_ts = 0.0
        unchanged = not (ts > since_ts)

    if unchanged:
        return jsonify({'ok': True, 'changed': False, 'timestamp': ts, 'version': ver})

    state = get_state_snapshot()
    state['canvas'] = dict(_canvas_meta)
    # 兼容旧数据：仍残留 dataURL 的图层替换为图片端点 URL，不内嵌 base64
    for l in state.get('layers', []):
        if isinstance(l, dict) and l.get('dataURL'):
            if not l.get('url'):
                l['url'] = '/api/shared-canvas/layer/' + str(l.get('id'))
            l['dataURL'] = ''
    return jsonify({
        'ok': True,
        'state': state,
        'timestamp': ts,
        'version': state.get('_ver', ver),
        'changed': True,
    })


@app.route('/api/shared-canvas', methods=['POST'])
def api_push_shared_canvas():
    """推送画布更新（HTTP 降级通道，WS 断开时使用）。

    支持 _mode: 'full'（全量替换）或 'incremental'（按 id 合并）；
    增量模式可带 _removed: {key: [id, ...]} 实现删除同步（P1-1）。
    """
    data = request.get_json(silent=True) or {}
    updates = data.get('updates', data)
    mode = updates.get('_mode', data.get('_mode', 'incremental'))
    removed = updates.get('_removed') or data.get('_removed') or {}

    ver = get_version()
    for key in ['strokes', 'layers', 'tokens', 'texts', 'fog']:
        items = updates.get(key)
        removed_ids = removed.get(key) or []
        if items is None and not removed_ids:
            continue

        if mode == 'full' or key == 'fog':
            ver = update_shared_canvas(key, items or [])
        elif key == 'strokes':
            if items:
                ver = append_shared_strokes(items)
            if removed_ids:
                ver = remove_shared_strokes(removed_ids)
        else:
            ver = apply_incremental(key, items or [], removed_ids)

    # 有变更则通知 WS 在线客户端拉取（HTTP 推送方通常自身无 WS 连接）
    return jsonify({'ok': True, 'timestamp': get_shared_canvas_ts(), 'version': ver})


@app.route('/api/shared-canvas/layer-image', methods=['POST'])
def api_upload_layer_image():
    """上传图层图片（P0-1：图片与画布状态分离，只在状态中保存 URL）。

    请求体: {id: <图层id>, dataURL: 'data:image/png;base64,...'}
    响应:  {ok: true, url: '/maps/layers/xxx.png'}
    """
    data = request.get_json(silent=True) or {}
    layer_id = data.get('id')
    data_url = data.get('dataURL', '')
    if layer_id is None or not data_url.startswith('data:image'):
        return jsonify({'ok': False, 'error': '参数无效：需要 id 和 dataURL'}), 400
    url = save_layer_image(layer_id, data_url)
    if not url:
        return jsonify({'ok': False, 'error': '图片保存失败（格式不支持或磁盘错误）'}), 500
    return jsonify({'ok': True, 'url': url})


@app.route('/api/shared-canvas/layer/<int:layer_id>', methods=['GET'])
def api_get_layer_image(layer_id):
    """获取共享画布图层图片（兼容端点；dataURL 为空时回退到 maps/layers/ 文件）。"""
    canvas = get_shared_canvas()
    for layer in canvas.get('layers', []):
        if layer.get('id') == layer_id:
            data_url = layer.get('dataURL', '')
            if data_url:
                import base64
                try:
                    header, b64 = data_url.split(',', 1)
                    img_data = base64.b64decode(b64)
                except Exception:
                    break  # dataURL 损坏，尝试文件回退
                else:
                    mime = 'image/png'
                    if 'image/jpeg' in header or 'image/jpg' in header:
                        mime = 'image/jpeg'
                    elif 'image/gif' in header:
                        mime = 'image/gif'
                    elif 'image/webp' in header:
                        mime = 'image/webp'
                    return app.response_class(img_data, mimetype=mime)
            # dataURL 为空或损坏，回退到 maps/layers/ 文件
            for ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
                fpath = MAPS_DIR / 'layers' / f'{layer_id}{ext}'
                if fpath.exists() and fpath.is_file():
                    mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                                '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp'}
                    return send_file(str(fpath), mimetype=mime_map.get(ext, 'image/png'))
            break
    abort(404)


# ━━━ 战斗状态同步 ━━━
_combat_state: dict | None = None  # 共享的战斗状态
_combat_state_ts: float = 0.0


def get_active():
    global active_char_id
    if active_char_id is None:
        return None
    return get_character(active_char_id)


# ━━━ 页面路由 ━━━

@app.route('/favicon.ico')
def favicon():
    """将 .ico 请求重定向到 SVG 图标"""
    return redirect('/static/favicon.svg')

@app.route('/')
def index():
    """官网门户首页"""
    return render_template('portal/index.html')


@app.route('/test4')
def test4_page():
    """北境雪原"""
    return render_template('test4.html')


@app.route('/north-expedition')
def north_expedition_page():
    """北境探索 — 独立生存冒险页面"""
    return render_template('north-expedition.html')


@app.route('/user')
def user_page():
    """用户中心"""
    return render_template('portal/user.html')


@app.route('/character')
def character_page():
    """角色卡编辑"""
    return render_template('character.html')


@app.route('/combat')
def combat_page():
    """战斗追踪器"""
    return render_template('combat.html')


@app.route('/reference')
def reference_page():
    """参考文档"""
    return render_template('reference.html')


@app.route('/map')
def map_page():
    """战术地图"""
    return render_template('map.html')


@app.route('/spells')
def spells_page():
    """法术管理"""
    return render_template('spells.html')


@app.route('/events')
def events_page():
    """随机事件表"""
    return render_template('events.html', event_groups=_EVENT_GROUPS)


@app.route('/chat')
def chat_page():
    """聊天室（禁用缓存，避免页面/脚本更新后浏览器仍渲染旧版）"""
    resp = _make_response(render_template('chat.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/dice3d')
def dice3d_page():
    """3D掷骰主页面"""
    return render_template('dice3d-e.html')


# ━━━ API 路由 ━━━

@app.route('/api/roll', methods=['POST'])
def api_roll():
    """掷骰 API"""
    data = request.get_json()
    expr = data.get('expression', '').strip()
    if not expr:
        return jsonify({'error': '缺少表达式'}), 400

    try:
        result = roll(expr)
        crit = 'success' if result.is_crit_success else ('failure' if result.is_crit_failure else None)
        # 获取当前活跃角色名
        char_name = ''
        active_char = get_active()
        if active_char:
            char_name = active_char.get('name', '')
        return jsonify({
            'expression': result.expression,
            'rolls': result.rolls,
            'kept_rolls': result.kept_rolls,
            'modifier': result.modifier,
            'total': result.total,
            'advantage': result.advantage,
            'is_crit_success': result.is_crit_success,
            'is_crit_failure': result.is_crit_failure,
            'char_name': char_name,
            'formatted': format_dice_result(
                result.expression, result.rolls, result.total,
                modifier=result.modifier, advantage=result.advantage,
                is_crit=crit,
            ).replace('\033[1m', '<b>').replace('\033[0m', '</b>')
             .replace('\033[31m', '<span style="color:red">').replace('\033[0m', '</span>')
             .replace('\033[32m', '<span style="color:green">').replace('\033[0m', '</span>')
             .replace('\033[36m', '<span style="color:cyan">').replace('\033[0m', '</span>'),
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/check', methods=['POST'])
def api_check():
    """检定 API"""
    data = request.get_json()
    target = data.get('target', '').strip()
    advantage = data.get('advantage')  # True/False/None
    char_id = data.get('char_id')  # 可选：指定角色（优先级高于全局 active_char_id）

    char = None
    if char_id:
        try:
            char = get_character(int(char_id))
        except (ValueError, TypeError):
            char = get_character(char_id)
    ability_mod = 0
    prof_bonus = 0
    label = target

    skill = normalize_skill(target)
    ability = normalize_ability(target)

    if skill:
        ability_for_skill = get_ability_for_skill(skill)
        label = f"{skill}({ability_for_skill})"
        if char:
            ability_key_map = {
                '力量': 'str', '敏捷': 'dex', '体质': 'con',
                '智力': 'int', '感知': 'wis', '魅力': 'cha',
            }
            abbr = ability_key_map.get(ability_for_skill, 'str')
            ability_score = char['abilities'].get(abbr, 10)
            ability_mod = ability_modifier(ability_score)
            skill_profs = char.get('skill_proficiencies', {})
            if skill in skill_profs and skill_profs[skill].get('is_proficient'):
                prof_bonus = char['proficiency_bonus']
                if skill_profs[skill].get('is_expertise'):
                    prof_bonus *= 2
    elif ability:
        if char:
            ability_key_map = {
                '力量': 'str', '敏捷': 'dex', '体质': 'con',
                '智力': 'int', '感知': 'wis', '魅力': 'cha',
            }
            abbr = ability_key_map.get(ability, 'str')
            ability_score = char['abilities'].get(abbr, 10)
            ability_mod = ability_modifier(ability_score)
            save_profs = char.get('save_proficiencies', {})
            if ability in save_profs and save_profs[ability].get('is_proficient'):
                prof_bonus = char['proficiency_bonus']

    # 执行检定
    from core.dice_engine import roll_ability_check
    result = roll_ability_check(ability_mod, prof_bonus, advantage)
    total_mod = ability_mod + prof_bonus

    # 获取所有 d20 投掷结果（优势/劣势时有2个）
    d20_rolls = result.rolls if result.rolls else [0]
    kept_roll = result.total - total_mod  # 反推实际使用的 d20 值
    # 对于优势：d20_rolls 包含两次投掷，最终取最高/最低

    char_name = char.get('name', '') if char else ''
    return jsonify({
        'label': label,
        'd20_roll': result.rolls[-1] if result.rolls else 0,
        'd20_rolls': d20_rolls,
        'kept_roll': max(d20_rolls) if advantage is True else (min(d20_rolls) if advantage is False else d20_rolls[0]),
        'ability_mod': ability_mod,
        'prof_bonus': prof_bonus,
        'total_mod': total_mod,
        'total': result.total,
        'advantage': advantage,
        'is_crit_success': result.is_crit_success,
        'is_crit_failure': result.is_crit_failure,
        'char_name': char_name,
    })


@app.route('/api/characters', methods=['GET'])
def api_list_characters():
    """列出角色和分组。DM 可看到全部；PL 只能看到自己创建/导入的角色和分组。"""
    name = request.args.get('name', '').strip()
    role = request.args.get('role', 'PL')
    client_ip = request.remote_addr or ''

    # DM 可看全部（明确声明为 DM 才放行，不再按 IP 自动提权）
    if role == 'DM':
        chars = list_characters()
        groups = list_character_groups()
    elif name:
        # PL 有用户名 → 看自己创建的 + DM 公开的角色
        chars = list_characters(created_by=name, include_public=True)
        groups = list_character_groups(created_by=name)
    else:
        # PL 未设置用户名 → 看不到任何角色（安全性）
        chars = []
        groups = []

    return jsonify({'characters': chars, 'groups': groups})


@app.route('/api/characters/hp', methods=['GET'])
def api_characters_hp():
    """批量获取角色 HP（轻量，供地图死亡标识轮询）"""
    ids_arg = request.args.get('ids', '')
    id_list = []
    for part in ids_arg.split(','):
        try:
            id_list.append(int(part))
        except (TypeError, ValueError):
            continue
    if not id_list:
        return jsonify({'ok': True, 'hp': []})
    from core.models import Character
    chars = Character.query.filter(Character.id.in_(id_list)).all()
    return jsonify({'ok': True, 'hp': [
        {'id': c.id, 'hp_current': c.hp_current, 'hp_max': c.hp_max}
        for c in chars
    ]})


@app.route('/api/characters/reorder', methods=['POST'])
def api_reorder_characters():
    """角色列表拖动排序。接收按新顺序排列的角色 ID 列表。"""
    data = request.get_json(silent=True) or {}
    ordered_ids = data.get('ids', [])
    if not ordered_ids or not isinstance(ordered_ids, list):
        return jsonify({'error': '请提供角色ID列表'}), 400

    from core.models import Character as _CharModel
    for i, char_id in enumerate(ordered_ids):
        _CharModel.query.filter_by(id=char_id).update({'sort_order': i})
    db.session.commit()
    return jsonify({'ok': True})


# ━━━ 角色分组管理 ━━━

@app.route('/api/character-groups', methods=['POST'])
def api_create_group():
    """创建角色分组"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '新分组').strip()
    if not name:
        return jsonify({'error': '分组名不能为空'}), 400
    created_by = data.get('created_by', '')
    gid = create_character_group(name, created_by)
    return jsonify({'ok': True, 'id': gid, 'name': name})


@app.route('/api/character-groups/<int:group_id>', methods=['PUT'])
def api_update_group(group_id):
    """更新分组（重命名）"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '分组名不能为空'}), 400
    update_character_group(group_id, name=name)
    return jsonify({'ok': True})


@app.route('/api/character-groups/<int:group_id>', methods=['DELETE'])
def api_delete_group(group_id):
    """删除分组，角色移回未分组"""
    delete_character_group(group_id)
    return jsonify({'ok': True})


@app.route('/api/character-groups/reorder', methods=['POST'])
def api_reorder_groups():
    """分组拖动排序"""
    data = request.get_json(silent=True) or {}
    ordered_ids = data.get('ids', [])
    if not ordered_ids or not isinstance(ordered_ids, list):
        return jsonify({'error': '请提供分组ID列表'}), 400
    for i, gid in enumerate(ordered_ids):
        update_character_group(gid, sort_order=i)
    return jsonify({'ok': True})


@app.route('/api/character/<int:char_id>/group', methods=['PUT'])
def api_set_char_group(char_id):
    """将角色移入/移出分组"""
    data = request.get_json(silent=True) or {}
    group_id = data.get('group_id')  # None 表示移回未分组
    set_character_group(char_id, group_id)
    return jsonify({'ok': True})


@app.route('/api/character/<name_or_id>', methods=['GET'])
def api_get_character(name_or_id):
    """获取角色详情"""
    try:
        char_id = int(name_or_id)
    except ValueError:
        char_id = name_or_id

    char = get_character(char_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    # 格式化角色卡
    char['formatted'] = format_character_sheet(char)
    return jsonify(char)


@app.route('/api/character', methods=['POST'])
def api_create_character():
    """创建角色"""
    data = request.get_json()
    name = data.get('name', '')
    level = data.get('level', 1)
    cls = data.get('class', '')
    race = data.get('race', '')
    background = data.get('background', '')
    created_by = data.get('created_by', '')
    group_id = data.get('group_id')

    if not name:
        return jsonify({'error': '角色名不能为空'}), 400

    char_id = create_character(name, level, cls, race, background, created_by=created_by, group_id=group_id)
    char = get_character(char_id)

    global active_char_id
    active_char_id = char_id

    return jsonify({'id': char_id, 'name': name, 'formatted': format_character_sheet(char)})


@app.route('/api/character/<name_or_id>', methods=['PUT'])
def api_update_character(name_or_id):
    """更新角色"""
    data = request.get_json()
    try:
        char_id = int(name_or_id)
    except ValueError:
        char = get_character(name_or_id)
        if not char:
            return jsonify({'error': '角色不存在'}), 404
        char_id = char['id']

    # 处理属性更新
    for key, value in data.items():
        if key.upper() in ('STR', 'DEX', 'CON', 'INT', 'WIS', 'CHA', '力量', '敏捷', '体质', '智力', '感知', '魅力'):
            set_ability(char_id, key, int(value))
        elif key in ('name', 'level', 'class', 'race', 'background', 'hp_max', 'hp_current', 'ac', 'speed'):
            update_character(char_id, **{key: value})

    char = get_character(char_id)
    return jsonify({
        'success': True,
        'formatted': format_character_sheet(char),
    })


def _resolve_char(name_or_id: str) -> dict | None:
    """按名字或ID查找角色（URL参数始终为字符串）"""
    try:
        char_id = int(name_or_id)
    except ValueError:
        char_id = name_or_id
    return get_character(char_id)


@app.route('/api/character/<name>/hp', methods=['POST'])
def api_adjust_hp(name):
    """调整 HP。setAbsolute=true 时直接设为指定值"""
    char = _resolve_char(name)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json()
    amount = data.get('amount', 0)
    if data.get('setAbsolute'):
        from core.character import set_hp
        result = set_hp(char['id'], hp_current=amount)
    else:
        result = adjust_hp(char['id'], amount)
    return jsonify(result)


@app.route('/api/character/<name>/spell_slot', methods=['POST'])
def api_use_spell_slot(name):
    """消耗法术位"""
    char = _resolve_char(name)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json()
    level = str(data.get('level', 1))
    result = use_spell_slot(char['id'], level)
    return jsonify(result)


@app.route('/api/character/<name>/longrest', methods=['POST'])
def api_long_rest(name):
    """长休"""
    char = _resolve_char(name)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    result = long_rest(char['id'])
    return jsonify(result)


# ━━━ 属性/技能/豁免 API ━━━

@app.route('/api/character/<name_or_id>/ability', methods=['PUT'])
def api_set_ability(name_or_id):
    """设置属性值"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    ability = data.get('ability', '').strip().lower()
    score = int(data.get('score', 10))

    valid = {'str', 'dex', 'con', 'int', 'wis', 'cha',
             '力量', '敏捷', '体质', '智力', '感知', '魅力'}
    if ability not in valid:
        return jsonify({'error': f'无效属性: {ability}'}), 400
    if score < 1 or score > 30:
        return jsonify({'error': '属性值范围 1-30'}), 400

    from core.character import set_ability as _set_ability
    ok = _set_ability(char['id'], ability, score)
    if not ok:
        return jsonify({'error': '设置失败'}), 400

    # 重新获取更新后的角色
    updated = get_character(char['id'])
    return jsonify({
        'success': True,
        'ability': ability,
        'score': score,
        'mod': updated['ability_mods'].get(ability, 0) if updated.get('ability_mods') else 0,
    })


@app.route('/api/character/<name_or_id>/skill', methods=['POST'])
def api_toggle_skill(name_or_id):
    """设置技能熟练和加值"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    skill_name = data.get('skill', '').strip()
    is_proficient = bool(data.get('proficient', True))
    is_expertise = bool(data.get('expertise', False))
    bonus = int(data.get('bonus', 0) or 0)

    skill = normalize_skill(skill_name)
    if not skill:
        return jsonify({'error': f'无效技能: {skill_name}'}), 400

    from core.character import set_skill_proficiency as _set_skill
    ok = _set_skill(char['id'], skill, is_proficient or is_expertise, is_expertise, bonus)
    if not ok:
        return jsonify({'error': '设置失败'}), 400

    return jsonify({
        'success': True,
        'skill': skill,
        'is_proficient': is_proficient or is_expertise,
        'is_expertise': is_expertise,
        'bonus': bonus,
    })


@app.route('/api/character/<name_or_id>/save-prof', methods=['POST'])
def api_toggle_save_prof(name_or_id):
    """设置豁免熟练和加值"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    ability_name = data.get('ability', '').strip()
    is_proficient = bool(data.get('proficient', True))
    save_bonus = int(data.get('save_bonus', 0) or 0)

    ability = normalize_ability(ability_name)
    if not ability:
        return jsonify({'error': f'无效属性: {ability_name}'}), 400

    from core.character import set_save_proficiency as _set_save
    ok = _set_save(char['id'], ability, is_proficient, save_bonus)
    if not ok:
        return jsonify({'error': '设置失败'}), 400

    return jsonify({
        'success': True, 'ability': ability,
        'is_proficient': is_proficient, 'save_bonus': save_bonus,
    })


# ━━━ 角色特性 API（职业能力/专长/种族特性/特殊能力/其他）━━━

@app.route('/api/character/<name_or_id>/features/<category>', methods=['GET'])
def api_get_features(name_or_id, category):
    """获取角色某类特性"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404
    features = char.get('features', {}).get(category, [])
    return jsonify({'features': features})


@app.route('/api/character/<name_or_id>/features/<category>', methods=['POST'])
def api_add_feature(name_or_id, category):
    """添加角色特性"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '特性名不能为空'}), 400

    from core.character import add_feature as _add_feat
    fid = _add_feat(char['id'], category, name, data.get('description', ''))
    return jsonify({'success': True, 'id': fid, 'name': name})


@app.route('/api/character/<name_or_id>/features/<int:feature_id>', methods=['PUT'])
def api_update_feature(name_or_id, feature_id):
    """更新特性"""
    data = request.get_json(silent=True) or {}
    from core.character import update_feature as _upd_feat
    ok = _upd_feat(feature_id, **{k: v for k, v in data.items() if k in ('name', 'description', 'category')})
    if not ok:
        return jsonify({'error': '更新失败'}), 400
    return jsonify({'success': True})


@app.route('/api/character/<name_or_id>/features/<int:feature_id>', methods=['DELETE'])
def api_delete_feature(name_or_id, feature_id):
    """删除特性"""
    from core.character import delete_feature as _del_feat
    ok = _del_feat(feature_id)
    if not ok:
        return jsonify({'error': '删除失败'}), 404
    return jsonify({'success': True})


# ━━━ 背景信息 API ━━━

@app.route('/api/character/<name_or_id>/background', methods=['PUT'])
def api_update_background(name_or_id):
    """批量更新角色背景信息"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    allowed = ['personality_traits', 'personality_traits_ext', 'ideals', 'bonds', 'flaws',
               'background_feature', 'appearance', 'origin', 'languages', 'tool_proficiencies', 'backstory']
    from core.character import set_background
    for key in allowed:
        if key in data:
            set_background(char['id'], **{key: data[key]})
    return jsonify({'success': True})


# ━━━ 已准备法术 API ━━━

@app.route('/api/character/<name_or_id>/prepared-spell', methods=['POST'])
def api_add_prepared_spell(name_or_id):
    """添加已准备法术（从法术书中选择或直接添加）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    spell_name = data.get('name', '').strip()
    spell_level = int(data.get('level', 0))

    if not spell_name:
        return jsonify({'error': '法术名不能为空'}), 400

    from core.character import add_prepared_spell as _add_spell
    sid = _add_spell(char['id'], spell_name, spell_level)
    return jsonify({'success': True, 'spell_id': sid, 'name': spell_name, 'level': spell_level})


@app.route('/api/character/<name_or_id>/prepared-spell/<int:spell_id>', methods=['DELETE'])
def api_remove_prepared_spell(name_or_id, spell_id):
    """删除已准备法术"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    from core.models import PreparedSpell as _PS
    _PS.query.filter(_PS.id == spell_id, _PS.character_id == char['id']).delete()
    db.session.commit()
    return jsonify({'success': True})


# ━━━ 法术书（已学习法术）API ━━━

@app.route('/api/character/<name_or_id>/learned-spell', methods=['POST'])
def api_add_learned_spell(name_or_id):
    """添加法术到法术书"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    spell_name = data.get('name', '').strip()
    if not spell_name:
        return jsonify({'error': '法术名不能为空'}), 400

    from core.character import add_learned_spell as _add_learned
    sid = _add_learned(
        char['id'], spell_name,
        spell_level=int(data.get('level', 0)),
        school=data.get('school', ''),
        casting_time=data.get('casting_time', ''),
        range=data.get('range', ''),
        duration=data.get('duration', ''),
        components=data.get('components', ''),
        ritual=data.get('ritual', '否'),
        concentration=data.get('concentration', '否'),
        description=data.get('description', ''),
        source=data.get('source', '自定义'),
    )
    return jsonify({'success': True, 'spell_id': sid, 'name': spell_name})


@app.route('/api/character/<name_or_id>/learned-spell/<int:spell_id>', methods=['DELETE'])
def api_remove_learned_spell(name_or_id, spell_id):
    """从法术书删除法术"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    from core.character import remove_learned_spell as _remove_learned
    ok = _remove_learned(spell_id)
    if not ok:
        return jsonify({'error': '法术不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/spell-detail/<path:name>', methods=['GET'])
def api_spell_detail(name):
    """获取法术详细描述"""
    spell = chm_get_spell_detail(name)
    if not spell:
        # 回退到 SRD
        spell = search_spell(name)
        if not spell:
            return jsonify({'error': '未找到法术'}), 404
        return jsonify({
            'name': spell.get('name', name),
            'name_en': '',
            'level': spell.get('level', '?'),
            'school': spell.get('school', '?'),
            'casting_time': spell.get('casting_time', '?'),
            'range': spell.get('range', '?'),
            'duration': spell.get('duration', '?'),
            'components': spell.get('components', '?'),
            'ritual': spell.get('ritual', '否'),
            'concentration': spell.get('concentration', '否'),
            'classes': spell.get('classes', '?'),
            'source': spell.get('source', '?'),
            'detail_text': spell.get('description', spell.get('detail', '')),
        })

    return jsonify({
        'name': spell.get('name_cn', spell.get('name', '?')),
        'name_en': spell.get('name_en', ''),
        'level': spell.get('level', '?'),
        'school': spell.get('school', '?'),
        'casting_time': spell.get('casting_time', '?'),
        'range': spell.get('range', ''),
        'duration': spell.get('duration', ''),
        'components': ' '.join(filter(None, [
            'V' if spell.get('verbal') in ('V', '✓') else '',
            'S' if spell.get('somatic') in ('S', '✓') else '',
            'M' if spell.get('material') in ('M', '✓') else '',
        ])) or '—',
        'ritual': '是' if spell.get('ritual') in ('✓', '是', 'R') else '否',
        'concentration': '是' if spell.get('concentration') in ('✓', '是', 'C') else '否',
        'classes': spell.get('classes', '?'),
        'source': spell.get('source', '?'),
        'detail_text': spell.get('detail_text', ''),
    })

@app.route('/api/character/<name_or_id>/weapon', methods=['POST'])
def api_add_weapon(name_or_id):
    """添加武器"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '武器名不能为空'}), 400

    from core.character import add_weapon as _add_weapon
    try:
        wid = _add_weapon(
            char['id'], name,
            attack_bonus=int(data.get('attack_bonus', 0)),
            damage_dice=data.get('damage', ''),
            damage_type=data.get('damage_type', ''),
            description=data.get('description', ''),
            effect=data.get('effect', ''),
        )
    except Exception as e:
        return jsonify({'error': f'添加失败: {str(e)}'}), 500
    return jsonify({'success': True, 'weapon_id': wid, 'name': name})


@app.route('/api/character/<name_or_id>/weapon/<int:weapon_id>', methods=['PUT'])
def api_update_weapon(name_or_id, weapon_id):
    """更新武器字段（名称/命中/伤害/类型/描述/效果）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    field = data.get('field', '')
    value = data.get('value', '')

    allowed = ['name', 'attack_bonus', 'damage', 'damage_type', 'description', 'effect']
    if field not in allowed:
        return jsonify({'error': f'不允许修改字段: {field}'}), 400

    from core.models import Weapon as _WP
    # 字段名与模型列名映射（damage → damage_dice 等）
    col_map = {'name': 'name', 'attack_bonus': 'attack_bonus', 'damage': 'damage_dice',
               'damage_type': 'damage_type', 'description': 'description', 'effect': 'effect'}
    col = col_map.get(field)
    if col is None:
        return jsonify({'error': f'不允许修改字段: {field}'}), 400
    n = _WP.query.filter(_WP.id == weapon_id, _WP.character_id == char['id']).update({col: value})
    db.session.commit()
    if n == 0:
        return jsonify({'error': '武器不存在或不属于该角色'}), 404
    return jsonify({'success': True})


@app.route('/api/character/<name_or_id>/weapon/<int:weapon_id>', methods=['DELETE'])
def api_remove_weapon(name_or_id, weapon_id):
    """删除武器"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    from core.character import remove_weapon as _remove_weapon
    ok = _remove_weapon(weapon_id)
    if not ok:
        return jsonify({'error': '武器不存在或不属于该角色'}), 404
    return jsonify({'success': True})


# ━━━ 物品 API ━━━

@app.route('/api/character/<name_or_id>/item', methods=['POST'])
def api_add_item(name_or_id):
    """添加物品"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '物品名不能为空'}), 400

    from core.character import add_item as _add_item
    try:
        iid = _add_item(
            char['id'], name,
            quantity=int(data.get('quantity', 1)),
            weight=float(data.get('weight', 0)),
            location=data.get('location', '背包'),
            description=data.get('description', ''),
            effect=data.get('effect', ''),
        )
    except Exception as e:
        return jsonify({'error': f'添加失败: {str(e)}'}), 500
    return jsonify({'success': True, 'item_id': iid, 'name': name})


@app.route('/api/character/<name_or_id>/item/<int:item_id>', methods=['PUT'])
def api_update_item(name_or_id, item_id):
    """更新物品（数量/字段均可）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    field = data.get('field', '')
    value = data.get('value')

    from core.character import update_item_quantity as _update_qty, remove_item as _remove_item

    # 字段编辑模式
    if field:
        allowed = ['item_name', 'quantity', 'location', 'weight', 'description', 'effect']
        if field not in allowed:
            return jsonify({'error': f'不允许修改字段: {field}'}), 400
        from core.models import InventoryItem as _II
        n = _II.query.filter(_II.id == item_id, _II.character_id == char['id']).update({field: value})
        db.session.commit()
        if n == 0:
            return jsonify({'error': '物品不存在'}), 404
        return jsonify({'success': True})

    # 数量模式（兼容旧接口）
    qty = int(data.get('quantity', 1))
    if qty <= 0:
        ok = _remove_item(item_id)
        if not ok:
            return jsonify({'error': '物品不存在'}), 404
        return jsonify({'success': True, 'deleted': True})
    else:
        ok = _update_qty(item_id, qty)
        if not ok:
            return jsonify({'error': '物品不存在'}), 404
        return jsonify({'success': True, 'quantity': qty})


@app.route('/api/character/<name_or_id>/item/<int:item_id>', methods=['DELETE'])
def api_remove_item(name_or_id, item_id):
    """删除物品"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    from core.character import remove_item as _remove_item
    ok = _remove_item(item_id)
    if not ok:
        return jsonify({'error': '物品不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/character/<name_or_id>/inventory/stack', methods=['POST'])
def api_stack_inventory(name_or_id):
    """合并同名物品"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    from core.character import stack_inventory as _stack
    merged = _stack(char['id'])
    return jsonify({'success': True, 'merged': merged})


# ━━━ 钱币 API ━━━

@app.route('/api/character/<name_or_id>/coin', methods=['POST'])
def api_adjust_coin(name_or_id):
    """调整钱币（正数增加，负数减少）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json() or {}
    coin_type = data.get('coin_type', '').strip().lower()
    amount = int(data.get('amount', 0))

    if not coin_type or amount == 0:
        return jsonify({'error': '请提供 coin_type 和 amount'}), 400

    valid = {'cp', 'sp', 'ep', 'gp', 'pp', '铜币', '银币', '金银币', '金币', '白金币'}
    if coin_type not in valid:
        return jsonify({'error': f'无效币种: {coin_type}。可用: cp/sp/ep/gp/pp'}), 400

    from core.character import adjust_coin as _adjust_coin
    result = _adjust_coin(char['id'], coin_type, amount)
    if 'error' in result:
        return jsonify({'error': result['error']}), 400
    return jsonify({'success': True, **result})


# ━━━ 头像 API ━━━

@app.route('/api/character/<name_or_id>', methods=['DELETE'])
def api_delete_character(name_or_id):
    """删除角色"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    char_name = char['name']
    char_id = char['id']
    delete_character(char_id)

    # 审计日志
    ip = request.remote_addr or 'unknown'
    audit_log('character.delete', username=char_name, ip=ip,
              detail=f'删除角色 id={char_id} name={char_name}')

    global active_char_id
    if active_char_id == char_id:
        active_char_id = None

    return jsonify({'success': True, 'name': char_name})


@app.route('/api/character/<name_or_id>/copy', methods=['POST'])
def api_copy_character(name_or_id):
    """复制角色"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json(silent=True) or {}
    new_name = data.get('name', '').strip()

    from core.character import copy_character as _copy
    new_id = _copy(char['id'], new_name)
    new_char = get_character(new_id)
    return jsonify({'id': new_id, 'name': new_char['name']})


@app.route('/api/character/<name_or_id>/field', methods=['PUT'])
def api_update_character_field(name_or_id):
    """更新角色单个字段（AC/速度/熟练/身高/体重/被动察觉等）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json(silent=True) or {}
    field = data.get('field', '').strip()
    value = data.get('value')

    if not field:
        return jsonify({'error': '缺少 field 参数'}), 400

    # 允许更新的字段
    int_fields = {
        'ac', 'speed', 'proficiency_bonus', 'passive_perception', 'initiative_bonus',
        'level', 'spell_save_dc', 'spell_attack_bonus', 'hp_max', 'hp_current', 'temp_hp',
    }
    str_fields = {
        'height', 'weight_field', 'name', 'class', 'race', 'subrace',
        'alignment', 'faith', 'gender', 'resistances', 'key_abilities',
    }
    if field not in int_fields and field not in str_fields:
        return jsonify({'error': f'不允许修改字段: {field}'}), 400

    # 数值字段
    if field in int_fields:
        try:
            value = int(value)
        except (TypeError, ValueError):
            return jsonify({'error': f'{field} 必须是整数'}), 400
    else:
        value = str(value)

    # HP限制：当前HP不能超过上限
    if field == 'hp_current':
        hp_max = char.get('hp_max', 10)
        value = max(0, min(value, hp_max))
    elif field == 'hp_max':
        value = max(1, value)

    update_character(char['id'], **{field: value})
    return jsonify({'success': True, 'field': field, 'value': value})


@app.route('/api/character/<name_or_id>/visibility', methods=['POST'])
def api_set_character_visibility(name_or_id):
    """DM 设置角色公开/私有。公开角色对全体玩家可见（PL 只读）。"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    role = (request.args.get('role') or '').strip()
    if role != 'DM':
        return jsonify({'error': '仅 DM 可以设置角色可见性'}), 403

    data = request.get_json(silent=True) or {}
    is_public = bool(data.get('is_public', False))
    update_character(char['id'], is_public=is_public)
    return jsonify({'success': True, 'is_public': is_public})


@app.route('/api/character/<name_or_id>/spell-slots', methods=['PUT'])
def api_update_spell_slots(name_or_id):
    """直接设置法术位（x/y格式，x为当前已用，y为最大）"""
    char = _resolve_char(name_or_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    data = request.get_json(silent=True) or {}
    slots = data.get('slots', {})  # {'1': {'max': 4, 'used': 2}, '2': {...}}

    from core.models import SpellSlot as _SS
    # 确保所有1-9环都存在（存在则更新，不存在则插入）
    for level in range(1, 10):
        level_str = str(level)
        slot_data = slots.get(level_str, {})
        max_slots = int(slot_data.get('max', 0))
        used_slots = int(slot_data.get('used', 0))
        used_slots = max(0, min(used_slots, max_slots))  # x不能超过y

        existing = _SS.query.filter_by(character_id=char['id'], slot_level=level_str).first()
        if existing:
            existing.max_slots = max_slots
            existing.used_slots = used_slots
        else:
            db.session.add(_SS(character_id=char['id'], slot_level=level_str,
                               max_slots=max_slots, used_slots=used_slots))
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/character/<name_or_id>/portrait', methods=['GET'])
def api_get_portrait(name_or_id):
    """获取角色头像图片"""
    try:
        char_id = int(name_or_id)
    except ValueError:
        char_id = name_or_id

    char = get_character(char_id)
    if not char:
        abort(404)

    path = char.get('portrait_path', '')
    if not path:
        abort(404)

    # 尝试解析路径
    resolved, _ = resolve_portrait_path(path)
    if not resolved:
        # 如果原始路径存在也接受
        if _os.path.exists(path) and _os.path.isfile(path):
            resolved = path
        else:
            abort(404)

    # 安全检查：限制在项目根目录内（尘封之卷-九子的注视/）
    project_root = _Path(__file__).parent.parent.parent.resolve()
    resolved_path = _Path(resolved).resolve()
    try:
        resolved_path.relative_to(project_root)
    except ValueError:
        # 不在项目目录内，拒绝访问
        abort(403)

    # 确定 mimetype
    ext = resolved_path.suffix.lower()
    mimetypes = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp', '.bmp': 'image/bmp',
    }
    mimetype = mimetypes.get(ext, 'application/octet-stream')

    resp = send_file(str(resolved_path), mimetype=mimetype)
    # 头像不缓存：上传新头像后立即显示新图，避免浏览器命中旧缓存
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/api/character/<name_or_id>/portrait', methods=['PUT'])
def api_set_portrait(name_or_id):
    """设置角色头像路径"""
    data = request.get_json()
    path = data.get('path', '').strip()
    if not path:
        return jsonify({'error': '缺少路径参数'}), 400

    try:
        char_id = int(name_or_id)
    except ValueError:
        char = get_character(name_or_id)
        if not char:
            return jsonify({'error': '角色不存在'}), 404
        char_id = char['id']

    # 尝试解析路径
    resolved, tried = resolve_portrait_path(path)
    if resolved:
        update_character(char_id, portrait_path=resolved)
        return jsonify({'success': True, 'path': resolved})
    else:
        # 保存原始路径，但给出警告（含调试信息）
        tried_str = '\n已尝试搜索:\n' + '\n'.join(tried[-5:]) if tried else ''
        update_character(char_id, portrait_path=path)
        return jsonify({
            'warning': f'文件未找到，已保存原始路径{tried_str}',
            'path': path,
            'tried': tried[-5:] if tried else []
        })


@app.route('/api/character/<name_or_id>/portrait', methods=['DELETE'])
def api_clear_portrait(name_or_id):
    """清除角色头像"""
    try:
        char_id = int(name_or_id)
    except ValueError:
        char = get_character(name_or_id)
        if not char:
            return jsonify({'error': '角色不存在'}), 404
        char_id = char['id']

    update_character(char_id, portrait_path='')
    return jsonify({'success': True})


@app.route('/api/character/<name_or_id>/portrait/upload', methods=['POST'])
def api_upload_portrait(name_or_id):
    """上传角色头像图片到资源库头像子文件夹，自动命名为角色名。"""
    try:
        char_id = int(name_or_id)
    except ValueError:
        char = get_character(name_or_id)
        if not char:
            return jsonify({'error': '角色不存在'}), 404
        char_id = char['id']

    char = get_character(char_id)
    if not char:
        return jsonify({'error': '角色不存在'}), 404

    if 'file' not in request.files:
        return jsonify({'error': '请选择图片文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    ext = _os.path.splitext(file.filename)[1].lower()
    if ext not in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'):
        return jsonify({'error': f'不支持的图片格式: {ext}'}), 400

    # 保存到 resources/头像/<角色名>_<角色id><ext>
    # 文件名带角色 id：不同角色同名也不会互相覆盖头像（PL 与 DM 同名角色同理）
    avatar_dir = RESOURCES_DIR / '头像'
    avatar_dir.mkdir(exist_ok=True)
    safe_name = ''.join(c for c in char['name'] if c not in r'<>:"/\|?*')
    filename = f'{safe_name}_{char_id}{ext}'
    save_path = avatar_dir / filename

    try:
        file.save(str(save_path))
    except Exception as e:
        return jsonify({'error': f'保存失败: {e}'}), 500

    # 更新数据库中的头像路径
    update_character(char_id, portrait_path=str(save_path))

    return jsonify({
        'success': True,
        'path': str(save_path),
        'url': f'/resources/头像/{filename}',
    })


@app.route('/api/spells/search', methods=['GET'])
def api_search_spells_list():
    """模糊搜索法术列表（按匹配质量排序）"""
    query = request.args.get('q', '').strip()
    if len(query) < 1:
        return jsonify({'error': '请输入搜索关键词'}), 400

    from core.chm_search import search_spell as _search_spell, _calc_spell_score

    scored_items = []
    seen_names = set()

    # 1. CHM法术（自带评分排序）
    for s in _search_spell(query):
        name = s.get('name_cn', s.get('name', '?'))
        if name in seen_names: continue
        score = _calc_spell_score(query, name, s.get('name_en', ''), s.get('name', ''), s.get('tags', ''))
        seen_names.add(name)
        scored_items.append((score, {
            'name': name,
            'name_en': s.get('name_en', ''),
            'level': s.get('level', ''),
            'school': s.get('school', ''),
            'casting_time': s.get('casting_time', ''),
            'components': ' '.join(filter(None, [
                'V' if s.get('verbal') in ('V', '✓') else '',
                'S' if s.get('somatic') in ('S', '✓') else '',
                'M' if s.get('material') in ('M', '✓') else '',
            ])) or '—',
            'ritual': '是' if s.get('ritual') in ('✓', '是', 'R') else '否',
            'concentration': '是' if s.get('concentration') in ('✓', '是', 'C') else '否',
            'classes': s.get('classes', ''),
            'source': s.get('source', ''),
            'detail_link': s.get('detail_link', ''),
            'detail': f"{s.get('level','?')}环 {s.get('school','?')} | {s.get('classes','?')}",
        }))

    # 2. 自定义法术（评分排序）
    for e in _load_custom('spells'):
        name = e['name']
        if name in seen_names: continue
        score = _calc_spell_score(query, name, e.get('name_en', ''))
        if score > 0:
            seen_names.add(name)
            scored_items.append((score, {
                'name': name, 'name_en': e.get('name_en', ''),
                'level': e.get('level', ''), 'school': e.get('school', ''),
                'classes': e.get('classes', ''), 'source': '自定义',
                'detail': e.get('description', '')[:100],
            }))

    # 按得分降序
    scored_items.sort(key=lambda x: x[0], reverse=True)
    items = [item for _, item in scored_items]

    return jsonify({'query': query, 'results': items[:200], 'total': len(items)})


@app.route('/api/spell/<name>', methods=['GET'])
def api_search_spell(name):
    """查询法术（自定义数据优先）"""
    # 1. 先检查自定义法术
    for e in _load_custom('spells'):
        if e['name'].lower() == name.lower():
            return jsonify({
                'name': e['name'],
                'name_en': e.get('name_en', ''),
                'level': e.get('level', ''),
                'school': e.get('school', ''),
                'casting_time': e.get('casting_time', ''),
                'range': e.get('range', ''),
                'duration': e.get('duration', ''),
                'components': e.get('components', ''),
                'ritual': e.get('ritual', ''),
                'concentration': e.get('concentration', ''),
                'classes': e.get('classes', ''),
                'detail_text': e.get('description', ''),
                'source': e.get('source', '自定义'),
            })

    # 2. CHM搜索
    spell = search_spell(name)
    if not spell:
        return jsonify({'error': '未找到法术'}), 404
    return jsonify(spell)


@app.route('/api/monsters/search', methods=['GET'])
def api_search_monsters():
    """模糊搜索怪物列表（按匹配质量排序——匹配字数越多越靠前）"""
    query = request.args.get('q', '').strip()
    if len(query) < 1:
        return jsonify({'error': '请输入搜索关键词'}), 400

    from core.chm_search import search_monster as _search_monster, _calc_match_score

    scored_items = []  # (score, item_dict)
    seen_names = set()

    def _make_item(name, name_en='', cr='', size='', mtype='', source='', detail='', detail_text='', detail_link=''):
        return {
            'name': name, 'name_en': name_en, 'cr': cr,
            'size': size, 'type': mtype, 'source': source,
            'detail': detail or f"CR:{cr or '?'} {size or ''} {mtype or ''}",
            'detail_text': detail_text,
            'detail_link': detail_link,
        }

    # 1. 自定义怪物（评分排序）
    for e in _load_custom('monsters'):
        name = e['name']
        if name in seen_names: continue
        score = _calc_match_score(query, name, e.get('name_en', ''))
        if score > 0:
            seen_names.add(name)
            scored_items.append((score, _make_item(
                name, e.get('name_en', ''), e.get('cr', ''), e.get('size', ''),
                e.get('type', ''), '自定义', '', e.get('detail_text', ''),
            )))

    # 2. CHM 怪物（已自带评分排序）
    for m in _search_monster(query):
        name = m.get('name_cn', m.get('name', '?'))
        if name in seen_names: continue
        score = _calc_match_score(query, name, m.get('name_en', ''))
        if score > 0:
            seen_names.add(name)
            scored_items.append((score, _make_item(
                name, m.get('name_en', ''), m.get('cr', ''), m.get('size', ''),
                m.get('type', ''), m.get('source', ''),
                detail_link=m.get('detail_link', ''),
            )))

    # 3. SRD 怪物（评分排序）
    try:
        from utils.data_loader import load_monsters as _load_srd
        for m in _load_srd():
            name = m.get('name', '')
            if name in seen_names: continue
            score = _calc_match_score(query, name, m.get('name_en', ''))
            if score > 0:
                seen_names.add(name)
                scored_items.append((score, _make_item(
                    name, m.get('name_en', ''), m.get('cr', ''), m.get('size', ''),
                    m.get('type', ''), 'SRD',
                )))
    except Exception:
        pass

    # 按得分降序排列
    scored_items.sort(key=lambda x: x[0], reverse=True)
    items = [item for _, item in scored_items]

    return jsonify({'query': query, 'results': items[:200], 'total': len(items)})


@app.route('/api/monster/<name>', methods=['GET'])
def api_search_monster(name):
    """查询怪物（自定义数据优先，CHM 次之，SRD 兜底）"""
    # 1. 先检查自定义怪物（精确/包含匹配优先）
    custom_monsters = _load_custom('monsters')
    for e in custom_monsters:
        if e['name'].lower() == name.lower():
            return jsonify({
                'name': e['name'],
                'name_en': e.get('name_en', ''),
                'cr': e.get('cr', '?'),
                'size': e.get('size', '?'),
                'type': e.get('type', '?'),
                'source': e.get('source', '自定义'),
                'legendary': e.get('legendary', ''),
                'detail_text': e.get('detail_text', ''),
                'source_db': 'custom',
            })
    for e in custom_monsters:
        if name.lower() in e['name'].lower():
            return jsonify({
                'name': e['name'],
                'name_en': e.get('name_en', ''),
                'cr': e.get('cr', '?'),
                'size': e.get('size', '?'),
                'type': e.get('type', '?'),
                'source': e.get('source', '自定义'),
                'legendary': e.get('legendary', ''),
                'detail_text': e.get('detail_text', ''),
                'source_db': 'custom',
            })

    # 2. 从 CHM 获取完整详情（含 detail_text）
    monster = chm_get_monster_detail(name)
    if monster:
        return jsonify({
            'name': monster.get('name_cn', monster.get('name', '?')),
            'name_en': monster.get('name_en', ''),
            'cr': monster.get('cr', '?'),
            'size': monster.get('size', '?'),
            'type': monster.get('type', '?'),
            'source': monster.get('source', '?'),
            'legendary': monster.get('legendary', ''),
            'detail_text': monster.get('detail_text', ''),
            'source_db': 'chm',
        })

    # 3. CHM 模糊搜索
    chm_results = chm_search_monster(name)
    if chm_results:
        m = chm_results[0]
        detail_text = ''
        if m.get('detail_link'):
            from core.chm_search import read_detail_page
            detail_text = read_detail_page(m['detail_link']) or ''
        return jsonify({
            'name': m.get('name_cn', m.get('name', '?')),
            'name_en': m.get('name_en', ''),
            'cr': m.get('cr', '?'),
            'size': m.get('size', '?'),
            'type': m.get('type', '?'),
            'source': m.get('source', '?'),
            'legendary': m.get('legendary', ''),
            'detail_text': detail_text,
            'source_db': 'chm',
        })

    # 4. 回退到 SRD
    monster = search_monster(name)
    if not monster:
        return jsonify({'error': '未找到怪物'}), 404
    return jsonify({**monster, 'source_db': 'srd'})


@app.route('/api/search', methods=['GET'])
def api_search():
    """综合搜索（CHM 资料库 + 项目文件 + 自定义资料库，按匹配质量排序）"""
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify({'error': '关键词至少2个字'}), 400

    from core.chm_search import (
        search_all_combined as _combined,
        search_project_files as _search_files,
        _calc_match_score,
        _calc_spell_score,
    )

    scored_items = []  # (score, item_dict)
    seen = set()

    def _add(score, item):
        name = item.get('name', '')
        if name in seen: return
        seen.add(name)
        scored_items.append((score, item))

    # 1. 自定义怪物优先（评分排序，排在最前面）
    for e in _load_custom('monsters'):
        name = e['name']
        score = _calc_match_score(query, name, e.get('name_en', ''))
        if score > 0:
            _add(score + 200, {  # 自定义数据高优先级
                'type': 'monster',
                'name': name,
                'name_en': e.get('name_en', ''),
                'detail': e.get('detail_text', '')[:100],
                'cr': e.get('cr', ''),
                'source': e.get('source', '自定义'),
            })

    # 2. 自定义法术优先（评分排序）
    for e in _load_custom('spells'):
        name = e['name']
        score = _calc_spell_score(query, name, e.get('name_en', ''))
        if score > 0:
            _add(score + 200, {
                'type': 'spell',
                'name': name,
                'name_en': e.get('name_en', ''),
                'detail': e.get('description', '')[:100],
                'level': e.get('level', ''),
                'school': e.get('school', ''),
                'source': e.get('source', '自定义'),
            })

    # 3. CHM综合搜索（自带评分，去重）
    for r in _combined(query):
        rtype = r.get('type', '')
        name = r.get('name_cn', r.get('name', r.get('title', '?')))
        base_score = 85
        _add(base_score, {
            'type': rtype,
            'name': name,
            'name_en': r.get('name_en', ''),
            'detail': r.get('detail', r.get('snippet', '')),
            'path': r.get('path', r.get('detail_link', '')),
            'cr': r.get('cr', ''),
            'level': r.get('level', ''),
            'school': r.get('school', ''),
            'source': r.get('source', ''),
        })

    # 4. 项目文件
    files = _search_files(query)
    for i, f in enumerate(files[:10]):
        name = f.get('name', '?')
        snippet = f.get('snippet', '') or ''
        cat = f.get('cat_label', '')
        parent = f.get('parent', '')
        detail_parts = []
        if snippet: detail_parts.append(snippet[:200])
        if cat: detail_parts.append(cat)
        if parent: detail_parts.append(parent)
        _add(max(0, 40 - i), {
            'type': 'file',
            'name': name,
            'detail': ' | '.join(detail_parts) if detail_parts else '',
            'path': f.get('path', ''),
            'ext': f.get('ext', ''),
            'snippet': snippet,
        })

    # 5. DND物品表（评分排序） + 自定义物品优先
    item_data = _load_items()
    for it in item_data:
        name = it['name']
        score = _calc_match_score(query, name, it.get('name_en', ''))
        if score > 0 and len(scored_items) < 200:
            detail_parts = []
            if it['price']: detail_parts.append(f"💰{it['price']}")
            if it['damage']: detail_parts.append(f"⚔️{it['damage']}")
            if it['weight']: detail_parts.append(f"⚖️{it['weight']}")
            if it['type']: detail_parts.append(f"📦{it['type']}")
            _add(score + 30, {
                'type': 'item',
                'name': name,
                'detail': ' | '.join(detail_parts),
            })

    # 6. 自定义物品
    for e in _load_custom('items'):
        name = e['name']
        score = _calc_match_score(query, name)
        if score > 0 and len(scored_items) < 200:
            _add(score + 40, {
                'type': 'item',
                'name': name,
                'detail': e.get('description', '')[:100],
                'source': '自定义',
            })

    # 按得分降序排列
    scored_items.sort(key=lambda x: x[0], reverse=True)
    items = [item for _, item in scored_items]

    return jsonify({'query': query, 'results': items[:200], 'total': len(items)})


@app.route('/api/conditions', methods=['GET'])
def api_conditions():
    """获取状态列表"""
    return jsonify(load_conditions())


@app.route('/api/skills', methods=['GET'])
def api_skills():
    """获取技能列表"""
    return jsonify(SKILL_TO_ABILITY)


@app.route('/api/abilities', methods=['GET'])
def api_abilities():
    """获取属性列表"""
    return jsonify(ABILITY_ORDER)


@app.route('/api/character/from-monster', methods=['POST'])
def api_create_from_monster():
    """从怪物数据创建角色"""
    data = request.get_json()
    name = data.get('name', '')
    if not name:
        return jsonify({'error': '缺少怪物名称'}), 400

    # 估算属性
    cr_str = data.get('cr', '0')
    try:
        cr = float(cr_str)
    except ValueError:
        cr = 0

    # CR → 属性估算（粗略映射）
    import math
    prof_bonus = max(2, math.ceil((cr + 4) / 4)) if cr > 0 else 2
    # 估算 HP（每CR约15HP，至少10）
    hp_est = max(10, int(cr * 15 + 10)) if cr > 0 else 10
    # 估算 AC
    ac_est = min(22, 10 + int(cr / 2) + 3)
    # 估算属性值（基于CR粗略估算）
    base_score = min(20, 10 + int(cr))
    abilities = {
        'str': base_score, 'dex': min(20, base_score - 1),
        'con': min(20, base_score + 1), 'int': max(8, base_score - 2),
        'wis': min(20, base_score - 1), 'cha': max(8, base_score - 2),
    }

    size = data.get('size', '中型')
    mtype = data.get('type', '')
    race_str = f'{size} {mtype}'.strip() or '未知'
    created_by = data.get('created_by', '')

    char_id = create_character(name, level=max(1, int(cr)), cls='怪物/NPC', race=race_str, created_by=created_by)
    char = get_character(char_id)

    # 设置属性
    for key, score in abilities.items():
        set_ability(char_id, key, score)

    # 设置 HP 和 AC
    update_character(char_id, hp_max=hp_est, hp_current=hp_est, ac=ac_est)

    # ━━ 将无法直接填入的怪物详细数据写入背景特性 ━━
    detail_text = data.get('detail_text', '')
    source = data.get('source', '')
    monster_size = data.get('size', '')
    monster_type = data.get('type', '')
    legendary = data.get('legendary', '')

    bg_parts = []
    if detail_text:
        bg_parts.append(f"【怪物数据】\n{detail_text}")
    else:
        info_parts = []
        if source: info_parts.append(f"来源：{source}")
        if monster_size: info_parts.append(f"体型：{monster_size}")
        if monster_type: info_parts.append(f"类型：{monster_type}")
        if legendary == '有': info_parts.append("传奇动作：有")
        if cr_str: info_parts.append(f"挑战等级：{cr_str}")
        if info_parts:
            bg_parts.append(f"【怪物数据】\n{' | '.join(info_parts)}")
            # 估算属性信息
            bg_parts.append(f"估算属性：力{abilities['str']} 敏{abilities['dex']} 体{abilities['con']} 智{abilities['int']} 感{abilities['wis']} 魅{abilities['cha']}")

    if bg_parts:
        from core.character import set_background
        set_background(char_id, background_feature='\n\n'.join(bg_parts))

    # 设置活跃角色
    global active_char_id
    active_char_id = char_id

    char = get_character(char_id)
    return jsonify({
        'id': char_id,
        'name': name,
        'hp_max': hp_est,
        'hp_current': hp_est,
        'ac': ac_est,
        'formatted': format_character_sheet(char),
    })


@app.route('/api/character/<name>/use', methods=['POST'])
def api_set_active(name):
    """设置活跃角色"""
    char = _resolve_char(name)
    if not char:
        return jsonify({'error': '角色不存在'}), 404
    global active_char_id
    active_char_id = char['id']
    # 同步 CLI 的活跃角色（随机事件自动添加物品需要）
    _bot.set_active_char(active_char_id)
    return jsonify({'success': True, 'name': char['name']})


@app.route('/api/character/import', methods=['POST'])
def api_import_character():
    """从 Excel 文件导入角色。

    接受 multipart/form-data，字段名 'file'。
    返回导入的角色信息。
    """
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件（字段名: file）'}), 400

    file = request.files['file']
    if not file or not file.filename or file.filename.strip() == '':
        return jsonify({'error': '未选择文件'}), 400

    # 检查文件扩展名
    try:
        ext = _os.path.splitext(file.filename)[1].lower()
    except Exception:
        return jsonify({'error': '无法解析文件名'}), 400
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'error': f'不支持的文件类型: {ext}，请上传 .xlsx 或 .xls 文件'}), 400

    # 保存临时文件
    import tempfile
    tmp = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        file.save(tmp.name)
        tmp.close()

        # 预览
        preview = import_and_print_summary(tmp.name)
        # 导入（Excel 解析结果需扁平化）
        raw = import_character_from_excel(tmp.name)
        basic = raw.get('basic', {})
        combat = raw.get('combat', {})
        abilities = raw.get('abilities', {})
        spell_info = raw.get('spell_info', {})
        spell_slots = raw.get('spell_slots', {})
        coins = raw.get('coins', {})
        bg_data = raw.get('background', {})
        data = {
            'name': basic.get('name', '未命名'),
            'player': basic.get('player', ''),
            'class': basic.get('class', ''),
            'level': basic.get('level', 1),
            'race': basic.get('race', ''),
            'subrace': basic.get('subrace', ''),
            'background': _json.dumps(bg_data, ensure_ascii=False) if bg_data else '',
            'alignment': basic.get('alignment', ''),
            'faith': basic.get('faith', ''),
            'gender': basic.get('gender', ''),
            'age': basic.get('age', ''),
            'height': basic.get('height', ''),
            'weight': basic.get('weight', ''),
            'hp_max': combat.get('hp_max', 10),
            'hp_current': combat.get('hp_current', 10),
            'ac': combat.get('ac', 10),
            'initiative_bonus': combat.get('initiative_bonus', 0),
            'speed': combat.get('speed', 30),
            'proficiency_bonus': basic.get('proficiency_bonus', 2),
            'hd_count': combat.get('hd_count', 1),
            'hit_dice': combat.get('hd_type', '1d8'),
            'xp': basic.get('xp', 0),
            'key_abilities': raw.get('key_abilities', basic.get('key_abilities', '')),
            'resistances': raw.get('resistances', basic.get('resistances', '')),
            'passive_perception': combat.get('passive_perception', 10),
            'spellcasting_ability': spell_info.get('spellcasting_ability', ''),
            'spell_attack_bonus': spell_info.get('spell_attack_bonus', 0),
            'spell_save_dc': spell_info.get('spell_save_dc', 10),
            'prepared_spell_count': spell_info.get('prepared_spell_count', 0),
            'portrait_path': raw.get('portrait_path', ''),
            'abilities': abilities,
            'save_proficiencies': raw.get('save_proficiencies', {}),
            'skill_proficiencies': raw.get('skill_proficiencies', {}),
            'weapons': raw.get('weapons', []),
            'armor': raw.get('armor', {}),
            'spell_slots': spell_slots,
            'prepared_spells': raw.get('prepared_spells', []),
            'coins': coins,
            'weight_data': raw.get('weight', {}),
            'background_data': bg_data,
            'inventory': raw.get('inventory', []),
            'features': raw.get('features', []),
        }
        created_by = request.form.get('created_by', '')
        char_id = import_from_excel_data(data, source_file=_os.path.abspath(tmp.name), created_by=created_by)
        char = get_character(char_id)

        # 设为活跃角色
        global active_char_id
        active_char_id = char_id

        # 调试：提取到的特性数据
        features_raw = raw.get('features', [])
        features_debug = [{'name': f['name'], 'cat': f['category'],
                           'desc': (f.get('description','') or '')[:80]}
                          for f in features_raw[:10]]
        return jsonify({
            'id': char_id,
            'name': char['name'],
            'level': char.get('level', 1),
            'class': char.get('class', ''),
            'race': char.get('race', ''),
            'hp_current': char.get('hp_current', 0),
            'hp_max': char.get('hp_max', 0),
            'ac': char.get('ac', 10),
            'preview': preview,
            'formatted': format_character_sheet(char),
            '_debug': {
                'features_count': len(features_raw),
                'features_sample': features_debug,
            },
        })
    except Exception as e:
        return jsonify({'error': f'导入失败: {type(e).__name__}: {e}'}), 500
    finally:
        if tmp is not None:
            try:
                _os.unlink(tmp.name)
            except Exception:
                pass


# ━━━ 北境酒馆 Excel 导入（独立，不写入 characters 表）━━━

@app.route('/api/tavern/character/import', methods=['POST'])
def api_tavern_character_import():
    """北境酒馆专用：解析 Excel 角色卡并返回完整数据，不写入数据库。

    与 /api/character/import 的区别：只解析 Excel，不调用 import_from_excel_data，
    因此不会写入 characters 表，跑团平台不可见。
    """
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件（字段名: file）'}), 400

    file = request.files['file']
    if not file or not file.filename or file.filename.strip() == '':
        return jsonify({'error': '未选择文件'}), 400

    try:
        ext = _os.path.splitext(file.filename)[1].lower()
    except Exception:
        return jsonify({'error': '无法解析文件名'}), 400
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'error': f'不支持的文件类型: {ext}，请上传 .xlsx 或 .xls 文件'}), 400

    import tempfile
    tmp = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        file.save(tmp.name)
        tmp.close()

        # 预览
        preview = import_and_print_summary(tmp.name)
        # 解析 Excel
        raw = import_character_from_excel(tmp.name)
        basic = raw.get('basic', {})
        combat = raw.get('combat', {})
        abilities = raw.get('abilities', {})
        spell_info = raw.get('spell_info', {})
        spell_slots = raw.get('spell_slots', {})
        coins = raw.get('coins', {})
        bg_data = raw.get('background', {})
        weapons = raw.get('weapons', [])
        inventory = raw.get('inventory', [])
        features = raw.get('features', [])

        # 转换武器格式
        weapons_out = []
        for w in weapons:
            weapons_out.append({
                'name': w.get('name', ''),
                'bonus': w.get('attack_bonus', 0),
                'damage': w.get('damage_dice', ''),
                'type': w.get('damage_type', ''),
            })

        # 转换物品格式
        inventory_out = []
        for it in inventory:
            inventory_out.append({
                'name': it.get('item_name', it.get('name', '')),
                'qty': it.get('quantity', 1),
                'location': it.get('location', '背包'),
            })

        # 转换特性格式（中文 category → 北境 key）
        cat_map = {
            '职业能力': 'class_feature', '种族特性': 'racial_trait',
            '专长': 'feat', '特殊能力': 'special_ability', '其他': 'other',
        }
        features_out = []
        for f in features:
            cat = f.get('category', '其他')
            features_out.append({
                'name': f.get('feature_name', f.get('name', '')),
                'cat': cat_map.get(cat, 'other'),
                'desc': f.get('description', ''),
            })

        return jsonify({
            'ok': True,
            'name': basic.get('name', '未命名'),
            'level': basic.get('level', 1),
            'class': basic.get('class', ''),
            'race': basic.get('race', ''),
            'hp_current': combat.get('hp_current', 10),
            'hp_max': combat.get('hp_max', 10),
            'ac': combat.get('ac', 10),
            'speed': combat.get('speed', 30),
            'proficiency_bonus': basic.get('proficiency_bonus', 2),
            'passive_perception': combat.get('passive_perception', 10),
            'spell_save_dc': spell_info.get('spell_save_dc', 10),
            'spell_attack_bonus': spell_info.get('spell_attack_bonus', 0),
            'initiative_bonus': combat.get('initiative_bonus', 0),
            'abilities': {
                'str': abilities.get('str', 10), 'dex': abilities.get('dex', 10),
                'con': abilities.get('con', 10), 'int': abilities.get('int', 10),
                'wis': abilities.get('wis', 10), 'cha': abilities.get('cha', 10),
            },
            'skill_proficiencies': raw.get('skill_proficiencies', {}),
            'save_proficiencies': raw.get('save_proficiencies', {}),
            'weapons': weapons_out,
            'inventory': inventory_out,
            'coins': {'cp': coins.get('cp', 0), 'sp': coins.get('sp', 0), 'gp': coins.get('gp', 0)},
            'spell_slots': {
                '1': spell_slots.get('1', 0), '2': spell_slots.get('2', 0),
                '3': spell_slots.get('3', 0), '4': spell_slots.get('4', 0),
                '5': spell_slots.get('5', 0), '6': spell_slots.get('6', 0),
                '7': spell_slots.get('7', 0), '8': spell_slots.get('8', 0),
                '9': spell_slots.get('9', 0),
            },
            'prepared_spells': raw.get('prepared_spells', []),
            'features': features_out,
            'background': {
                'personality': bg_data.get('personality_traits', ''),
                'ideals': bg_data.get('ideals', ''),
                'bonds': bg_data.get('bonds', ''),
                'flaws': bg_data.get('flaws', ''),
                'appearance': bg_data.get('appearance', ''),
                'backstory': bg_data.get('backstory', bg_data.get('background_feature', '')),
            },
            'alignment': basic.get('alignment', ''),
            'faith': basic.get('faith', ''),
            'gender': basic.get('gender', ''),
            'height': basic.get('height', ''),
            'weight': basic.get('weight', ''),
            'languages': basic.get('languages', '通用语'),
            'key_abilities': raw.get('key_abilities', basic.get('key_abilities', '')),
            'resistances': raw.get('resistances', basic.get('resistances', '')),
            'preview': preview,
        })
    except Exception as e:
        return jsonify({'error': f'导入失败: {type(e).__name__}: {e}'}), 500
    finally:
        if tmp is not None:
            try:
                _os.unlink(tmp.name)
            except Exception:
                pass


# ━━━ 随机事件 API ━━━

def _call_event_command(cmd_func, args: str = '') -> tuple[str | None, str | None]:
    """调用事件命令函数，自动处理交互式 input 提示（web 环境默认通过）。

    同时捕获 stdout（CLI 代码中的 print() 调用），确保输出完整。
    """
    original_input = builtins.input
    original_stdout = sys.stdout
    builtins.input = lambda prompt='': 'y'
    # 同步活跃角色到 CLI（确保事件物品自动添加到角色背包）
    if active_char_id is not None or _bot._active_char_id is not None:
        cid = active_char_id or _bot._active_char_id
        if cid:
            _bot.set_active_char(cid)
    captured = _io.StringIO()
    sys.stdout = captured
    try:
        output, error = cmd_func(args)
        printed = captured.getvalue()
        if printed and output:
            output = printed.rstrip('\n') + '\n' + output
        elif printed:
            output = printed.rstrip('\n')
        return output, error
    finally:
        builtins.input = original_input
        sys.stdout = original_stdout


def _strip_ansi(text: str) -> str:
    """去除 ANSI 颜色码，转换为 HTML 友好格式。"""
    # 粗体: \033[1m...\033[0m → <b>...</b>
    text = _re.sub(r'\x1b\[1m(.*?)\x1b\[0m', r'<b>\1</b>', text)
    # 红色
    text = _re.sub(r'\x1b\[31m(.*?)\x1b\[0m', r'<span style="color:#f44336">\1</span>', text)
    # 绿色
    text = _re.sub(r'\x1b\[32m(.*?)\x1b\[0m', r'<span style="color:#4caf50">\1</span>', text)
    # 青色
    text = _re.sub(r'\x1b\[36m(.*?)\x1b\[0m', r'<span style="color:#00bcd4">\1</span>', text)
    # 黄色
    text = _re.sub(r'\x1b\[33m(.*?)\x1b\[0m', r'<span style="color:#ffd700">\1</span>', text)
    # 剩余的裸 \033[0m（关闭标记）
    text = text.replace('\x1b[0m', '')
    # 裸 \033[1m（开始粗体，无配对的情况）
    text = text.replace('\x1b[1m', '<b>')
    # 裸 ANSI 码（未匹配的颜色开始）
    text = _re.sub(r'\x1b\[31m', '<span style="color:#f44336">', text)
    text = _re.sub(r'\x1b\[32m', '<span style="color:#4caf50">', text)
    text = _re.sub(r'\x1b\[36m', '<span style="color:#00bcd4">', text)
    text = _re.sub(r'\x1b\[33m', '<span style="color:#ffd700">', text)
    return text


@app.route('/api/events/<command>', methods=['POST'])
def api_event(command: str):
    """随机事件掷表 API。

    URL 参数:
        command  — 事件命令名 (如 zy100, ts100, hj6 等)
    Body (JSON):
        args     — 可选，命令参数（wp 命令使用，传入物品编号）

    返回掷表结果（HTML 格式，ANSI 颜色码已转换）。
    """
    if command not in _EVENT_COMMANDS:
        return jsonify({'error': f'未知事件命令: .{command}'}), 404

    data = request.get_json(silent=True) or {}
    cmd_args = data.get('args', '').strip()

    label, dice, cmd_func = _EVENT_COMMANDS[command]

    # 确保 CLI 活跃角色已同步
    if active_char_id is not None:
        _bot.set_active_char(active_char_id)

    try:
        output, error = _call_event_command(cmd_func, cmd_args)
    except Exception as e:
        return jsonify({'error': f'{type(e).__name__}: {e}'}), 500

    if error:
        return jsonify({'error': error}), 400

    # 转换为 HTML
    html_output = _strip_ansi(output or '')
    # 将换行转为 <br>，保留缩进空格
    html_output = html_output.replace('\n', '<br>')
    html_output = html_output.replace('  ', '&nbsp;&nbsp;')

    return jsonify({
        'command': command,
        'label': label,
        'dice': dice,
        'output': output or '',
        'html': html_output,
    })


@app.route('/api/events', methods=['GET'])
def api_list_events():
    """列出所有可用事件命令及分组。"""
    groups = []
    for group_label, items in _EVENT_GROUPS:
        group_items = []
        for cmd, label, dice, desc in items:
            group_items.append({
                'command': cmd,
                'label': label,
                'dice': dice,
                'description': desc,
            })
        groups.append({'label': group_label, 'items': group_items})
    return jsonify({'groups': groups})


# ━━━ 事件统计数据 ━━━
# 数据模型: {用户名: {表名: {事件内容: 次数}}}
# 例如: {"张三": {"随机环境": {"雪坑（需要过被动察觉...）": 3, "暴风雪...": 2}}}
_event_stats_lock = _threading.Lock()


def _load_event_stats() -> dict:
    """从 PostgreSQL 加载所有用户的事件统计数据（原 event_stats.json）"""
    try:
        from core.models import EventStat as _ES
        rows = _ES.query.all()
        return {r.username: (r.data if isinstance(r.data, dict) else {}) for r in rows}
    except Exception:
        return {}


def _save_event_stats(stats: dict):
    """保存事件统计数据到 PostgreSQL"""
    try:
        from core.models import EventStat as _ES
        with db.engine.begin() as conn:
            conn.execute(_ES.__table__.delete())
            for uname, data in (stats or {}).items():
                conn.execute(_ES.__table__.insert().values(username=uname, data=data))
    except Exception:
        pass


def _extract_event_name(text: str) -> str:
    """从事件描述文本中提取简短的事件名称。

    例如 "商队（可交易）" → "商队"
         "雪坑（需要过被动察觉...）" → "雪坑"
         "冰原狼（1d2只）" → "冰原狼"
    """
    if not text:
        return text
    # 取括号前的内容作为事件名称
    idx = text.find('（')
    if idx > 0:
        return text[:idx].strip()
    idx = text.find('(')
    if idx > 0:
        return text[:idx].strip()
    return text.strip()


@app.route('/api/events/stats', methods=['POST'])
def api_record_event():
    """记录事件触发（含具体事件内容）。

    请求体: { name, command, events: [{table, content}] }
       table   — 表名（如"随机环境"）
       content — 事件内容描述（将提取短名称存储）
    同一次掷表可能产生多个事件（含链式子表），一次性全部记录。
    """
    data = request.get_json(silent=True) or {}
    username = data.get('name', '').strip()

    # 清空统计
    if data.get('clear'):
        if not username:
            return jsonify({'error': '需要 name 参数'}), 400
        stats = _load_event_stats()
        if username in stats:
            del stats[username]
            _save_event_stats(stats)
            return jsonify({'ok': True})
        return jsonify({'ok': True, 'message': '用户无统计数据'})

    command = data.get('command', '')
    event_list = data.get('events', [])

    if not username:
        return jsonify({'error': '需要 name 参数'}), 400
    if not event_list:
        return jsonify({'error': '需要 events 参数'}), 400

    stats = _load_event_stats()
    user_stats = stats.setdefault(username, {})

    # 每次点击事件按钮计数一次（不含链式子事件）
    user_stats['_clicks'] = user_stats.get('_clicks', 0) + 1

    for evt in event_list:
        table = evt.get('table', '').strip()
        content = evt.get('content', '').strip()
        if not table or not content:
            continue
        short_name = _extract_event_name(content)
        table_events = user_stats.setdefault(table, {})
        table_events[short_name] = table_events.get(short_name, 0) + 1

    _save_event_stats(stats)
    return jsonify({'ok': True})


@app.route('/api/events/stats', methods=['GET'])
def api_get_event_stats():
    """获取指定用户的事件内容统计数据。

    URL 参数: name — 用户名
    返回:
      { name, total, groups: [{table, dice, total_in_table, events: [{content, count, probability}]}] }
      probability = 该事件次数 / 全局总事件次数 × 100
    """
    username = request.args.get('name', '').strip()
    if not username:
        return jsonify({'error': '需要 name 参数'}), 400

    stats = _load_event_stats()
    user_stats = stats.get(username, {})

    # 总点击次数（不含链式子事件）
    total_clicks = user_stats.get('_clicks', 0)
    # 兼容旧数据：没有 _clicks 时用事件总和
    if total_clicks == 0:
        for table_events in user_stats.values():
            if isinstance(table_events, dict):
                total_clicks += sum(table_events.values())
    global_total = total_clicks

    # 构建分组数据
    groups = []
    # 按表的事件总数降序排列（排除 _clicks 等非字典字段）
    table_order = sorted(
        [(k, v) for k, v in user_stats.items() if isinstance(v, dict)],
        key=lambda x: -sum(x[1].values())
    )

    for table_name, table_events in table_order:
        if not isinstance(table_events, dict):
            continue
        table_total = sum(table_events.values())
        # 获取该表的骰子信息
        dice = ''
        for cmd, (label, d, _) in _EVENT_COMMANDS.items():
            if label == table_name:
                dice = d
                break

        events = []
        for content, count in sorted(table_events.items(), key=lambda x: -x[1]):
            table_prob = (count / table_total * 100) if table_total > 0 else 0
            global_prob = (count / global_total * 100) if global_total > 0 else 0
            events.append({
                'content': content,
                'count': count,
                'probability': round(table_prob, 1),
                'global_probability': round(global_prob, 1),
            })

        groups.append({
            'table': table_name,
            'dice': dice,
            'total_in_table': table_total,
            'events': events,
        })

    return jsonify({
        'name': username,
        'total': global_total,
        'groups': groups,
    })


# ━━━ DND 物品表加载 ━━━
import openpyxl
import os as _os

_item_cache: list[dict] | None = None

def _load_items():
    """加载 DND 物品表 Excel"""
    global _item_cache
    if _item_cache is not None:
        return _item_cache

    script_dir = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    path = _os.path.join(script_dir, 'data', 'DND5E物品表.xlsx')
    if not _os.path.exists(path):
        _item_cache = []
        return _item_cache

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        items = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[1]:
                continue
            items.append({
                'id': str(row[0] or ''),
                'name': str(row[1] or '').strip(),
                'price': str(row[2] or '').strip(),
                'damage': str(row[3] or '').strip(),
                'weight': str(row[4] or '').strip(),
                'type': str(row[5] or '').strip(),
            })
    except Exception:
        items = []

    # 合并随机事件表中的物品（29项），自动去重
    seen_names = {it['name'].lower() for it in items}
    for num, desc in sorted(_bot.ITEM_TABLE.items()):
        # 提取物品名和括号内描述
        name = desc.split('（')[0].split('(')[0].strip()
        # 提取括号内容作为描述
        detail = ''
        for sep in ['（', '(']:
            if sep in desc:
                rest = desc.split(sep, 1)[1]
                end = rest.rfind('）') if '）' in rest else rest.rfind(')')
                if end > 0:
                    detail = rest[:end].strip()
                else:
                    detail = rest.strip().rstrip('）').rstrip(')')
                break
        if name.lower() not in seen_names:
            seen_names.add(name.lower())
            items.append({
                'id': f'wp{num}',
                'name': name,
                'price': '',
                'damage': '',
                'weight': '',
                'type': '随机事件物品',
                'detail': detail,
            })

    _item_cache = items
    return items


# ━━━ 规则/文档详情 API ━━━

@app.route('/api/rule-detail', methods=['GET'])
def api_rule_detail():
    """读取 CHM 页面的完整内容"""
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'error': '缺少 path 参数'}), 400

    from core.chm_search import read_detail_page as _read
    content = _read(path)
    if content is None:
        return jsonify({'error': '无法读取页面'}), 404
    return jsonify({'path': path, 'content': content[:50000]})


@app.route('/api/file-content', methods=['GET'])
def api_file_content():
    """读取项目文件内容"""
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'error': '缺少 path 参数'}), 400

    # 安全检查：解析真实路径并确保在项目根目录内
    script_dir = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    project_root = _os.path.dirname(script_dir) if _os.path.basename(script_dir) == '骰娘' else script_dir
    full_path = _os.path.realpath(_os.path.join(project_root, path))
    if not full_path.startswith(_os.path.realpath(project_root) + _os.sep):
        return jsonify({'error': '路径越权'}), 403
    if not _os.path.exists(full_path):
        return jsonify({'error': '文件不存在'}), 404

    ext = _os.path.splitext(full_path)[1].lower()

    # 图片文件 → 返回文件内容供前端直接展示
    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'):
        return jsonify({'path': path, 'type': 'image', 'content': f'/api/file-raw?path={path}'})

    try:
        if ext == '.txt':
            with open(full_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        elif ext in ('.md', '.markdown'):
            with open(full_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        elif ext == '.docx':
            try:
                import docx
                doc = docx.Document(full_path)
                content = '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
            except ImportError:
                content = '[需要安装 python-docx 库来读取 .docx 文件]'
        else:
            try:
                with open(full_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
            except Exception:
                return jsonify({'error': f'不支持的文件类型: {ext}'}), 400
    except Exception as e:
        return jsonify({'error': f'读取失败: {str(e)}'}), 500

    return jsonify({'path': path, 'content': content[:50000]})


@app.route('/api/file-raw', methods=['GET'])
def api_file_raw():
    """直接返回项目文件的原始内容（图片等）"""
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'error': '缺少 path 参数'}), 400

    script_dir = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    project_root = _os.path.dirname(script_dir) if _os.path.basename(script_dir) == '骰娘' else script_dir
    full_path = _os.path.realpath(_os.path.join(project_root, path))
    if not full_path.startswith(_os.path.realpath(project_root) + _os.sep):
        return jsonify({'error': '路径越权'}), 403
    if not _os.path.exists(full_path):
        abort(404)

    ext = _os.path.splitext(full_path)[1].lower()
    mimetypes = {
        '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
        '.pdf': 'application/pdf',
    }
    mimetype = mimetypes.get(ext, 'application/octet-stream')
    return send_file(full_path, mimetype=mimetype)


# ━━━ 物品/装备搜索 API ━━━

@app.route('/api/items/search', methods=['GET'])
def api_search_items():
    """搜索 DND 物品表（武器/装备/物品）"""
    query = request.args.get('q', '').strip()
    if len(query) < 1:
        return jsonify({'error': '请输入搜索关键词'}), 400

    items = _load_items()
    qlower = query.lower()
    results = []
    for item in items:
        name = item['name']
        dtype = item['type']
        if qlower in name.lower() or qlower in dtype.lower():
            detail_parts = []
            if item.get('price'): detail_parts.append(f"💰{item['price']}")
            if item.get('damage'): detail_parts.append(f"⚔️{item['damage']}")
            if item.get('weight'): detail_parts.append(f"⚖️{item['weight']}")
            if item.get('type'): detail_parts.append(f"📦{item['type']}")
            # 随机事件物品的描述（括号内内容）
            if item.get('detail'): detail_parts.append(item['detail'])
            results.append({
                'name': name,
                'type': 'item',
                'detail': ' | '.join(detail_parts),
                'price': item.get('price', ''),
                'damage': item.get('damage', ''),
                'weight': item.get('weight', ''),
                'item_type': item.get('type', ''),
            })
        if len(results) >= 50:
            break

    return jsonify({'query': query, 'results': results, 'total': len(results)})


# ━━━ 自定义资料库 API ━━━
import json as _json

_CUSTOM_DIR = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), 'data', 'custom')


def _load_custom(kind: str) -> list[dict]:
    path = _os.path.join(_CUSTOM_DIR, f'{kind}.json')
    if not _os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return _json.load(f)
    except Exception:
        return []


def _save_custom(kind: str, data: list[dict]):
    _os.makedirs(_CUSTOM_DIR, exist_ok=True)
    path = _os.path.join(_CUSTOM_DIR, f'{kind}.json')
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(data, f, ensure_ascii=False, indent=2)


@app.route('/api/custom/<kind>', methods=['GET'])
def api_list_custom(kind):
    """列出自定义条目（spells / monsters / items）"""
    if kind not in ('spells', 'monsters', 'items'):
        return jsonify({'error': '无效类型'}), 400
    return jsonify({'results': _load_custom(kind)})


@app.route('/api/custom/<kind>', methods=['POST'])
def api_add_custom(kind):
    """添加自定义条目"""
    if kind not in ('spells', 'monsters', 'items'):
        return jsonify({'error': '无效类型'}), 400
    data = request.get_json() or {}
    if not data.get('name', '').strip():
        return jsonify({'error': '名称不能为空'}), 400

    existing = _load_custom(kind)
    entry = {'name': data['name'].strip()}
    if kind == 'spells':
        entry.update({
            'name_en': data.get('name_en', ''),
            'level': data.get('level', ''),
            'school': data.get('school', ''),
            'casting_time': data.get('casting_time', ''),
            'range': data.get('range', ''),
            'duration': data.get('duration', ''),
            'components': data.get('components', ''),
            'ritual': data.get('ritual', '否'),
            'concentration': data.get('concentration', '否'),
            'classes': data.get('classes', ''),
            'description': data.get('description', ''),
            'source': '自定义',
        })
    elif kind == 'monsters':
        entry.update({
            'name_en': data.get('name_en', ''),
            'cr': data.get('cr', ''),
            'size': data.get('size', ''),
            'type': data.get('type', ''),
            'legendary': data.get('legendary', ''),
            'detail_text': data.get('detail_text', ''),
            'source': '自定义',
        })
    elif kind == 'items':
        entry['description'] = data.get('description', '')
    existing.append(entry)
    _save_custom(kind, existing)
    return jsonify({'success': True, 'name': entry['name']})


@app.route('/api/custom/<kind>', methods=['DELETE'])
def api_delete_custom(kind):
    """删除自定义条目"""
    if kind not in ('spells', 'monsters', 'items'):
        return jsonify({'error': '无效类型'}), 400
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'error': '缺少 name 参数'}), 400
    existing = _load_custom(kind)
    new_list = [e for e in existing if e.get('name') != name]
    if len(new_list) == len(existing):
        return jsonify({'error': '未找到该条目'}), 404
    _save_custom(kind, new_list)
    return jsonify({'success': True})


# ━━━ 服务器端地图加载 ━━━

MAPS_DIR = _Path(__file__).parent.parent / 'maps'
MAPS_DIR.mkdir(exist_ok=True)
_ALLOWED_MAP_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}


@app.route('/maps/<path:filename>')
def serve_map(filename):
    """Serve map image files from the maps/ directory.

    如果文件不存在，尝试从共享画布图层状态中恢复（dataURL 回退）。
    """
    filepath = (MAPS_DIR / filename).resolve()
    if not str(filepath).startswith(str(MAPS_DIR.resolve())):
        return jsonify({'error': 'Access denied'}), 403
    if not filepath.exists() or not filepath.is_file():
        # 自愈：文件丢失时尝试从共享画布状态恢复
        # 文件名如 "layers/1.png" → 提取 layer_id=1
        if filename.startswith('layers/') or filename.startswith('layers\\'):
            try:
                name_part = filename.split('/')[-1].split('\\')[-1]
                layer_id = int(name_part.split('.')[0])
                canvas = get_shared_canvas()
                for layer in canvas.get('layers', []):
                    if layer.get('id') == layer_id:
                        data_url = layer.get('dataURL', '')
                        if data_url and data_url.startswith('data:image'):
                            import base64
                            try:
                                header, b64 = data_url.split(',', 1)
                                img_data = base64.b64decode(b64)
                                # 异步写回磁盘
                                _threading.Thread(target=save_layer_image, args=(layer_id, data_url), daemon=True).start()
                                mime = 'image/png'
                                if 'image/jpeg' in header or 'image/jpg' in header:
                                    mime = 'image/jpeg'
                                elif 'image/gif' in header:
                                    mime = 'image/gif'
                                elif 'image/webp' in header:
                                    mime = 'image/webp'
                                return app.response_class(img_data, mimetype=mime)
                            except Exception:
                                pass
                        break
            except (ValueError, IndexError):
                pass
        return jsonify({'error': 'File not found'}), 404
    ext = filepath.suffix.lower()
    mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
                '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
    return send_file(str(filepath), mimetype=mime_map.get(ext, 'image/png'))


@app.route('/api/server-maps')
def api_server_maps():
    """List all map files in the server maps/ directory."""
    maps = []
    for f in sorted(MAPS_DIR.iterdir()):
        if f.is_file() and f.suffix.lower() in _ALLOWED_MAP_EXTS:
            size_kb = f.stat().st_size / 1024
            maps.append({
                'name': f.name,
                'size_kb': round(size_kb, 1),
                'url': f'/maps/{f.name}',
            })
    return jsonify({'maps': maps})


# ━━━ 服务器端资源库 ━━━

RESOURCES_DIR = _Path(__file__).parent.parent / 'resources'
RESOURCES_DIR.mkdir(exist_ok=True)

_RESOURCE_CATEGORIES = {
    '.pdf': ('📄 PDF', 'pdf'),
    '.jpg': ('🖼 图片', 'img'), '.jpeg': ('🖼 图片', 'img'), '.png': ('🖼 图片', 'img'),
    '.gif': ('🖼 图片', 'img'), '.webp': ('🖼 图片', 'img'), '.bmp': ('🖼 图片', 'img'),
    '.svg': ('🖼 图片', 'img'),
    '.txt': ('📝 文本', 'text'), '.md': ('📝 文本', 'text'),
    '.docx': ('📋 文档', 'doc'), '.doc': ('📋 文档', 'doc'), '.pptx': ('📋 文档', 'doc'), '.ppt': ('📋 文档', 'doc'),
    '.xlsx': ('📊 表格', 'sheet'), '.xls': ('📊 表格', 'sheet'), '.csv': ('📊 表格', 'sheet'),
    '.json': ('💾 存档', 'save'),
    '.mp3': ('🎵 音频', 'audio'), '.wav': ('🎵 音频', 'audio'), '.ogg': ('🎵 音频', 'audio'), '.flac': ('🎵 音频', 'audio'),
}


@app.route('/api/resources')
def api_resources():
    """List all files in the resources directory, optionally filtered by category."""
    category = request.args.get('cat', '')
    items = []
    try:
        files = sorted(RESOURCES_DIR.iterdir())
    except Exception as e:
        return jsonify({'error': f'无法读取资源目录: {e}', 'resources': []}), 500
    for f in files:
        if not f.is_file():
            continue
        try:
            ext = f.suffix.lower()
            info = _RESOURCE_CATEGORIES.get(ext, ('📦 其他', 'other'))
            if category and info[1] != category:
                continue
            size_kb = f.stat().st_size / 1024
            items.append({
                'name': f.name,
                'size_kb': round(size_kb, 1),
                'size_display': f'{size_kb:.1f}KB' if size_kb < 1024 else f'{size_kb/1024:.1f}MB',
                'ext': ext.lstrip('.'),
                'icon': info[0].split()[0],
                'category': info[1],
                'url': f'/resources/{f.name}',
            })
        except Exception:
            continue
    return jsonify({'resources': items})


@app.route('/resources/<path:filename>')
def serve_resource(filename):
    """Serve resource files from the resources/ directory."""
    filepath = (RESOURCES_DIR / filename).resolve()
    if not str(filepath).startswith(str(RESOURCES_DIR.resolve())):
        return jsonify({'error': 'Access denied'}), 403
    if not filepath.exists() or not filepath.is_file():
        return jsonify({'error': 'File not found'}), 404
    return send_file(str(filepath))


# ━━━ 聊天系统 API ━━━

def _is_dm_ip(ip: str) -> bool:
    """Check if the IP belongs to the DM (localhost or server's own LAN IP).

    注意：使用 frp/内网穿透时，所有远程连接的 remote_addr 都显示为 127.0.0.1，
    因此 _is_dm_ip() 对所有人返回 True。此时角色判定依赖用户选择而非 IP。
    """
    # 如果有多个用户共享同一 IP（frp 特征），IP 检测不可靠，回退到用户选择
    global _online_users
    same_ip_users = [n for n, u in _online_users.items() if u.get('ip') == ip]
    if len(same_ip_users) > 1:
        return False  # frp 环境：IP 不能区分用户，不自动提权

    if ip in ('127.0.0.1', 'localhost', '::1', '0:0:0:0:0:0:0:1'):
        return True
    try:
        import socket
        hostname = socket.gethostname()
        local_ips = socket.gethostbyname_ex(hostname)[2]
        if ip in local_ips:
            return True
    except Exception:
        pass
    return False


@app.route('/api/chat/send', methods=['POST'])
def api_chat_send():
    """Send a chat message."""
    global _dm_name, _dm_ip
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    text = data.get('text', '').strip()
    role = data.get('role', 'PL')
    color = data.get('color', '') or ''
    if not name or not text:
        return jsonify({'ok': False, 'error': 'Name and text required'})
    if len(text) > 2000:
        return jsonify({'ok': False, 'error': 'Message too long (max 2000 chars)'})

    client_ip = request.remote_addr or 'unknown'
    is_dm_ip = _is_dm_ip(client_ip)

    # 只有第一个以DM身份发言的localhost用户是DM，锁定后不更改
    # （PL角色发言不触发自动注册，避免PL被误识别为DM）
    if is_dm_ip and _dm_name is None and role == 'DM':
        _dm_name = name
        _dm_ip = client_ip

    # 只有已注册的DM本人才显示DM标识
    is_dm = (name == _dm_name)

    # 如果有人声称是DM的名字但不是DM的IP，拒绝
    if name == _dm_name and not is_dm and _dm_ip is not None:
        is_dm = False  # 不是真正的DM

    # 同步在线列表中的角色
    if name in _online_users:
        role = _online_users[name].get('role', role)
    msg = {
        'name': name,
        'text': text,
        'time': _time.strftime('%H:%M:%S'),
        'is_dm': is_dm,
        'color': color,
        'role': role,
        'ip': client_ip,
        '_ts': _time.time(),
    }
    _chat_messages.append(msg)

    # ━━ 解析 @提及 ━━
    import re as _re
    mentioned_names = set()
    for m in _re.finditer(r'@(\S+)', text):
        target = m.group(1).rstrip('，,。.！!？?：:）)】】》>')
        if target and target != name:  # 不提及自己
            mentioned_names.add(target)

    # 为每个被提及的在线用户创建通知
    for target_name in mentioned_names:
        if target_name in _online_users:
            _mentions.append({
                'target_name': target_name,
                'from_name': name,
                'text': text,
                'time': _time.strftime('%H:%M:%S'),
                '_ts': _time.time(),
            })
    # Trim old mentions
    while len(_mentions) > MAX_MENTIONS:
        _mentions.pop(0)
    # 裁剪并归档旧消息
    _trim_and_archive_chat()
    # 异步保存聊天记录到磁盘
    _threading.Thread(target=_save_chat_log, daemon=True).start()
    return jsonify({'ok': True, 'msg': msg})


@app.route('/api/chat/messages')
def api_chat_messages():
    """Get chat messages. ?since=<timestamp_float> for incremental polling."""
    since = request.args.get('since', '')
    if since:
        try:
            since_ts = float(since)
            new_msgs = [m for m in _chat_messages if m.get('_ts', 0) > since_ts]
            return jsonify({'ok': True, 'messages': new_msgs})
        except ValueError:
            pass
    return jsonify({'ok': True, 'messages': _chat_messages[-50:]})  # Last 50


# ━━━ 酒馆聊天系统 API ━━━

@app.route('/api/tavern/chat/send', methods=['POST'])
def api_tavern_chat_send():
    """发送酒馆聊天消息"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    text = data.get('text', '').strip()
    color = data.get('color', '#d4a050')
    role = data.get('role', 'PL')
    if not name or not text:
        return jsonify({'ok': False, 'error': 'Name and text required'})
    if len(text) > 2000:
        return jsonify({'ok': False, 'error': 'Message too long (max 2000 chars)'})

    msg = {
        'name': name,
        'text': text,
        'time': _time.strftime('%H:%M:%S'),
        'color': color,
        'role': role,
        '_ts': _time.time(),
    }
    _tavern_messages.append(msg)
    _trim_and_archive_tavern_chat()
    _threading.Thread(target=_save_tavern_chat_log, daemon=True).start()
    return jsonify({'ok': True, 'msg': msg})


@app.route('/api/tavern/chat/messages')
def api_tavern_chat_messages():
    """获取酒馆聊天消息。?since=<timestamp> 增量查询"""
    since = request.args.get('since', '')
    if since:
        try:
            since_ts = float(since)
            new_msgs = [m for m in _tavern_messages if m.get('_ts', 0) > since_ts]
            return jsonify({'ok': True, 'messages': new_msgs})
        except ValueError:
            pass
    return jsonify({'ok': True, 'messages': _tavern_messages[-50:]})  # Last 50


# ━━━ DM 状态 API ━━━

def _is_dm_request(name: str, client_ip: str, user_id=None) -> bool:
    """判断请求者是否为DM。优先级：user_id > name > IP（与 api_dm_status 逻辑一致）。"""
    global _dm_name, _dm_ip, _dm_user_id
    if _dm_name is not None:
        if user_id is not None and _dm_user_id is not None:
            return user_id == _dm_user_id
        if name:
            return name == _dm_name
        return _is_dm_ip(client_ip) and client_ip == _dm_ip
    return _is_dm_ip(client_ip)


# ━━━ DM 事件系统 API（发布事件 / 全平台弹窗通知 / DM事件列表）━━━

@app.route('/api/dm/event-publish', methods=['POST'])
def api_dm_event_publish():
    """DM 发布事件：聊天室同步显示事件消息 + 广播全平台弹窗通知。"""
    global _event_notifications
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    client_ip = request.remote_addr or 'unknown'

    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})
    if not _is_dm_request(name, client_ip):
        return jsonify({'ok': False, 'error': '只有DM可以发布事件'})
    if not title:
        return jsonify({'ok': False, 'error': '事件标题不能为空'})
    if not content:
        return jsonify({'ok': False, 'error': '事件内容不能为空'})
    if len(title) > 100:
        return jsonify({'ok': False, 'error': '事件标题过长（最多100字）'})
    if len(content) > 2000:
        return jsonify({'ok': False, 'error': '事件内容过长（最多2000字）'})

    # 指令式发布解析："事件 {身份}：{事件名} {内容}" ——
    # 标题形如 "{发布者名}：{事件名} {内容}" 时拆分为事件名+内容（身份用于消息署名）
    ev_title, ev_content = title, content
    _prefix = name + '：' if name else ''
    if not _prefix:
        _prefix = name + ':' if name else ''
    if _prefix and title.startswith(_prefix):
        rest = title[len(_prefix):].strip()
        if rest:
            if ' ' in rest:
                sp = rest.split(' ', 1)
                ev_title, ev_content = sp[0].strip(), (sp[1].strip() or content)
            else:
                ev_title = rest

    ts = _time.time()

    # 0. 创建发布历史记录（供 DM 管理页撤回/置顶）
    pub_id = max([e.get('id', 0) for e in _dm_published_events], default=0) + 1
    pub_record = {
        'id': pub_id,
        'title': ev_title,
        'content': ev_content,
        'publisher': name,
        'published_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
        'recalled_at': None,
        'pinned': False,
    }
    _dm_published_events.append(pub_record)
    _save_dm_published_events()

    # 1. 聊天室同步显示（事件消息带发布者署名，格式"事件名：事件内容"）
    _chat_messages.append({
        'name': name,
        'text': ev_title + '：' + ev_content,
        'time': _time.strftime('%H:%M:%S'),
        'is_dm': True,
        'color': '#ffd700',
        'event': True,
        'event_title': ev_title,
        'event_content': ev_content,
        'event_id': pub_id,
        'ip': 'system',
        '_ts': ts,
    })
    _trim_and_archive_chat()
    _threading.Thread(target=_save_chat_log, daemon=True).start()

    # 2. 全平台弹窗通知广播（各页面 event-popup.js 轮询拉取）
    _event_notifications.append({
        'event_id': pub_id,
        'title': ev_title,
        'content': ev_content,
        'publisher': name,
        'time': _time.strftime('%H:%M:%S'),
        '_ts': ts,
        'recalled': False,
        'pinned': False,
    })
    while len(_event_notifications) > MAX_EVENT_NOTIFICATIONS:
        _event_notifications.pop(0)

    return jsonify({'ok': True, 'event': pub_record})


@app.route('/api/events/notifications', methods=['GET'])
def api_event_notifications():
    """各页面轮询：获取该用户需要弹窗显示的新事件。仅已加入房间的用户返回事件。"""
    name = request.args.get('name', '').strip()
    since_str = request.args.get('since', '')
    if not name or name not in _online_users:
        return jsonify({'ok': True, 'events': []})
    since_ts = 0.0
    if since_str:
        try:
            since_ts = float(since_str)
        except ValueError:
            pass
    # 过滤已撤回的；未撤回事件仅返回新发布的
    events = [e for e in _event_notifications
              if e.get('_ts', 0) > since_ts and not e.get('recalled', False)]
    # 追加当前置顶中的事件（_is_pinned 标记，前端关闭后记录已读，不再重复弹）
    for ev in _dm_published_events:
        if ev.get('pinned') and not ev.get('recalled_at'):
            events.append({
                'event_id': ev['id'],
                'title': ev['title'],
                'content': ev['content'],
                'time': ev.get('published_at', '')[-8:],
                '_ts': 0,
                '_is_pinned': True,
                'recalled': False,
                'pinned': True,
            })
    return jsonify({'ok': True, 'events': events})


@app.route('/api/dm/events', methods=['GET'])
def api_dm_events_list():
    """获取DM事件列表（仅DM可见）。"""
    name = request.args.get('name', '').strip()
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以查看事件列表'})
    return jsonify({'ok': True, 'events': _dm_event_list})


@app.route('/api/dm/events', methods=['POST'])
def api_dm_events_save():
    """DM 保存事件到事件列表（带 id 则更新，否则新增）。玩家不可见。"""
    global _dm_event_list
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以保存事件'})
    if not title:
        return jsonify({'ok': False, 'error': '事件标题不能为空'})
    if not content:
        return jsonify({'ok': False, 'error': '事件内容不能为空'})
    if len(title) > 100 or len(content) > 2000:
        return jsonify({'ok': False, 'error': '标题或内容过长'})

    eid = data.get('id')
    if eid is not None:
        # 更新已有事件
        for ev in _dm_event_list:
            if ev.get('id') == int(eid):
                ev['title'] = title
                ev['content'] = content
                _save_dm_event_list()
                return jsonify({'ok': True, 'event': ev})
        return jsonify({'ok': False, 'error': '事件不存在'})

    # 新增
    new_id = max([ev.get('id', 0) for ev in _dm_event_list], default=0) + 1
    ev = {
        'id': new_id,
        'title': title,
        'content': content,
        'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
    }
    _dm_event_list.append(ev)
    _save_dm_event_list()
    return jsonify({'ok': True, 'event': ev})


@app.route('/api/dm/events/<int:eid>', methods=['DELETE'])
def api_dm_events_delete(eid):
    """DM 删除事件列表中的事件。"""
    global _dm_event_list
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or request.args.get('name') or '').strip()
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以删除事件'})
    for i, ev in enumerate(_dm_event_list):
        if ev.get('id') == eid:
            _dm_event_list.pop(i)
            _save_dm_event_list()
            return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': '事件不存在'})


# ━━━ DM 已发布事件管理（撤回 / 置顶 / 发布历史）━━━

@app.route('/api/dm/published-events', methods=['GET'])
def api_dm_published_events():
    """DM 获取已发布事件历史（含撤回/置顶状态）。仅DM可见。"""
    name = request.args.get('name', '').strip()
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以查看发布历史'})
    return jsonify({'ok': True, 'events': list(reversed(_dm_published_events))})


@app.route('/api/dm/events/<int:pid>/recall', methods=['POST'])
def api_dm_event_recall(pid):
    """DM 撤回已发布事件：标记撤回，聊天室对应消息显示"已撤回"，全平台弹窗不再推送。"""
    global _chat_messages, _event_notifications
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or request.args.get('name') or '').strip()
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以撤回事件'})

    pub = None
    for ev in _dm_published_events:
        if ev.get('id') == pid:
            pub = ev
            break
    if not pub:
        return jsonify({'ok': False, 'error': '事件不存在'})
    if pub.get('recalled_at'):
        return jsonify({'ok': False, 'error': '该事件已撤回'})

    pub['recalled_at'] = _time.strftime('%Y-%m-%d %H:%M:%S')
    pub['pinned'] = False
    _save_dm_published_events()

    # 聊天室对应消息标记撤回（内存 + 磁盘存档 + 归档文件）。
    # 同时刷新 _ts：已打开聊天页的玩家通过增量轮询收到该消息，
    # 前端据此把界面上的事件消息移除（撤回后完全不可见）。
    recall_ts = _time.time()
    for msg in _chat_messages:
        if msg.get('event_id') == pid:
            msg['_recalled'] = True
            msg['_ts'] = recall_ts
    _threading.Thread(target=_save_chat_log, daemon=True).start()
    # 归档表中的对应消息同步标记（历史回放不再显示）
    try:
        from core.models import ChatArchive as _CA
        with db.engine.begin() as conn:
            rows = conn.execute(_CA.__table__.select().where(_CA.channel == 'chat')).fetchall()
            for r in rows:
                amsgs = r.messages if isinstance(r.messages, list) else []
                a_changed = False
                for msg in amsgs:
                    if msg.get('event_id') == pid and not msg.get('_recalled'):
                        msg['_recalled'] = True
                        a_changed = True
                if a_changed:
                    conn.execute(_CA.__table__.update().where(_CA.id == r.id).values(messages=amsgs))
    except Exception:
        pass

    # 弹窗通知标记撤回（轮询端已过滤 recalled）
    for nt in _event_notifications:
        if nt.get('event_id') == pid:
            nt['recalled'] = True
            nt['pinned'] = False

    return jsonify({'ok': True})


@app.route('/api/dm/events/<int:pid>/pin', methods=['POST'])
def api_dm_event_pin(pid):
    """DM 置顶/取消置顶已发布事件：置顶期间全平台轮询持续返回该事件（公告式展示）。"""
    global _event_notifications
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or request.args.get('name') or '').strip()
    pinned = bool(data.get('pinned'))
    if not name or not _is_dm_request(name, request.remote_addr or 'unknown'):
        return jsonify({'ok': False, 'error': '只有DM可以置顶事件'})

    pub = None
    for ev in _dm_published_events:
        if ev.get('id') == pid:
            pub = ev
            break
    if not pub:
        return jsonify({'ok': False, 'error': '事件不存在'})
    if pub.get('recalled_at'):
        return jsonify({'ok': False, 'error': '已撤回的事件无法置顶'})
    pub['pinned'] = pinned
    _save_dm_published_events()
    for nt in _event_notifications:
        if nt.get('event_id') == pid:
            nt['pinned'] = pinned
    return jsonify({'ok': True, 'pinned': pinned})

# ━━━ 骰点广播到聊天室 ━━━

@app.route('/api/dice-broadcast', methods=['POST'])
def api_dice_broadcast():
    """将掷骰结果广播到聊天室。hidden=true 时只有DM可见。"""
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    text = data.get('text', '').strip()
    hidden = data.get('hidden', False)
    color = data.get('color', '#888888')
    role = data.get('role', 'PL')

    if not name or not text:
        return jsonify({'ok': False, 'error': 'Name and text required'})

    msg = {
        'name': name,
        'text': text,
        'time': _time.strftime('%H:%M:%S'),
        'is_dm': False,
        'color': color,
        'role': role,
        'dice_roll': True,
        'hidden': hidden,
        'ip': 'system',
        '_ts': _time.time(),
    }
    _chat_messages.append(msg)
    _trim_and_archive_chat()
    _threading.Thread(target=_save_chat_log, daemon=True).start()
    return jsonify({'ok': True})


# ━━━ @提及通知 API ━━━

@app.route('/api/mentions', methods=['GET'])
def api_get_mentions():
    """获取当前用户的@提及通知。?name=<用户名>&since=<时间戳>"""
    global _mentions
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})

    since = request.args.get('since', '')
    since_ts = 0.0
    if since:
        try:
            since_ts = float(since)
        except ValueError:
            pass

    # 获取该用户未读的提及
    user_mentions = [m for m in _mentions
                     if m['target_name'] == name and m['_ts'] > since_ts]

    return jsonify({
        'ok': True,
        'mentions': user_mentions,
        'count': len(user_mentions),
    })


@app.route('/api/mentions/clear', methods=['POST'])
def api_clear_mentions():
    """清除某个用户的所有提及通知。"""
    global _mentions
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})

    _mentions = [m for m in _mentions if m['target_name'] != name]
    return jsonify({'ok': True})


@app.route('/api/dm-status', methods=['GET'])
def api_dm_status():
    """返回当前DM信息。按 user_id > name > IP 优先级匹配。"""
    global _dm_name, _dm_ip, _dm_user_id
    client_ip = request.remote_addr or 'unknown'
    name = request.args.get('name', '').strip()
    user_id_str = request.args.get('user_id', '').strip()

    if _dm_name is not None:
        # 优先按 user_id 匹配（改名不丢DM）
        if user_id_str and _dm_user_id is not None:
            try:
                is_dm = (int(user_id_str) == _dm_user_id)
            except (ValueError, TypeError):
                is_dm = (name == _dm_name)
        elif name:
            is_dm = (name == _dm_name)
        else:
            is_dm = _is_dm_ip(client_ip) and (client_ip == _dm_ip)
    elif name or user_id_str:
        # 带身份信息的请求：未注册DM前一律按非DM处理，
        # 身份由加入房间时选择的角色决定（PL不会被误判为DM）
        is_dm = False
    else:
        # 无身份信息（如覆盖层未加入时）：保留本地IP兜底体验
        is_dm = _is_dm_ip(client_ip)

    return jsonify({
        'is_dm': is_dm,
        'dm_name': _dm_name,
        'client_ip': client_ip,
    })


# ━━━ 房间/在线用户 API ━━━

def _prune_stale_users():
    """清理过期在线记录（超过10秒无心跳视为离线），并发送退出消息。"""
    global _dm_name, _dm_user_id, _dm_ip
    _now = _time.time()
    stale_names = [n for n, u in _online_users.items() if _now - u.get('last_heartbeat', 0) > 10]
    for n in stale_names:
        # 注意：DM身份不随离线清除（只清除在线记录），
        # 避免短暂离开（切页面/关标签页）导致身份丢失来回切换；
        # DM身份仅在显式选择PL身份或服务重启时重置
        del _online_users[n]
        # 发送退出消息（与 api_room_leave 保持一致）
        if _dm_name and n != _dm_name:
            _chat_messages.append({
                'name': '系统',
                'text': f'🔴 {n} 退出房间',
                'time': _time.strftime('%H:%M:%S'),
                'is_dm': False,
                'color': '#888888',
                'ip': 'system',
                '_ts': _time.time(),
                'system': True,
            })


@app.route('/api/room/join', methods=['POST'])
def api_room_join():
    """用户加入房间。"""
    global _dm_name, _dm_ip, _dm_user_id, _online_users
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    color = data.get('color', '#00bcd4')
    role = data.get('role', 'PL')
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})

    client_ip = request.remote_addr or 'unknown'
    _prune_stale_users()

    is_new = name not in _online_users

    # ID 唯一性检测：同名且同IP（刷新页面）或旧记录已过期则允许
    if not is_new:
        existing = _online_users.get(name, {})
        existing_ip = existing.get('ip', '')
        if existing_ip and existing_ip == client_ip:
            del _online_users[name]
            is_new = True
        else:
            return jsonify({'ok': False, 'error': f'「{name}」已被使用，请更换ID'})

    # 绑定官网注册用户（可选）
    user_id = data.get('user_id')
    if user_id is not None:
        try:
            user_id = int(user_id)
        except (ValueError, TypeError):
            user_id = None

    # 心跳自动重连（auto=true）：不触碰DM身份、不改变在线角色，
    # 避免多标签页/短暂离线导致的身份波动
    auto = bool(data.get('auto', False))
    if auto and name in _online_users:
        role = _online_users[name].get('role', role)

    # DM 身份判定：优先通过 user_id 识别，其次通过 _dm_name
    # 同一 user_id 改名后仍保持 DM 身份
    if role == 'DM' and not auto:
        if _dm_name is None:
            _dm_name = name
            _dm_ip = client_ip
            _dm_user_id = user_id
        elif _dm_ip is not None and client_ip == _dm_ip:
            # 同一DM（同IP）改名/重新选择DM身份：接管并更新名字，
            # 避免标签页会话名不一致导致被静默降级为PL
            _dm_name = name
            _dm_ip = client_ip
            _dm_user_id = user_id
        elif user_id is not None and _dm_user_id is not None and user_id == _dm_user_id:
            # 同一用户改名：更新 DM 显示名，保持身份
            _dm_name = name
            _dm_ip = client_ip
        elif _dm_name != name:
            role = 'PL'  # 另有其人已是DM
    elif role == 'PL' and not auto and _dm_name == name:
        # 用户显式选择PL身份：放弃DM身份（防止PL被误判为DM）
        _dm_name = None
        _dm_ip = None
        _dm_user_id = None

    _online_users[name] = {
        'ip': client_ip,
        'color': color,
        'role': role,
        'user_id': user_id,          # 关联 users 表主键
        'last_heartbeat': _time.time(),
        'joined_at': _time.time(),
    }

    # 发送系统消息（仅首次加入，重连不重复通知）
    _now = _time.time()
    was_offline = name not in [n for n, u in _online_users.items() if _now - u.get('last_heartbeat', 0) <= 15]
    if was_offline and _dm_name and name != _dm_name:
        _chat_messages.append({
            'name': '系统',
            'text': f'🔵 {name} 进入房间',
            'time': _time.strftime('%H:%M:%S'),
            'is_dm': False,
            'color': '#888888',
            'ip': 'system',
            '_ts': _time.time(),
            'system': True,
        })

    return jsonify({
        'ok': True,
        'is_new': is_new,
        'is_dm': (name == _dm_name),  # 只有实际DM才返回true
        'dm_name': _dm_name,
        'online_count': len(_online_users),
        # 选择了DM但被降级为PL时提示（另有其人已是DM）
        'role_downgraded': (data.get('role', 'PL') == 'DM' and role == 'PL'),
    })


@app.route('/api/room/heartbeat', methods=['POST'])
def api_room_heartbeat():
    """心跳：更新在线状态，返回在线用户列表。用户不在列表中时返回 need_rejoin 标志。"""
    global _online_users, _dm_name
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    color = data.get('color', '#00bcd4')
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})

    # 清理超时用户，确保离线玩家从所有人的在线列表中消失
    _prune_stale_users()

    if name in _online_users:
        _online_users[name]['last_heartbeat'] = _time.time()
        if color:
            _online_users[name]['color'] = color
    else:
        # 不自动重新注册——避免 leaveRoom 后又被心跳拉回来导致重复退出消息
        return jsonify({
            'ok': True,
            'need_rejoin': True,
            'online_users': [
                {'name': n, 'color': u['color'], 'role': u.get('role', 'PL')}
                for n, u in _online_users.items()
            ],
            'online_count': len(_online_users),
        })

    return jsonify({
        'ok': True,
        'online_users': [
            {'name': n, 'color': u['color'], 'role': u.get('role', 'PL')}
            for n, u in _online_users.items()
        ],
        'online_count': len(_online_users),
    })


@app.route('/api/room/leave', methods=['POST'])
def api_room_leave():
    """用户离开房间。"""
    global _online_users, _dm_name, _dm_ip, _dm_user_id
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})

    if name in _online_users:
        # DM身份不随离开清除（只移除在线记录），避免切换页面导致身份丢失
        del _online_users[name]

        if _dm_name and name != _dm_name:
            _chat_messages.append({
                'name': '系统',
                'text': f'🔴 {name} 退出房间',
                'time': _time.strftime('%H:%M:%S'),
                'is_dm': False,
                'color': '#888888',
                'ip': 'system',
                '_ts': _time.time(),
                'system': True,
            })

    return jsonify({'ok': True, 'online_count': len(_online_users)})


# ━━━ 战斗状态同步 API（带磁盘持久化）━━

def _load_combat_state():
    """启动时从 PostgreSQL 恢复战斗状态（原 combat_state.json）。"""
    try:
        from core.models import CombatState as _CS
        row = _CS.query.filter_by(id=1).first()
        if row and isinstance(row.data, dict) and row.data.get('combatants') is not None:
            return row.data, row.ts or 0.0
    except Exception:
        pass
    return {'combatants': [], 'round': ''}, 0.0


def _save_combat_state():
    """保存战斗状态到 PostgreSQL（单行 upsert）。"""
    global _combat_state, _combat_state_ts
    try:
        from core.models import CombatState as _CS
        data = dict(_combat_state) if _combat_state else {'combatants': [], 'round': ''}
        data['_ts'] = _combat_state_ts
        with db.engine.begin() as conn:
            exists = conn.execute(_CS.__table__.select().where(_CS.id == 1)).first()
            if exists:
                conn.execute(_CS.__table__.update().where(_CS.id == 1).values(data=data, ts=_combat_state_ts))
            else:
                conn.execute(_CS.__table__.insert().values(id=1, data=data, ts=_combat_state_ts))
    except Exception as e:
        print(f'[combat-state] 保存失败: {e}')


# 模块加载时恢复战斗状态（ORM 查询需要 Flask app context）
with app.app_context():
    _combat_state, _combat_state_ts = _load_combat_state()
if _combat_state.get('combatants'):
    print(f'[combat-state] 已从磁盘恢复战斗状态 ({len(_combat_state["combatants"])} 名参战者)')


@app.route('/api/combat-state', methods=['GET'])
def api_get_combat_state():
    """获取共享的战斗状态（从内存返回，首次启动从磁盘加载，不再是 None）。"""
    global _combat_state, _combat_state_ts
    return jsonify({'ok': True, 'state': _combat_state, 'timestamp': _combat_state_ts})


@app.route('/api/combat-state', methods=['POST'])
def api_push_combat_state():
    """推送战斗状态（用服务器时间戳标记版本），自动持久化到磁盘。"""
    global _combat_state, _combat_state_ts
    data = request.get_json(silent=True) or {}
    state = data.get('state', {})
    if not isinstance(state, dict):
        state = {}
    # 版本保护：客户端基于旧状态的推送不覆盖新状态（防网络延迟回弹）
    base_ts = float(state.get('_ts', 0) or 0)
    if not state.get('_force') and base_ts < _combat_state_ts - 0.001:
        return jsonify({
            'ok': False, 'conflict': True,
            'state': _combat_state,
            'timestamp': _combat_state_ts,
        }), 409
    _combat_state_ts = _time.time()
    state['_ts'] = _combat_state_ts
    _combat_state = state
    _save_combat_state()
    return jsonify({'ok': True, 'timestamp': _combat_state_ts})


# ━━━ 资源上传 API ━━━

@app.route('/api/resources/upload', methods=['POST'])
def api_upload_resource():
    """上传资源文件到服务器资源库。

    接受 multipart/form-data，字段名 'file'。
    支持图片、PDF、文本、Excel等格式。
    """
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件（字段名: file）'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    # 白名单：只允许特定格式
    ext = _os.path.splitext(file.filename)[1].lower()
    allowed = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp',  # 图片
               '.pdf',                                              # PDF
               '.doc', '.docx',                                     # 文档
               '.xls', '.xlsx',                                     # 表格
               '.txt', '.md',                                       # 文本
               '.json',                                             # 存档
               '.zip'}                                              # 压缩包（工坊投稿）
    if ext not in allowed:
        return jsonify({'error': f'不支持的文件格式: {ext}，仅支持图片/PDF/文档/表格/TXT'}), 400

    # 限制文件大小 (50MB)
    max_size = 50 * 1024 * 1024
    file.seek(0, 2)  # 移动到文件末尾
    size = file.tell()
    file.seek(0)     # 回到开头
    if size > max_size:
        return jsonify({'error': f'文件过大 ({size/1024/1024:.1f}MB)，最大50MB'}), 400

    # 确保目录存在
    RESOURCES_DIR.mkdir(exist_ok=True)

    # 清理文件名：替换 Windows/Linux 非法字符
    import re as _re
    raw_name = file.filename
    # 替换路径分隔符和非法字符为下划线
    safe_name = _re.sub(r'[\\/:*?"<>|]', '_', raw_name)
    # 合并连续下划线，去除首尾下划线和空格
    safe_name = _re.sub(r'_+', '_', safe_name).strip('_ .')
    if not safe_name:
        safe_name = 'unnamed'
    # 保留扩展名（如果原来的扩展名还在）
    if '.' in raw_name:
        raw_ext = _os.path.splitext(raw_name)[1]
        safe_base = _os.path.splitext(safe_name)[0]
        safe_name = safe_base + raw_ext

    dest = RESOURCES_DIR / safe_name
    counter = 1
    name_base, name_ext = _os.path.splitext(safe_name)
    while dest.exists():
        safe_name = f"{name_base}_{counter}{name_ext}"
        dest = RESOURCES_DIR / safe_name
        counter += 1

    try:
        file.save(str(dest))
    except Exception as e:
        return jsonify({'error': f'保存失败: {e}'}), 500

    # 确定文件类别
    info = _RESOURCE_CATEGORIES.get(ext, ('📦 其他', 'other'))
    size_kb = dest.stat().st_size / 1024

    return jsonify({
        'success': True,
        'name': safe_name,
        'original_name': file.filename,
        'size_kb': round(size_kb, 1),
        'size_display': f'{size_kb:.1f}KB' if size_kb < 1024 else f'{size_kb/1024:.1f}MB',
        'ext': ext.lstrip('.'),
        'icon': info[0].split()[0],
        'category': info[1],
        'url': f'/resources/{safe_name}',
    })


# ━━━ 每周话题 API（官网公开）━━━
# 数据文件 data/topics.json: {"topics": [{id, title, content, images[], author, created_at, updated_at}]}
# 内容由管理员在后台发布，content 支持 HTML（含图片标签）

_topics_lock = _threading.Lock()
_comment_rate: dict[str, float] = {}  # 评论防刷: {username: 最近评论时间}
COMMENT_RATE_SECONDS = 15  # 同一用户两次评论最小间隔

# ━━━ 平台公告（管理员发布 → 平台页面顶部横幅）━━━
_announcements_lock = _threading.Lock()


def _load_announcements() -> dict:
    """从 PostgreSQL 加载公告数据（原 data/announcements.json）。"""
    try:
        from core.models import Announcement as _AN
        rows = _AN.query.order_by(_AN.id).all()
        return {'announcements': [{'id': r.id, 'title': r.title, 'content': r.content,
                                   'created_at': r.created_at, 'active': bool(r.active)}
                                  for r in rows]}
    except Exception:
        return {'announcements': []}


def _save_announcements(data: dict):
    """保存公告数据到 PostgreSQL（全删全插保留 id）。"""
    try:
        from core.models import Announcement as _AN
        anns = data.get('announcements', []) if isinstance(data, dict) else []
        with db.engine.begin() as conn:
            conn.execute(_AN.__table__.delete())
            for a in anns:
                conn.execute(_AN.__table__.insert().values(
                    id=a.get('id'), title=a.get('title', ''),
                    content=a.get('content', ''), created_at=a.get('created_at', ''),
                    active=bool(a.get('active', True))))
    except Exception:
        pass


@app.route('/api/announcements')
def api_announcements():
    """公开接口：返回启用中的公告（最新 3 条，供各页面顶部横幅展示）。"""
    with _announcements_lock:
        data = _load_announcements()
    anns = [a for a in data.get('announcements', []) if a.get('active', True)]
    anns.sort(key=lambda a: a.get('id', 0), reverse=True)
    return jsonify({'ok': True, 'announcements': anns[:3]})


@app.route('/api/announcements/history')
def api_announcements_history():
    """公开接口：返回全部启用中的公告（按 id 倒序），供历史公告页展示。"""
    with _announcements_lock:
        data = _load_announcements()
    anns = [a for a in data.get('announcements', []) if a.get('active', True)]
    anns.sort(key=lambda a: a.get('id', 0), reverse=True)
    return jsonify({'ok': True, 'announcements': anns})


@app.route('/announcements')
def announcements_page():
    """官网历史公告页。"""
    return render_template('portal/announcements.html')


# ━━━ 更新日志（历史版本页，读取 骰娘/更新日志/ 目录） ━━━
_CHANGELOG_DIR = _Path(__file__).parent.parent / '更新日志'
_CHANGELOG_RE = _re.compile(r'^(\d{4}-\d{2}-\d{2})_(\d+)_(.+?)\.txt$')


@app.route('/api/changelog')
def api_changelog_list():
    """公开接口：列出 更新日志/ 目录下全部日志（按日期+序号数字倒序）。"""
    entries = []
    try:
        parsed = []
        for f in _CHANGELOG_DIR.glob('*.txt'):
            m = _CHANGELOG_RE.match(f.name)
            if not m:
                continue
            parsed.append((f, m))
        # 按 (日期, 序号) 数字倒序：文件名字符串倒序在序号≥10 时排序错误（_9_ > _10_）
        for f, m in sorted(parsed, key=lambda pm: (pm[1].group(1), int(pm[1].group(2))), reverse=True):
            entries.append({
                'filename': f.name,
                'date': m.group(1),
                'seq': int(m.group(2)),
                'title': m.group(3),
                'size': f.stat().st_size,
            })
    except Exception:
        pass
    return jsonify({'ok': True, 'entries': entries})


@app.route('/api/changelog/<path:filename>')
def api_changelog_content(filename):
    """公开接口：读取单个更新日志内容（纯文本；防路径穿越）。"""
    fname = _os.path.basename(filename)
    if not fname.endswith('.txt') or not _CHANGELOG_RE.match(fname):
        return jsonify({'ok': False, 'error': '文件名不合法'}), 400
    fp = _CHANGELOG_DIR / fname
    try:
        content = fp.read_text(encoding='utf-8', errors='replace')
    except FileNotFoundError:
        return jsonify({'ok': False, 'error': '文件不存在'}), 404
    except Exception:
        return jsonify({'ok': False, 'error': '读取失败'}), 500
    return jsonify({'ok': True, 'filename': fname, 'content': content})


@app.route('/changelog')
def changelog_page():
    """官网历史版本更新日志页。"""
    return render_template('portal/changelog.html')


@app.route('/guide')
def guide_page():
    """官网新手使用指南页（静态内容）。"""
    return render_template('portal/guide.html')


@app.route('/post/<int:post_id>')
def post_detail_page(post_id):
    """酒馆帖子独立详情页（收藏/列表跳转用）。"""
    return render_template('portal/post_detail.html', post_id=post_id)


@app.route('/workshop/<int:item_id>')
def workshop_detail_page(item_id):
    """工坊投稿独立详情页（收藏/列表跳转用）。"""
    return render_template('portal/workshop_detail.html', item_id=item_id)


@app.route('/notifications')
def notifications_page():
    """消息通知独立页（被评论/被点赞/@提及提醒）。"""
    return render_template('portal/notifications.html')


@app.route('/api/admin/announcements', methods=['GET'])
def api_admin_announcements_list():
    """管理员读取全部公告（含已停用），按 id 倒序。"""
    user = _get_current_user()
    if not user or not (user.is_admin if user.is_admin is not None else False):
        return jsonify({'ok': False, 'error': '需要管理员权限'}), 403
    with _announcements_lock:
        data = _load_announcements()
    anns = sorted(data.get('announcements', []), key=lambda a: a.get('id', 0), reverse=True)
    return jsonify({'ok': True, 'announcements': anns})


@app.route('/api/admin/announcements', methods=['POST'])
def api_admin_announcements():
    """管理员发布/更新/启停/删除公告。
    请求: {action: 'add'|'toggle'|'delete', title, content, id}
    """
    user = _get_current_user()
    if not user or not (user.is_admin if user.is_admin is not None else False):
        return jsonify({'ok': False, 'error': '需要管理员权限'}), 403
    data = request.get_json(silent=True) or {}
    action = data.get('action', '')
    with _announcements_lock:
        store = _load_announcements()
        anns = store.setdefault('announcements', [])
        if action == 'add':
            title = (data.get('title') or '').strip()
            content = (data.get('content') or '').strip()
            if not title or not content:
                return jsonify({'ok': False, 'error': '标题与内容不能为空'})
            if len(title) > 60 or len(content) > 500:
                return jsonify({'ok': False, 'error': '标题最多60字，内容最多500字'})
            new_id = max([a.get('id', 0) for a in anns], default=0) + 1
            anns.append({
                'id': new_id,
                'title': title,
                'content': content,
                'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
                'active': True,
            })
            _save_announcements(store)
            return jsonify({'ok': True, 'announcement': anns[-1]})
        if action == 'toggle':
            ann_id = data.get('id')
            for a in anns:
                if a.get('id') == ann_id:
                    a['active'] = not a.get('active', True)
                    _save_announcements(store)
                    return jsonify({'ok': True, 'active': a['active']})
            return jsonify({'ok': False, 'error': '公告不存在'})
        if action == 'delete':
            ann_id = data.get('id')
            before = len(anns)
            anns[:] = [a for a in anns if a.get('id') != ann_id]
            if len(anns) != before:
                _save_announcements(store)
                return jsonify({'ok': True})
            return jsonify({'ok': False, 'error': '公告不存在'})
    return jsonify({'ok': False, 'error': '未知操作'})


def _load_topics() -> dict:
    """从 PostgreSQL 加载每周话题（原 data/topics.json）。"""
    try:
        from core.models import Topic as _TP
        rows = _TP.query.order_by(_TP.id).all()
        return {'topics': [{'id': r.id, 'title': r.title, 'content': r.content,
                            'images': r.images or [], 'author': r.author or '',
                            'created_at': r.created_at or '', 'updated_at': r.updated_at or '',
                            'comments': r.comments or []} for r in rows]}
    except Exception:
        return {'topics': []}


def _save_topics(data: dict):
    """保存每周话题到 PostgreSQL（全删全插保留 id）。"""
    try:
        from core.models import Topic as _TP
        topics = data.get('topics', []) if isinstance(data, dict) else []
        with db.engine.begin() as conn:
            conn.execute(_TP.__table__.delete())
            for t in topics:
                conn.execute(_TP.__table__.insert().values(
                    id=t.get('id'), title=t.get('title', ''), content=t.get('content', ''),
                    images=t.get('images') or [], author=t.get('author', ''),
                    created_at=t.get('created_at', ''), updated_at=t.get('updated_at', ''),
                    comments=t.get('comments') or []))
    except Exception:
        pass


@app.route('/api/stats/overview')
def api_stats_overview():
    """平台统计（公开）：主题/帖子/用户/工坊投稿/投掷次数。"""
    result = {'topics': 0, 'posts': 0, 'users': 0, 'workshop': 0, 'dice': 0}
    try:
        with _community_lock:
            cposts = _load_json_file(_COMMUNITY_FILE, [])
            witems = _load_json_file(_WORKSHOP_FILE, [])
        result['topics'] = len(cposts)
        result['posts'] = len(cposts)  # 帖子总数 = 酒馆帖子数（评论不计入）
        result['workshop'] = len(witems)
    except Exception:
        pass
    try:
        result['users'] = _UserModel.query.count()
    except Exception:
        pass
    try:
        with _stats_lock:
            dice_stats = _load_stats()
        result['dice'] = sum(s.get('total', 0) for s in dice_stats.values())
    except Exception:
        pass
    return jsonify({'ok': True, **result})


@app.route('/api/topics')
def api_topics_list():
    """每周话题列表（公开，按创建时间倒序）。"""
    with _topics_lock:
        data = _load_topics()
    topics = []
    for t in data.get('topics', []):
        text = _re.sub(r'<[^>]+>', '', t.get('content', ''))  # 去 HTML 标签
        topics.append({
            'id': t.get('id', 0),
            'title': t.get('title', ''),
            'summary': text.strip()[:80],
            'author': t.get('author', ''),
            'created_at': t.get('created_at', ''),
            'updated_at': t.get('updated_at', ''),
            'image_count': len(t.get('images', [])),
            'cover': (t.get('images') or [''])[0],
        })
    return jsonify({'ok': True, 'topics': topics})


@app.route('/topics/<int:topic_id>')
def topic_detail_page(topic_id):
    """每周话题独立详情页（含评论区）。"""
    with _topics_lock:
        data = _load_topics()
    topic = next((t for t in data.get('topics', []) if t.get('id') == topic_id), None)
    if not topic:
        return jsonify({'error': '话题不存在'}), 404
    return render_template('topic_detail.html', topic_id=topic_id)


@app.route('/api/topics/<int:topic_id>')
def api_topic_detail(topic_id):
    """每周话题详情（公开，含评论）。"""
    with _topics_lock:
        data = _load_topics()
    for t in data.get('topics', []):
        if t.get('id') == topic_id:
            t.setdefault('comments', [])
            return jsonify({'ok': True, 'topic': t})
    return jsonify({'ok': False, 'error': '话题不存在'}), 404


@app.route('/api/topics/<int:topic_id>/comments', methods=['POST'])
def api_topic_comment(topic_id):
    """发表评论（需登录）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    data = request.get_json(silent=True) or {}
    content = (data.get('content') or '').strip()
    if not content:
        return jsonify({'ok': False, 'error': '评论内容不能为空'}), 400
    if len(content) > 500:
        return jsonify({'ok': False, 'error': '评论最多 500 字'}), 400

    with _topics_lock:
        # 防刷：同一用户 15 秒内只能评论一次
        now = _time.time()
        last = _comment_rate.get(user.username, 0)
        if now - last < COMMENT_RATE_SECONDS:
            return jsonify({'ok': False, 'error': '评论太频繁，请稍后再试'}), 429
        _comment_rate[user.username] = now

        data_topics = _load_topics()
        topic = next((t for t in data_topics.get('topics', []) if t.get('id') == topic_id), None)
        if not topic:
            return jsonify({'ok': False, 'error': '话题不存在'}), 404

        comments = topic.setdefault('comments', [])
        comment = {
            'id': max([c.get('id', 0) for c in comments], default=0) + 1,
            'username': user.username,
            'content': content,
            'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        comments.append(comment)
        _save_topics(data_topics)

    return jsonify({'ok': True, 'comment': comment})


@app.route('/api/topics/<int:topic_id>/comments/<int:comment_id>', methods=['DELETE'])
def api_topic_comment_delete(topic_id, comment_id):
    """删除评论（仅作者本人或管理员）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    with _topics_lock:
        data_topics = _load_topics()
        topic = next((t for t in data_topics.get('topics', []) if t.get('id') == topic_id), None)
        if not topic:
            return jsonify({'ok': False, 'error': '话题不存在'}), 404

        comments = topic.get('comments', [])
        comment = next((c for c in comments if c.get('id') == comment_id), None)
        if not comment:
            return jsonify({'ok': False, 'error': '评论不存在'}), 404

        is_admin_user = bool(user.is_admin if user.is_admin is not None else False)
        if comment.get('username') != user.username and not is_admin_user:
            return jsonify({'ok': False, 'error': '只能删除自己的评论'}), 403

        topic['comments'] = [c for c in comments if c.get('id') != comment_id]
        _save_topics(data_topics)

    return jsonify({'ok': True})


# ━━━ 冒险者酒馆 + 创意工坊 API（真实存储）━━━
# 数据文件：data/community.json（帖子）、data/workshop.json（工坊投稿）
# 板块定义为静态结构，主题数/帖子数为真实统计

_COMMUNITY_FILE = _Path(__file__).parent.parent / 'data' / 'community.json'
_WORKSHOP_FILE = _Path(__file__).parent.parent / 'data' / 'workshop.json'
_community_lock = _threading.Lock()

# 板块静态定义
_COMMUNITY_BOARDS = [
    {'id': 'discuss', 'icon': '💭', 'name': '综合讨论区', 'desc': '跑团相关的任何话题'},
    {'id': 'rules', 'icon': '📖', 'name': '规则研讨堂', 'desc': 'D&D 5E 规则深度讨论'},
    {'id': 'recruit', 'icon': '📯', 'name': '跑团招募版', 'desc': '寻找 DM 或玩家'},
    {'id': 'battle', 'icon': '📜', 'name': '冒险战报馆', 'desc': '记录你的每一次冒险'},
    {'id': 'tech', 'icon': '🛠️', 'name': '技术支持区', 'desc': '平台使用问题、Bug反馈'},
]


def _load_json_file(path: _Path, default: list) -> list:
    """加载酒馆帖子/工坊投稿（2026-08-15 起从 PostgreSQL 读取）。

    path 参数保留兼容（_COMMUNITY_FILE/_WORKSHOP_FILE 识别来源）。
    """
    try:
        if path == _COMMUNITY_FILE:
            from core.models import CommunityPost as _CP
            rows = _CP.query.order_by(_CP.id).all()
            return [{'id': r.id, 'board': r.board, 'title': r.title, 'content': r.content,
                     'images': r.images or [], 'author': r.author or '',
                     'created_at': r.created_at or '', 'reply_count': r.reply_count or 0,
                     'comments': r.comments or [], 'likes': r.likes or []} for r in rows]
        if path == _WORKSHOP_FILE:
            from core.models import WorkshopItem as _WI
            rows = _WI.query.order_by(_WI.id).all()
            return [{'id': r.id, 'title': r.title, 'desc': r.desc, 'file_url': r.file_url or '',
                     'author': r.author or '', 'created_at': r.created_at or '',
                     'cat': r.cat or 'user', 'comments': r.comments or [], 'likes': r.likes or []}
                    for r in rows]
    except Exception:
        pass
    return default


def _save_json_file(path: _Path, data: list):
    """保存酒馆帖子/工坊投稿到 PostgreSQL（全删全插保留 id）。"""
    try:
        if path == _COMMUNITY_FILE:
            from core.models import CommunityPost as _CP
            with db.engine.begin() as conn:
                conn.execute(_CP.__table__.delete())
                for p in data:
                    conn.execute(_CP.__table__.insert().values(
                        id=p.get('id'), board=p.get('board', 'discuss'),
                        title=p.get('title', ''), content=p.get('content', ''),
                        images=p.get('images') or [], author=p.get('author', ''),
                        created_at=p.get('created_at', ''), reply_count=p.get('reply_count', 0),
                        comments=p.get('comments') or [], likes=p.get('likes') or []))
        elif path == _WORKSHOP_FILE:
            from core.models import WorkshopItem as _WI
            with db.engine.begin() as conn:
                conn.execute(_WI.__table__.delete())
                for s in data:
                    conn.execute(_WI.__table__.insert().values(
                        id=s.get('id'), title=s.get('title', ''), desc=s.get('desc', ''),
                        file_url=s.get('file_url', ''), author=s.get('author', ''),
                        created_at=s.get('created_at', ''), cat=s.get('cat', 'user'),
                        comments=s.get('comments') or [], likes=s.get('likes') or []))
    except Exception:
        pass


@app.route('/api/community/boards')
def api_community_boards():
    """板块列表 + 真实帖子统计。"""
    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
    boards = []
    for b in _COMMUNITY_BOARDS:
        b_posts = [p for p in posts if p.get('board') == b['id']]
        replies = sum(p.get('reply_count', 0) for p in b_posts)
        boards.append({
            **b,
            'topics': len(b_posts),
            'posts': len(b_posts) + replies,
            'last': b_posts[0].get('created_at', '') if b_posts else '',
            'lastBy': b_posts[0].get('author', '') if b_posts else '',
            'lastTopic': b_posts[0].get('title', '') if b_posts else '',
        })
    return jsonify({'ok': True, 'boards': boards})


def _fav_count_map() -> dict:
    """favorites.json → {(type, item_id): 收藏数}，供帖子/工坊列表统计。"""
    with _favorites_lock:
        store = _load_favorites()
    all_favs = store.get('favorites', [])
    cnt: dict = {}
    for fv in all_favs:
        key = (fv.get('type'), fv.get('item_id'))
        cnt[key] = cnt.get(key, 0) + 1
    return cnt


def _enrich_post(p: dict, fav_map: dict, user) -> dict:
    """帖子附加 like_count / fav_count / liked / author_avatar。"""
    likes = p.get('likes') or []
    if not isinstance(likes, list):
        likes = []
    return {
        **p,
        'like_count': len(likes),
        'fav_count': fav_map.get(('post', p.get('id')), 0),
        'liked': (user.username in likes) if user else False,
        'author_avatar': _user_avatar(p.get('author')),
    }


def _enrich_workshop_item(it: dict, fav_map: dict, user) -> dict:
    """工坊投稿附加 like_count / fav_count / liked / author_avatar。"""
    likes = it.get('likes') or []
    if not isinstance(likes, list):
        likes = []
    return {
        **it,
        'like_count': len(likes),
        'fav_count': fav_map.get(('workshop', it.get('id')), 0),
        'liked': (user.username in likes) if user else False,
        'author_avatar': _user_avatar(it.get('author')),
    }


def _user_avatar(username: str) -> str:
    """按用户名查头像 URL（查不到返回空串，前端显示默认占位）。"""
    if not username:
        return ''
    u = _UserModel.query.filter_by(username=username).first()
    return (u.avatar_url or '') if u else ''


@app.route('/api/community/boards/<board_id>/threads')
def api_community_threads(board_id):
    """板块帖子列表（按发布时间倒序）。

    支持搜索与分页：?q=关键词&page=1&page_size=20
    q 匹配标题/内容/作者（忽略大小写）；不传 page 时返回全部（兼容旧调用）。
    每帖附带 like_count / fav_count / liked / author_avatar。
    """
    if board_id not in {b['id'] for b in _COMMUNITY_BOARDS}:
        return jsonify({'ok': False, 'error': '板块不存在'}), 404
    q = (request.args.get('q') or '').strip().lower()
    page = request.args.get('page', type=int)
    page_size = min(request.args.get('page_size', type=int) or 20, 100)
    user = _get_current_user()
    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
    threads = [p for p in posts if p.get('board') == board_id]
    if q:
        threads = [p for p in threads
                   if q in (p.get('title') or '').lower()
                   or q in (p.get('content') or '').lower()
                   or q in (p.get('author') or '').lower()]
    total = len(threads)
    has_more = False
    if page is not None:
        page = max(page, 1)
        page_size = max(page_size, 1)
        start = (page - 1) * page_size
        threads = threads[start:start + page_size]
        has_more = start + len(threads) < total
    fav_map = _fav_count_map()
    out = [_enrich_post(p, fav_map, user) for p in threads]
    return jsonify({'ok': True, 'threads': out, 'total': total, 'page': page, 'has_more': has_more})


def _notify(user_id, ntype: str, actor: str, content: str, link: str, group_key: str = ''):
    """写入消息通知（接收者/操作者/动作文本/跳转链接/聚合键）。失败静默不影响主流程。

    content 为动作文本（不含 actor）；同 group_key 未读通知自动聚合（count+1，
    展示时 "N 人"+content）。兼容旧调用（无 group_key）。
    """
    try:
        from core.models import Notification as _NT
        now = _time.strftime('%Y-%m-%d %H:%M:%S')
        if group_key:
            existing = _NT.query.filter_by(user_id=user_id, group_key=group_key, read=False).first()
            if existing:
                existing.count = (existing.count or 1) + 1
                existing.actor = actor
                existing.created_at = now
                db.session.commit()
                return
        db.session.add(_NT(user_id=user_id, type=ntype, actor=actor, content=content,
                           link=link, read=False, group_key=group_key, count=1,
                           created_at=now))
        db.session.commit()
    except Exception as _e:
        _app_logger.error(f'通知写入失败 user_id={user_id} type={ntype}: {_e}')


def _notify_author(author_name: str, ntype: str, actor: str, content: str, link: str,
                   group_key: str = ''):
    """按作者用户名通知（作者无官网账号或为自己操作时跳过）。"""
    if not author_name:
        return
    author = _UserModel.query.filter_by(username=author_name).first()
    if author and author.username != actor:
        _notify(author.id, ntype, actor, content, link, group_key)
    elif not author:
        _app_logger.warning(f'通知未发送：找不到用户 [{author_name}]（帖子/投稿作者名与账号用户名不一致，可能改过名）')


_MENTION_RE = _re.compile(r'@([\w一-龥-]{2,30})')


def _notify_mentions(text: str, actor: str, parent_title: str, link: str):
    """解析评论内容中的 @用户名 定向通知（排除自己；被提及者无账号跳过）。"""
    try:
        seen = set()
        for m in _MENTION_RE.finditer(text or ''):
            uname = m.group(1)
            if uname in seen:
                continue
            seen.add(uname)
            target = _UserModel.query.filter_by(username=uname).first()
            if target and target.username != actor:
                _notify(target.id, 'mention', actor,
                        f'在评论中提及了你：《{parent_title[:30]}》',
                        link, group_key=f'mention:{link}')
    except Exception:
        pass


@app.route('/api/notifications')
def api_my_notifications():
    """我的消息通知列表（登录，最新 50 条）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    from core.models import Notification as _NT
    rows = _NT.query.filter_by(user_id=user.id).order_by(_NT.id.desc()).limit(50).all()
    return jsonify({'ok': True, 'notifications': [{
        'id': r.id, 'type': r.type, 'actor': r.actor, 'content': r.content,
        'link': r.link, 'read': bool(r.read), 'count': r.count or 1,
        'created_at': r.created_at or '',
    } for r in rows]})


@app.route('/api/notifications/unread-count')
def api_notifications_unread_count():
    """未读通知数（用户中心红点）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': True, 'count': 0})
    from core.models import Notification as _NT
    count = _NT.query.filter_by(user_id=user.id, read=False).count()
    return jsonify({'ok': True, 'count': count})


@app.route('/api/notifications/read', methods=['POST'])
def api_notifications_mark_read():
    """标记已读：{id} 单条或 {all: true} 全部。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    from core.models import Notification as _NT
    data = request.get_json(silent=True) or {}
    if data.get('all'):
        _NT.query.filter_by(user_id=user.id, read=False).update({'read': True})
        db.session.commit()
        return jsonify({'ok': True})
    nid = data.get('id')
    if nid is not None:
        _NT.query.filter_by(id=nid, user_id=user.id).update({'read': True})
        db.session.commit()
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': '参数错误'}), 400


# ━━━ 好友系统（2026-08-18，双向好友：申请→接受→成为好友）━━━

@app.route('/api/users/search')
def api_user_search():
    """按用户名或用户 ID 搜索用户（好友添加前置，登录）。返回与自己的好友关系状态。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'ok': False, 'error': '请输入用户名或用户ID'}), 400
    from core.models import FriendRequest as _FR
    targets = []
    if q.isdigit():
        t = _UserModel.query.filter_by(id=int(q)).first()
        if t:
            targets = [t]
    else:
        exact = _UserModel.query.filter_by(username=q).first()
        if exact:
            targets = [exact]
        else:
            targets = _UserModel.query.filter(_UserModel.username.ilike('%' + q + '%')).limit(10).all()
    now = _time.strftime('%Y-%m-%d %H:%M:%S')
    users = []
    for t in targets:
        if t.id == user.id:
            continue  # 不返回自己
        pair = ((_FR.user_id == user.id) & (_FR.target_id == t.id)) | \
               ((_FR.user_id == t.id) & (_FR.target_id == user.id))
        if _FR.query.filter(pair, _FR.status == 'accepted').first():
            relation = 'friend'
        elif _FR.query.filter_by(user_id=user.id, target_id=t.id, status='pending').first():
            relation = 'pending_out'
        elif _FR.query.filter_by(user_id=t.id, target_id=user.id, status='pending').first():
            relation = 'pending_in'
        else:
            relation = 'none'
        users.append({'id': t.id, 'username': t.username,
                      'avatar_url': t.avatar_url or '', 'relation': relation})
    return jsonify({'ok': True, 'users': users})


@app.route('/api/friends/request', methods=['POST'])
def api_friend_request():
    """发送好友申请（按用户名）。反向 pending 自动转接受互加。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    data = request.get_json(silent=True) or {}
    target_name = (data.get('target_username') or '').strip()
    if not target_name:
        return jsonify({'ok': False, 'error': '请输入对方用户名'}), 400
    target = _UserModel.query.filter_by(username=target_name).first()
    if not target:
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    if target.id == user.id:
        return jsonify({'ok': False, 'error': '不能添加自己为好友'}), 400
    from core.models import FriendRequest as _FR
    now = _time.strftime('%Y-%m-%d %H:%M:%S')
    pair = ((_FR.user_id == user.id) & (_FR.target_id == target.id)) | \
           ((_FR.user_id == target.id) & (_FR.target_id == user.id))
    accepted = _FR.query.filter(pair, _FR.status == 'accepted').first()
    if accepted:
        return jsonify({'ok': False, 'error': '你们已经是好友了'}), 400
    my_pending = _FR.query.filter_by(user_id=user.id, target_id=target.id,
                                     status='pending').first()
    if my_pending:
        return jsonify({'ok': False, 'error': '好友申请已发送，等待对方处理'}), 400
    # 对方已向我发过 pending：直接互加
    reverse = _FR.query.filter_by(user_id=target.id, target_id=user.id,
                                  status='pending').first()
    if reverse:
        reverse.status = 'accepted'
        reverse.responded_at = now
        db.session.commit()
        _notify(reverse.user_id, 'friend_accept', user.username,
                '接受了你的好友请求', '/user')
        return jsonify({'ok': True, 'accepted': True, 'message': '对方已向你发送过申请，你们已互相成为好友'})
    db.session.add(_FR(user_id=user.id, target_id=target.id, status='pending',
                       created_at=now, responded_at=''))
    db.session.commit()
    _notify(target.id, 'friend_request', user.username,
            '请求加你为好友', '/user')
    return jsonify({'ok': True, 'accepted': False, 'message': '好友申请已发送'})


@app.route('/api/friends/respond', methods=['POST'])
def api_friend_respond():
    """处理好友申请：{request_id, action: accept|reject}。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    data = request.get_json(silent=True) or {}
    from core.models import FriendRequest as _FR
    fr = _FR.query.filter_by(id=data.get('request_id'), target_id=user.id).first()
    if not fr or fr.status != 'pending':
        return jsonify({'ok': False, 'error': '申请不存在或已处理'}), 404
    action = data.get('action')
    if action not in ('accept', 'reject'):
        return jsonify({'ok': False, 'error': '参数错误'}), 400
    now = _time.strftime('%Y-%m-%d %H:%M:%S')
    fr.status = 'accepted' if action == 'accept' else 'rejected'
    fr.responded_at = now
    db.session.commit()
    if action == 'accept':
        _notify(fr.user_id, 'friend_accept', user.username,
                '接受了你的好友请求', '/user')
    return jsonify({'ok': True})


def _friend_to_dict(other_id: int):
    """好友信息（含用户名与头像，供列表展示）。"""
    other = db.session.get(_UserModel, other_id)
    if not other:
        return None
    return {'id': other.id, 'username': other.username,
            'avatar_url': other.avatar_url or ''}


@app.route('/api/friends')
def api_friend_list():
    """我的好友列表（accepted 双向）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    from core.models import FriendRequest as _FR
    rows = _FR.query.filter(
        ((_FR.user_id == user.id) | (_FR.target_id == user.id)),
        _FR.status == 'accepted',
    ).all()
    friends = []
    for r in rows:
        other_id = r.target_id if r.user_id == user.id else r.user_id
        info = _friend_to_dict(other_id)
        if info:
            friends.append(info)
    return jsonify({'ok': True, 'friends': friends})


@app.route('/api/friends/pending')
def api_friend_pending():
    """待处理请求：收到的（可接受/拒绝）+ 已发出的。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    from core.models import FriendRequest as _FR
    incoming = _FR.query.filter_by(target_id=user.id, status='pending').all()
    outgoing = _FR.query.filter_by(user_id=user.id, status='pending').all()
    received = []
    for r in incoming:
        info = _friend_to_dict(r.user_id)
        if info:
            info['request_id'] = r.id
            info['created_at'] = r.created_at or ''
            received.append(info)
    sent = []
    for r in outgoing:
        info = _friend_to_dict(r.target_id)
        if info:
            info['request_id'] = r.id
            sent.append(info)
    return jsonify({'ok': True, 'received': received, 'sent': sent})


@app.route('/api/friends/<int:other_id>', methods=['DELETE'])
def api_friend_delete(other_id):
    """删除好友（删除双方 accepted 记录）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    from core.models import FriendRequest as _FR
    _FR.query.filter(
        ((_FR.user_id == user.id) & (_FR.target_id == other_id)) |
        ((_FR.user_id == other_id) & (_FR.target_id == user.id)),
        _FR.status == 'accepted',
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/community/posts/<int:post_id>/like', methods=['POST'])
def api_post_like(post_id):
    """点赞/取消点赞帖子（需登录）。点赞人存 posts 对象 likes 数组。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
        post = next((p for p in posts if p.get('id') == post_id), None)
        if not post:
            return jsonify({'ok': False, 'error': '帖子不存在'}), 404
        likes = post.get('likes')
        if not isinstance(likes, list):
            likes = []
            post['likes'] = likes
        if user.username in likes:
            likes.remove(user.username)
            liked = False
        else:
            likes.append(user.username)
            liked = True
            # 被点赞通知（自己点赞自己不通知；同帖聚合）
            _notify_author(post.get('author'), 'post_like', user.username,
                           f'点赞了你的帖子《{post.get("title", "")[:30]}》',
                           f'/post/{post_id}', group_key=f'post_like:{post_id}')
        _save_json_file(_COMMUNITY_FILE, posts)
    return jsonify({'ok': True, 'liked': liked, 'like_count': len(likes)})


@app.route('/api/community/posts/<int:post_id>')
def api_community_post_detail(post_id):
    """帖子详情（公开）。附带点赞数/收藏数/当前用户点赞态/作者头像。"""
    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
    post = next((p for p in posts if p.get('id') == post_id), None)
    if not post:
        return jsonify({'ok': False, 'error': '帖子不存在'}), 404
    board_name = post.get('board')
    for b in _COMMUNITY_BOARDS:
        if b['id'] == post.get('board'):
            board_name = b['name']
            break
    user = _get_current_user()
    fav_map = _fav_count_map()
    enriched = _enrich_post(post, fav_map, user)
    return jsonify({'ok': True, 'post': {**enriched, 'boardName': board_name}})


@app.route('/api/community/posts', methods=['POST'])
def api_community_post():
    """发布帖子（需登录）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    data = request.get_json(silent=True) or {}
    board = (data.get('board') or '').strip()
    title = (data.get('title') or '').strip()
    content = (data.get('content') or '').strip()
    images = data.get('images') or []
    if not isinstance(images, list):
        images = []
    images = [str(i).strip()[:300] for i in images if str(i).strip()][:9]

    if board not in {b['id'] for b in _COMMUNITY_BOARDS}:
        return jsonify({'ok': False, 'error': '请选择板块'}), 400
    if not title:
        return jsonify({'ok': False, 'error': '请填写帖子标题'}), 400
    if len(title) > 100:
        return jsonify({'ok': False, 'error': '标题最多 100 字'}), 400
    if not content:
        return jsonify({'ok': False, 'error': '请填写帖子内容'}), 400
    if len(content) > 5000:
        return jsonify({'ok': False, 'error': '内容最多 5000 字'}), 400

    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
        post = {
            'id': max([p.get('id', 0) for p in posts], default=0) + 1,
            'board': board,
            'title': title,
            'content': content,
            'images': images,
            'author': user.username,
            'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
            'reply_count': 0,
        }
        posts.insert(0, post)
        _save_json_file(_COMMUNITY_FILE, posts)

    return jsonify({'ok': True, 'post': post})


@app.route('/api/community/posts/<int:post_id>', methods=['DELETE'])
def api_community_post_delete(post_id):
    """删除帖子（仅作者本人或管理员）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
        post = next((p for p in posts if p.get('id') == post_id), None)
        if not post:
            return jsonify({'ok': False, 'error': '帖子不存在'}), 404
        is_admin_user = bool(user.is_admin if user.is_admin is not None else False)
        if post.get('author') != user.username and not is_admin_user:
            return jsonify({'ok': False, 'error': '只能删除自己的帖子'}), 403
        posts = [p for p in posts if p.get('id') != post_id]
        _save_json_file(_COMMUNITY_FILE, posts)

    return jsonify({'ok': True})


@app.route('/api/workshop/items')
def api_workshop_items():
    """工坊列表（真实投稿）。

    支持搜索与分页：?q=关键词&page=1&page_size=12
    q 匹配标题/介绍/作者（忽略大小写）；不传 page 时返回全部（兼容旧调用）。
    每项附带 like_count / fav_count / liked / author_avatar。
    """
    q = (request.args.get('q') or '').strip().lower()
    page = request.args.get('page', type=int)
    page_size = min(request.args.get('page_size', type=int) or 12, 100)
    user = _get_current_user()
    with _community_lock:
        subs = _load_json_file(_WORKSHOP_FILE, [])
    if q:
        subs = [s for s in subs
                if q in (s.get('title') or '').lower()
                or q in (s.get('desc') or '').lower()
                or q in (s.get('author') or '').lower()]
    total = len(subs)
    has_more = False
    if page is not None:
        page = max(page, 1)
        page_size = max(page_size, 1)
        start = (page - 1) * page_size
        subs = subs[start:start + page_size]
        has_more = start + len(subs) < total
    fav_map = _fav_count_map()
    out = [_enrich_workshop_item(s, fav_map, user) for s in subs]
    return jsonify({'ok': True, 'items': out, 'total': total, 'page': page, 'has_more': has_more})


def _make_comment(username: str, content: str, comments: list) -> dict:
    """构造评论对象（id 递增）。"""
    return {
        'id': max([c.get('id', 0) for c in comments], default=0) + 1,
        'username': username,
        'content': content,
        'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
    }


def _comment_rate_check(username: str) -> str | None:
    """评论防刷（同一用户 15 秒间隔），返回错误信息或 None。"""
    now = _time.time()
    last = _comment_rate.get(username, 0)
    if now - last < COMMENT_RATE_SECONDS:
        return '评论太频繁，请稍后再试'
    _comment_rate[username] = now
    return None


@app.route('/api/community/posts/<int:post_id>/comments', methods=['POST'])
def api_community_post_comment(post_id):
    """帖子评论（需登录）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    data = request.get_json(silent=True) or {}
    content = (data.get('content') or '').strip()
    if not content:
        return jsonify({'ok': False, 'error': '评论内容不能为空'}), 400
    if len(content) > 500:
        return jsonify({'ok': False, 'error': '评论最多 500 字'}), 400

    rate_err = _comment_rate_check(user.username)
    if rate_err:
        return jsonify({'ok': False, 'error': rate_err}), 429

    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
        post = next((p for p in posts if p.get('id') == post_id), None)
        if not post:
            return jsonify({'ok': False, 'error': '帖子不存在'}), 404
        comments = post.setdefault('comments', [])
        comment = _make_comment(user.username, content, comments)
        comments.append(comment)
        _save_json_file(_COMMUNITY_FILE, posts)
        # 被评论通知（评论自己的帖子不通知；同帖聚合）
        _notify_author(post.get('author'), 'post_comment', user.username,
                       f'评论了你的帖子《{post.get("title", "")[:30]}》',
                       f'/post/{post_id}', group_key=f'post_comment:{post_id}')
        # @提及提醒
        _notify_mentions(content, user.username, post.get('title', ''), f'/post/{post_id}')

    return jsonify({'ok': True, 'comment': comment})


@app.route('/api/community/posts/<int:post_id>/comments/<int:comment_id>', methods=['DELETE'])
def api_community_post_comment_delete(post_id, comment_id):
    """删除帖子评论（作者本人或管理员）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    with _community_lock:
        posts = _load_json_file(_COMMUNITY_FILE, [])
        post = next((p for p in posts if p.get('id') == post_id), None)
        if not post:
            return jsonify({'ok': False, 'error': '帖子不存在'}), 404
        comments = post.get('comments', [])
        comment = next((c for c in comments if c.get('id') == comment_id), None)
        if not comment:
            return jsonify({'ok': False, 'error': '评论不存在'}), 404
        is_admin_user = bool(user.is_admin if user.is_admin is not None else False)
        if comment.get('username') != user.username and not is_admin_user:
            return jsonify({'ok': False, 'error': '只能删除自己的评论'}), 403
        post['comments'] = [c for c in comments if c.get('id') != comment_id]
        _save_json_file(_COMMUNITY_FILE, posts)

    return jsonify({'ok': True})


@app.route('/api/workshop/items/<int:item_id>/comments', methods=['POST'])
def api_workshop_item_comment(item_id):
    """工坊投稿评论（需登录）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    data = request.get_json(silent=True) or {}
    content = (data.get('content') or '').strip()
    if not content:
        return jsonify({'ok': False, 'error': '评论内容不能为空'}), 400
    if len(content) > 500:
        return jsonify({'ok': False, 'error': '评论最多 500 字'}), 400

    rate_err = _comment_rate_check(user.username)
    if rate_err:
        return jsonify({'ok': False, 'error': rate_err}), 429

    with _community_lock:
        items = _load_json_file(_WORKSHOP_FILE, [])
        item = next((s for s in items if s.get('id') == item_id), None)
        if not item:
            return jsonify({'ok': False, 'error': '投稿不存在'}), 404
        comments = item.setdefault('comments', [])
        comment = _make_comment(user.username, content, comments)
        comments.append(comment)
        _save_json_file(_WORKSHOP_FILE, items)
        # 被评论通知（评论自己的投稿不通知；同投稿聚合）
        _notify_author(item.get('author'), 'workshop_comment', user.username,
                       f'评论了你的投稿《{item.get("title", "")[:30]}》',
                       f'/workshop/{item_id}', group_key=f'workshop_comment:{item_id}')
        # @提及提醒
        _notify_mentions(content, user.username, item.get('title', ''), f'/workshop/{item_id}')

    return jsonify({'ok': True, 'comment': comment})


@app.route('/api/workshop/items/<int:item_id>/comments/<int:comment_id>', methods=['DELETE'])
def api_workshop_item_comment_delete(item_id, comment_id):
    """删除工坊投稿评论（作者本人或管理员）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    with _community_lock:
        items = _load_json_file(_WORKSHOP_FILE, [])
        item = next((s for s in items if s.get('id') == item_id), None)
        if not item:
            return jsonify({'ok': False, 'error': '投稿不存在'}), 404
        comments = item.get('comments', [])
        comment = next((c for c in comments if c.get('id') == comment_id), None)
        if not comment:
            return jsonify({'ok': False, 'error': '评论不存在'}), 404
        is_admin_user = bool(user.is_admin if user.is_admin is not None else False)
        if comment.get('username') != user.username and not is_admin_user:
            return jsonify({'ok': False, 'error': '只能删除自己的评论'}), 403
        item['comments'] = [c for c in comments if c.get('id') != comment_id]
        _save_json_file(_WORKSHOP_FILE, items)

    return jsonify({'ok': True})


@app.route('/api/workshop/items/<int:item_id>')
def api_workshop_item_detail(item_id):
    """工坊投稿详情（含评论，公开）。附带点赞数/收藏数/当前用户点赞态/作者头像。"""
    with _community_lock:
        items = _load_json_file(_WORKSHOP_FILE, [])
    item = next((s for s in items if s.get('id') == item_id), None)
    if not item:
        return jsonify({'ok': False, 'error': '投稿不存在'}), 404
    user = _get_current_user()
    fav_map = _fav_count_map()
    return jsonify({'ok': True, 'item': _enrich_workshop_item(item, fav_map, user)})


@app.route('/api/workshop/items/<int:item_id>/like', methods=['POST'])
def api_workshop_item_like(item_id):
    """点赞/取消点赞工坊投稿（需登录）。点赞人存投稿对象 likes 数组。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    with _community_lock:
        items = _load_json_file(_WORKSHOP_FILE, [])
        item = next((s for s in items if s.get('id') == item_id), None)
        if not item:
            return jsonify({'ok': False, 'error': '投稿不存在'}), 404
        likes = item.get('likes')
        if not isinstance(likes, list):
            likes = []
            item['likes'] = likes
        if user.username in likes:
            likes.remove(user.username)
            liked = False
        else:
            likes.append(user.username)
            liked = True
            # 被点赞通知（自己点赞自己不通知；同投稿聚合）
            _notify_author(item.get('author'), 'workshop_like', user.username,
                           f'点赞了你的投稿《{item.get("title", "")[:30]}》',
                           f'/workshop/{item_id}', group_key=f'workshop_like:{item_id}')
        _save_json_file(_WORKSHOP_FILE, items)
    return jsonify({'ok': True, 'liked': liked, 'like_count': len(likes)})


@app.route('/api/workshop/items', methods=['POST'])
def api_workshop_submit():
    """创意工坊投稿（需登录，文件可选上传）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    data = request.get_json(silent=True) or {}
    title = (data.get('title') or '').strip()
    desc = (data.get('desc') or '').strip()
    file_url = (data.get('file_url') or '').strip()

    if not title:
        return jsonify({'ok': False, 'error': '请填写主题名称'}), 400
    if len(title) > 60:
        return jsonify({'ok': False, 'error': '主题名称最多 60 字'}), 400
    if not desc:
        return jsonify({'ok': False, 'error': '请填写内容介绍'}), 400
    if len(desc) > 2000:
        return jsonify({'ok': False, 'error': '内容介绍最多 2000 字'}), 400

    with _community_lock:
        subs = _load_json_file(_WORKSHOP_FILE, [])
        item = {
            'id': max([s.get('id', 0) for s in subs], default=0) + 1,
            'title': title,
            'desc': desc,
            'file_url': file_url,
            'author': user.username,
            'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
            'cat': 'user',
        }
        subs.insert(0, item)
        _save_json_file(_WORKSHOP_FILE, subs)

    return jsonify({'ok': True, 'item': item})


@app.route('/api/workshop/items/<int:item_id>', methods=['DELETE'])
def api_workshop_item_delete(item_id):
    """删除工坊投稿（仅作者本人或管理员）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    with _community_lock:
        subs = _load_json_file(_WORKSHOP_FILE, [])
        item = next((s for s in subs if s.get('id') == item_id), None)
        if not item:
            return jsonify({'ok': False, 'error': '投稿不存在'}), 404
        is_admin_user = bool(user.is_admin if user.is_admin is not None else False)
        if item.get('author') != user.username and not is_admin_user:
            return jsonify({'ok': False, 'error': '只能删除自己的投稿'}), 403
        subs = [s for s in subs if s.get('id') != item_id]
        _save_json_file(_WORKSHOP_FILE, subs)

    return jsonify({'ok': True})


# ━━━ 资源删除 API ━━━

@app.route('/api/resources/<path:filename>', methods=['DELETE'])
def api_delete_resource(filename):
    """删除服务器资源库中的文件。"""
    filepath = (RESOURCES_DIR / filename).resolve()
    if not str(filepath).startswith(str(RESOURCES_DIR.resolve())):
        return jsonify({'error': 'Access denied'}), 403
    if not filepath.exists() or not filepath.is_file():
        return jsonify({'error': 'File not found'}), 404
    try:
        filepath.unlink()
        return jsonify({'success': True, 'name': filename})
    except Exception as e:
        return jsonify({'error': f'删除失败: {e}'}), 500


# ━━━ Mod 系统 ━━━
from core.mod_loader import ModLoader

_MODS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'mods')
_mod_loader = ModLoader(_MODS_DIR)

# 应用启动时加载所有Mod
_manifests = _mod_loader.discover()
_loaded_count = 0
for _m in _manifests:
    if _mod_loader.load_mod(_m, app):
        _loaded_count += 1
print(f"[Mod] 发现 {len(_manifests)} 个Mod，成功加载 {_loaded_count} 个")


# ━━━ Mod 管理 API ━━━
@app.route('/api/mods')
def api_list_mods():
    """列出所有已加载的Mod"""
    all_mods = _mod_loader.get_loaded_mods()
    # 添加未加载的Mod信息
    loaded_ids = {m['id'] for m in all_mods}
    for m in _manifests:
        if m.get('id') not in loaded_ids:
            backend = m.get('backend', {})
            all_mods.append({
                'id': m.get('id', '?'),
                'name': m.get('name', '?'),
                'version': m.get('version', '?'),
                'description': m.get('description', ''),
                'author': m.get('author', ''),
                'backend_enabled': backend.get('enabled', False),
                'frontend_enabled': m.get('frontend', {}).get('enabled', False),
            })
    return jsonify({'mods': all_mods, 'loaded_count': _loaded_count})


@app.route('/api/mods/<mod_id>')
def api_get_mod(mod_id):
    """获取单个Mod的详细信息"""
    info = _mod_loader.loaded.get(mod_id)
    if info is None:
        return jsonify({'error': f'Mod不存在或未加载: {mod_id}'}), 404
    return jsonify({
        'id': mod_id,
        'name': info.get('name', mod_id),
        'version': info.get('version', '?'),
        'description': info.get('description', ''),
        'author': info.get('author', ''),
        'dependencies': info.get('dependencies', {}),
        'backend': info.get('backend', {}),
        'frontend': info.get('frontend', {}),
    })


@app.route('/api/mods/<mod_id>/reload', methods=['POST'])
def api_reload_mod(mod_id):
    """重新加载Mod"""
    if mod_id not in _mod_loader.loaded:
        return jsonify({'error': f'Mod未加载: {mod_id}'}), 404
    info = _mod_loader.loaded[mod_id]
    _mod_loader.unload_mod(mod_id)
    ok = _mod_loader.load_mod(info, app)
    return jsonify({'mod_id': mod_id, 'reloaded': ok})


# ━━━ Mod 管理页面 ━━━
@app.route('/mods')
def mods_page():
    """Mod 管理界面"""
    return render_template('mods.html')


# ━━━ 骰子数据统计 API ━━━
import os as _os
_stats_lock = _threading.Lock()

def _load_stats():
    """从 PostgreSQL 加载骰子统计（原 dice_stats.json）"""
    try:
        from core.models import DiceStat as _DS
        rows = _DS.query.all()
        return {r.username: {'total': r.total, 'crit20': r.crit20, 'crit1': r.crit1} for r in rows}
    except Exception:
        return {}

def _save_stats(stats):
    """保存骰子统计到 PostgreSQL"""
    try:
        from core.models import DiceStat as _DS
        with db.engine.begin() as conn:
            conn.execute(_DS.__table__.delete())
            for uname, data in (stats or {}).items():
                conn.execute(_DS.__table__.insert().values(
                    username=uname, total=data.get('total', 0),
                    crit20=data.get('crit20', 0), crit1=data.get('crit1', 0)))
    except Exception:
        pass

@app.route('/api/dice-stats', methods=['GET'])
def api_get_dice_stats():
    """获取所有玩家的骰子统计数据（排行榜）"""
    with _stats_lock:
        stats = _load_stats()
    # 计算每个玩家的概率并排序
    leaderboard = []
    for name, data in stats.items():
        total = data.get('total', 0)
        crit20 = data.get('crit20', 0)
        crit1 = data.get('crit1', 0)
        leaderboard.append({
            'name': name,
            'total': total,
            'crit20': crit20,
            'crit1': crit1,
            'rate20': round(crit20 / total * 100, 1) if total > 0 else 0,
            'rate1': round(crit1 / total * 100, 1) if total > 0 else 0,
        })
    leaderboard.sort(key=lambda x: x['crit20'], reverse=True)
    return jsonify({'ok': True, 'leaderboard': leaderboard[:50]})

@app.route('/api/dice-stats', methods=['POST'])
def api_record_dice_stats():
    """记录一次d20掷骰结果"""
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    roll_val = data.get('roll')  # d20结果值
    if not name or roll_val is None:
        return jsonify({'ok': False, 'error': '缺少参数'}), 400
    try:
        roll_val = int(roll_val)
    except (ValueError, TypeError):
        return jsonify({'ok': False}), 400

    with _stats_lock:
        stats = _load_stats()
        if name not in stats:
            stats[name] = {'total': 0, 'crit20': 0, 'crit1': 0}
        stats[name]['total'] += 1
        if roll_val == 20:
            stats[name]['crit20'] += 1
        elif roll_val == 1:
            stats[name]['crit1'] += 1
        _save_stats(stats)
    return jsonify({'ok': True})

@app.route('/api/dice-stats', methods=['DELETE'])
def api_clear_dice_stats():
    """清空所有骰子统计数据"""
    username = request.args.get('name', '').strip()
    ip = request.remote_addr or 'unknown'
    with _stats_lock:
        _save_stats({})
    audit_log('dice_stats.clear', username=username, ip=ip, detail='清空全部骰子统计')
    _app_logger.info(f'管理员清空骰子统计数据 (操作者={username}, IP={ip})')
    return jsonify({'ok': True})


# ━━━ 前端错误上报 API ━━━

@app.route('/api/error-report', methods=['POST'])
def api_error_report():
    """接收前端 JavaScript 错误上报。

    请求体:
      { message, source, lineno, colno, stack, url, page, username }
    """
    data = request.get_json(silent=True) or {}
    log_frontend_error(data)
    return jsonify({'ok': True})


# ━━━ 管理：聊天归档 ━━━

@app.route('/api/admin/chat-archives', methods=['GET'])
def api_list_chat_archives():
    """列出所有聊天归档行 (按时间倒序，最近20个)。"""
    archives = []
    try:
        from core.models import ChatArchive as _CA
        rows = _CA.query.filter_by(channel='chat').order_by(_CA.id.desc()).limit(20).all()
        archives = [{'filename': r.filename, 'size': 0,
                     'time': r.created_at or ''} for r in rows]
    except Exception:
        pass
    return jsonify({'ok': True, 'archives': archives})


@app.route('/api/admin/chat-archives', methods=['DELETE'])
def api_clean_chat_archives():
    """清理聊天归档行 (保留最近 N 天，默认30天，最小7天)。

    URL参数: ?days=30
    """
    days = request.args.get('days', '30')
    try:
        days = max(7, int(days))
    except ValueError:
        days = 30

    cleaned = 0
    try:
        from core.models import ChatArchive as _CA
        cutoff = _time.strftime('%Y-%m-%d %H:%M:%S', _time.localtime(_time.time() - days * 86400))
        cleaned = _CA.query.filter(_CA.channel == 'chat', _CA.created_at < cutoff).delete()
        db.session.commit()
    except Exception:
        pass

    username = request.args.get('name', '').strip()
    ip = request.remote_addr or 'unknown'
    audit_log('chat_archive.clean', username=username, ip=ip,
              detail=f'清理{days}天前的归档，删除{cleaned}个文件')
    _app_logger.info(f'聊天归档清理: 删除{cleaned}个文件 (保留{days}天)')
    return jsonify({'ok': True, 'cleaned': cleaned, 'retention_days': days})


# ━━━ 管理：统计清理与归档 ━━━

@app.route('/api/admin/archive-stats', methods=['POST'])
def api_archive_stats():
    """将当前统计数据归档（快照），然后重置。

    生成一个带时间戳的统计快照文件在 data/stats_archive/ 目录。
    """
    data_dir = _Path(__file__).parent.parent / 'data' / 'stats_archive'
    data_dir.mkdir(parents=True, exist_ok=True)

    ts = _time.strftime('%Y%m%d_%H%M%S')
    result = {'archived': []}

    # 归档骰子统计
    try:
        with _stats_lock:
            dice_stats = _load_stats()
            if dice_stats:
                snapshot_file = data_dir / f'dice_stats_{ts}.json'
                with open(snapshot_file, 'w', encoding='utf-8') as f:
                    _json.dump(dice_stats, f, ensure_ascii=False)
                result['archived'].append('dice_stats')
                _save_stats({})
    except Exception as e:
        _app_logger.error(f'归档骰子统计失败: {e}')

    # 归档事件统计
    try:
        event_stats = _load_event_stats()
        if event_stats:
            snapshot_file = data_dir / f'event_stats_{ts}.json'
            with open(snapshot_file, 'w', encoding='utf-8') as f:
                _json.dump(event_stats, f, ensure_ascii=False)
            result['archived'].append('event_stats')
            _save_event_stats({})
    except Exception as e:
        _app_logger.error(f'归档事件统计失败: {e}')

    username = request.args.get('name', '').strip()
    ip = request.remote_addr or 'unknown'
    audit_log('stats.archive', username=username, ip=ip,
              detail=f'归档统计快照 ({ts})')
    _app_logger.info(f'统计归档完成: {result["archived"]}')
    result['snapshot_time'] = ts
    return jsonify({'ok': True, **result})


@app.route('/api/admin/stats-summary', methods=['GET'])
def api_stats_summary():
    """获取当前运行时数据概况 (用于管理面板)。

    返回:
      { chat_count, dice_records, event_records, combat_state, online_users,
        frontend_errors_recent, disk_usage }
    """
    # 在线用户数
    _prune_stale_users()
    online_count = len(_online_users)

    # 聊天消息数
    chat_count = len(_chat_messages)

    # 骰子统计记录数
    with _stats_lock:
        dice_stats = _load_stats()
    dice_records = sum(s.get('total', 0) for s in dice_stats.values())

    # 事件统计记录数
    event_stats = _load_event_stats()
    event_records = sum(
        s.get('_clicks', 0) for s in event_stats.values()
        if isinstance(s, dict)
    )

    # 聊天归档数（PostgreSQL chat_archives 行）
    archive_count = 0
    archive_size = 0
    try:
        from core.models import ChatArchive as _CA
        archive_count = _CA.query.filter_by(channel='chat').count()
    except Exception:
        pass

    # 前端错误数（最近24小时）
    recent_errors = 0
    cutoff = _time.time() - 86400
    try:
        from utils.logger import _FRONTEND_ERROR_FILE
        if _FRONTEND_ERROR_FILE.exists():
            with open(_FRONTEND_ERROR_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        try:
                            err = _json.loads(line.strip())
                            if err.get('time', '') >= _time.strftime('%Y-%m-%d', _time.localtime(cutoff)):
                                recent_errors += 1
                        except _json.JSONDecodeError:
                            pass
    except Exception:
        pass

    return jsonify({
        'ok': True,
        'chat_count': chat_count,
        'dice_records': dice_records,
        'event_records': event_records,
        'combat_has_participants': bool(_combat_state and _combat_state.get('combatants')),
        'online_users': online_count,
        'chat_archives': archive_count,
        'chat_archives_size_mb': round(archive_size / 1048576, 2),
        'frontend_errors_24h': recent_errors,
    })


# ━━━ 地图 JSON 存档 API ━━━

_MAP_SAVES_DIR = _Path(__file__).parent.parent / '跑团存档' / '地图'
_MAP_SAVES_DIR.mkdir(parents=True, exist_ok=True)


@app.route('/api/map-saves', methods=['GET'])
def api_list_map_saves():
    """列出所有服务端地图存档。

    返回:
      { saves: [{name, filename, size, time}] }
    """
    saves = []
    try:
        for f in sorted(_MAP_SAVES_DIR.glob('*.json'), key=lambda x: x.stat().st_mtime, reverse=True):
            saves.append({
                'name': f.stem,
                'filename': f.name,
                'size': f.stat().st_size,
                'size_mb': round(f.stat().st_size / 1048576, 2),
                'time': _time.strftime('%Y-%m-%d %H:%M', _time.localtime(f.stat().st_mtime)),
            })
    except Exception:
        pass
    return jsonify({'ok': True, 'saves': saves})


@app.route('/api/map-saves', methods=['POST'])
def api_save_map():
    """将当前地图状态保存到服务端 JSON 文件。

    请求体: { name: "存档名称", state: {...} }
      name  — 存档名称（会清理非法字符）
      state — collectState() 的完整输出
    """
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    state = data.get('state')

    if not name:
        return jsonify({'error': '需要存档名称'}), 400
    if not state or not isinstance(state, dict):
        return jsonify({'error': '需要地图状态数据'}), 400

    # 清理文件名中的非法字符
    safe_name = ''.join(c for c in name if c not in r'<>:"/\|?*')[:100]
    if not safe_name:
        return jsonify({'error': '存档名称无效'}), 400

    # 添加服务端元数据
    state['_server_meta'] = {
        'saved_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
        'saved_by': data.get('username', ''),
        'filename': safe_name + '.json',
    }

    try:
        filepath = _MAP_SAVES_DIR / (safe_name + '.json')
        with open(filepath, 'w', encoding='utf-8') as f:
            _json.dump(state, f, ensure_ascii=False)
        file_size = filepath.stat().st_size

        username = data.get('username', '').strip()
        ip = request.remote_addr or 'unknown'
        audit_log('map.save', username=username, ip=ip,
                  detail=f'保存地图存档: {safe_name} ({file_size/1048576:.1f}MB)')
        _app_logger.info(f'地图存档已保存: {safe_name} ({file_size/1048576:.1f}MB)')

        return jsonify({
            'ok': True,
            'name': safe_name,
            'size': file_size,
            'size_mb': round(file_size / 1048576, 2),
        })
    except Exception as e:
        _app_logger.error(f'地图存档保存失败: {e}')
        return jsonify({'error': f'保存失败: {e}'}), 500


@app.route('/api/map-saves/<name>', methods=['GET'])
def api_load_map_save(name):
    """从服务端加载指定地图存档。

    返回存档的完整 JSON 内容。
    """
    safe_name = ''.join(c for c in name if c not in r'<>:"/\|?*')[:100]
    filepath = _MAP_SAVES_DIR / (safe_name + '.json')

    if not filepath.exists():
        return jsonify({'error': '存档不存在'}), 404

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            state = _json.load(f)
        return jsonify({'ok': True, 'name': safe_name, 'state': state})
    except Exception as e:
        return jsonify({'error': f'读取失败: {e}'}), 500


@app.route('/api/map-saves/<name>', methods=['DELETE'])
def api_delete_map_save(name):
    """删除服务端地图存档"""
    safe_name = ''.join(c for c in name if c not in r'<>:"/\|?*')[:100]
    filepath = _MAP_SAVES_DIR / (safe_name + '.json')

    if not filepath.exists():
        return jsonify({'error': '存档不存在'}), 404

    try:
        filepath.unlink()
        username = request.args.get('name', '').strip()
        ip = request.remote_addr or 'unknown'
        audit_log('map.delete', username=username, ip=ip,
                  detail=f'删除地图存档: {safe_name}')
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': f'删除失败: {e}'}), 500


# ━━━ 用户认证 API ━━━

from core.models import User as _UserModel, NorthSave as _NorthSave


def _get_current_user() -> _UserModel | None:
    """从 session 获取当前登录用户"""
    user_id = _flask_session.get('user_id')
    if user_id:
        return db.session.get(_UserModel, user_id)
    return None


# ━━━ 用户收藏（酒馆帖子 / 工坊）与我的评论 ━━━
_favorites_lock = _threading.Lock()


def _load_favorites() -> dict:
    """从 PostgreSQL 加载收藏（原 data/favorites.json）。"""
    try:
        from core.models import Favorite as _FV
        rows = _FV.query.order_by(_FV.id).all()
        return {'favorites': [{'user_id': r.user_id, 'type': r.type,
                               'item_id': r.item_id, 'created_at': r.created_at or ''}
                              for r in rows]}
    except Exception:
        return {'favorites': []}


def _save_favorites(data: dict):
    """保存收藏到 PostgreSQL（全删全插保留 id）。"""
    try:
        from core.models import Favorite as _FV
        favs = data.get('favorites', []) if isinstance(data, dict) else []
        with db.engine.begin() as conn:
            conn.execute(_FV.__table__.delete())
            for i, f in enumerate(favs, start=1):
                conn.execute(_FV.__table__.insert().values(
                    id=f.get('id', i), user_id=f.get('user_id'),
                    type=f.get('type', 'post'), item_id=f.get('item_id'),
                    created_at=f.get('created_at', '')))
    except Exception:
        pass


def _favorite_title(fav: dict) -> str:
    """补全收藏项的标题/作者（用于我的收藏展示）。

    帖子/工坊投稿存储在 JSON 文件（community.json / workshop.json），
    需从同源读取，否则 PostgreSQL 模型查不到导致收藏列表为空。
    """
    try:
        if fav.get('type') == 'post':
            with _community_lock:
                posts = _load_json_file(_COMMUNITY_FILE, [])
            p = next((x for x in posts if x.get('id') == fav.get('item_id')), None)
            if p:
                return {'title': p.get('title', ''), 'author': p.get('author', '')}
        elif fav.get('type') == 'workshop':
            with _community_lock:
                items = _load_json_file(_WORKSHOP_FILE, [])
            it = next((x for x in items if x.get('id') == fav.get('item_id')), None)
            if it:
                return {'title': it.get('title', ''), 'author': it.get('author', '')}
    except Exception:
        pass
    return {'title': '', 'author': ''}


@app.route('/api/favorites', methods=['GET'])
def api_my_favorites():
    """我的收藏列表（登录）：含标题/作者，最新在前。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    with _favorites_lock:
        favs = [f for f in _load_favorites().get('favorites', []) if f.get('user_id') == user.id]
    favs.sort(key=lambda f: f.get('created_at', ''), reverse=True)
    out = []
    for f in favs:
        info = _favorite_title(f)
        if info['title']:
            out.append({
                'type': f.get('type'),
                'item_id': f.get('item_id'),
                'title': info['title'],
                'author': info['author'],
                'created_at': f.get('created_at', ''),
            })
    return jsonify({'ok': True, 'favorites': out})


@app.route('/api/favorites', methods=['POST'])
def api_toggle_favorite():
    """收藏/取消收藏。请求: {type: 'post'|'workshop', id}"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    data = request.get_json(silent=True) or {}
    fav_type = data.get('type', '')
    item_id = data.get('id')
    if fav_type not in ('post', 'workshop') or item_id is None:
        return jsonify({'ok': False, 'error': '参数错误'}), 400
    with _favorites_lock:
        store = _load_favorites()
        favs = store.setdefault('favorites', [])
        exists = next((f for f in favs if f.get('user_id') == user.id
                       and f.get('type') == fav_type and f.get('item_id') == item_id), None)
        if exists:
            favs.remove(exists)
            _save_favorites(store)
            return jsonify({'ok': True, 'favorited': False})
        favs.append({
            'user_id': user.id,
            'type': fav_type,
            'item_id': item_id,
            'created_at': _time.strftime('%Y-%m-%d %H:%M:%S'),
        })
        _save_favorites(store)
        # 被收藏通知（自己收藏自己不通知；同项聚合；作者从 JSON 数据源读取）
        try:
            if fav_type == 'post':
                with _community_lock:
                    posts = _load_json_file(_COMMUNITY_FILE, [])
                target = next((p for p in posts if p.get('id') == item_id), None)
                if target:
                    _notify_author(target.get('author'), 'post_fav', user.username,
                                   f'收藏了你的帖子《{str(target.get("title", ""))[:30]}》',
                                   f'/post/{item_id}', group_key=f'post_fav:{item_id}')
            elif fav_type == 'workshop':
                with _community_lock:
                    items = _load_json_file(_WORKSHOP_FILE, [])
                target = next((s for s in items if s.get('id') == item_id), None)
                if target:
                    _notify_author(target.get('author'), 'workshop_fav', user.username,
                                   f'收藏了你的投稿《{str(target.get("title", ""))[:30]}》',
                                   f'/workshop/{item_id}', group_key=f'workshop_fav:{item_id}')
        except Exception:
            pass
        return jsonify({'ok': True, 'favorited': True})


@app.route('/api/my/comments')
def api_my_comments():
    """我的评论（登录）：遍历酒馆帖子与工坊评论，最新在前。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    out = []
    try:
        from core.models import CommunityPost as _CP
        for p in _CP.query.all():
            for c in (p.comments or []):
                if c.get('username') == user.username:
                    out.append({
                        'type': 'post', 'parent_id': p.id,
                        'parent_title': p.title or '', 'content': c.get('content', ''),
                        'created_at': c.get('created_at', ''),
                    })
    except Exception:
        pass
    try:
        from core.models import WorkshopItem as _WI
        for it in _WI.query.all():
            for c in (it.comments or []):
                if c.get('username') == user.username:
                    out.append({
                        'type': 'workshop', 'parent_id': it.id,
                        'parent_title': it.title or '', 'content': c.get('content', ''),
                        'created_at': c.get('created_at', ''),
                    })
    except Exception:
        pass
    out.sort(key=lambda c: c.get('created_at', ''), reverse=True)
    return jsonify({'ok': True, 'comments': out[:100]})


# ━━━ 全站在线状态 API（打开任意页面即在线，关闭全部页面才下线）━━━

def _prune_online_sessions(now: float = None) -> None:
    """清理心跳超时的在线会话（页面异常关闭/断网兜底）。"""
    now = now or _time.time()
    stale = [sid for sid, s in _online_sessions.items()
             if now - s['last_seen'] > ONLINE_SESSION_TIMEOUT]
    for sid in stale:
        _online_sessions.pop(sid, None)


@app.route('/api/online/heartbeat', methods=['POST'])
def api_online_heartbeat():
    """页面心跳：每 20 秒上报一次，表示该标签页仍打开。

    仅登录用户计入在线；未登录心跳仅用于移除该会话。
    """
    data = request.get_json(silent=True) or {}
    sid = (data.get('session_id') or '').strip()[:64]
    page = (data.get('page') or '')[:100]
    if not sid:
        return jsonify({'ok': False, 'error': '缺少 session_id'}), 400

    user = _get_current_user()
    with _online_lock:
        if user:
            _online_sessions[sid] = {
                'username': user.username,
                'user_id': user.id,
                'ip': request.remote_addr or '',
                'page': page,
                'last_seen': _time.time(),
            }
        else:
            _online_sessions.pop(sid, None)
    return jsonify({'ok': True})


@app.route('/api/online/leave', methods=['POST'])
def api_online_leave():
    """页面关闭：移除该标签页的在线会话。"""
    data = request.get_json(silent=True) or {}
    sid = (data.get('session_id') or '').strip()[:64]
    if sid:
        with _online_lock:
            _online_sessions.pop(sid, None)
    return jsonify({'ok': True})


@app.route('/api/online/users')
def api_online_users():
    """当前在线用户列表（按会话数降序）。"""
    with _online_lock:
        _prune_online_sessions()
        users: dict[str, dict] = {}
        for s in _online_sessions.values():
            entry = users.setdefault(s['username'], {'sessions': 0, 'page': s['page']})
            entry['sessions'] += 1
        online = [{'username': name, 'sessions': v['sessions'], 'page': v['page']}
                  for name, v in sorted(users.items(), key=lambda x: -x[1]['sessions'])]
    return jsonify({'ok': True, 'online': online, 'count': len(online)})


@app.route('/api/auth/register', methods=['POST'])
def api_auth_register():
    """用户注册。用户ID由数据库自动分配（注册顺序），不可修改。用户名可在用户中心修改。"""
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = (data.get('password') or '').strip()
    confirm_password = (data.get('confirm_password') or '').strip()

    if not username or not email or not password:
        return jsonify({'ok': False, 'error': '请填写所有字段'}), 400
    if len(username) < 2 or len(username) > 30:
        return jsonify({'ok': False, 'error': '用户名需2-30个字符'}), 400
    # 用户名白名单：字母/数字/下划线/连字符/中文，防 HTML/JS 注入与审计日志污染
    if not _re.fullmatch(r'[\w一-龥-]{2,30}', username):
        return jsonify({'ok': False, 'error': '用户名仅限中文、字母、数字、下划线、连字符'}), 400
    if len(password) < 9:
        return jsonify({'ok': False, 'error': '密码至少9位'}), 400
    if not (_re.search(r'[a-zA-Z]', password) and _re.search(r'\d', password)):
        return jsonify({'ok': False, 'error': '密码必须包含英文和数字'}), 400
    if password != confirm_password:
        return jsonify({'ok': False, 'error': '两次密码不一致'}), 400

    if _UserModel.query.filter_by(username=username).first():
        return jsonify({'ok': False, 'error': '用户名已存在'}), 409
    if _UserModel.query.filter_by(email=email).first():
        return jsonify({'ok': False, 'error': '邮箱已被注册'}), 409

    user = _UserModel(username=username, email=email)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    # 自动登录
    _flask_session['user_id'] = user.id
    _flask_session['username'] = user.username

    ip = request.remote_addr or 'unknown'
    audit_log('user.register', username=username, ip=ip, detail=f'新用户注册 id={user.id}')
    _app_logger.info(f'新用户注册: {username} ({email})')

    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    """用户登录（支持用户名或邮箱作为账号）"""
    data = request.get_json(silent=True) or {}
    account = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()

    if not account or not password:
        return jsonify({'ok': False, 'error': '请填写账号（用户名或邮箱）和密码'}), 400

    # 先按用户名查，查不到再按邮箱查（邮箱不区分大小写）
    user = _UserModel.query.filter_by(username=account).first()
    if not user:
        user = _UserModel.query.filter_by(email=account.lower()).first()
    if not user or not user.check_password(password):
        return jsonify({'ok': False, 'error': '用户名或密码错误'}), 401
    if not (user.is_active if user.is_active is not None else True):
        return jsonify({'ok': False, 'error': '账号已被禁用，请联系管理员'}), 403

    _flask_session['user_id'] = user.id
    _flask_session['username'] = user.username

    # 更新最后登录时间与 IP
    user.last_login = datetime.utcnow()
    user.last_login_ip = request.remote_addr or ''
    # 自动关联旧角色（created_by 匹配但 user_id 为空）
    from core.models import Character as _Char
    _Char.query.filter(
        _Char.user_id.is_(None),
        _Char.created_by == user.username,
    ).update({_Char.user_id: user.id}, synchronize_session=False)
    db.session.commit()

    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    """获取当前登录用户信息"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '未登录'}), 401
    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/auth/update-profile', methods=['POST'])
def api_auth_update_profile():
    """修改个人资料（用户名、邮箱、手机、简介）"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    data = request.get_json(silent=True) or {}
    new_username = (data.get('username') or '').strip()
    new_email = (data.get('email') or '').strip().lower()
    phone = (data.get('phone') or '').strip()
    bio = (data.get('bio') or '').strip()

    if new_username and new_username != user.username:
        if len(new_username) < 2 or len(new_username) > 30:
            return jsonify({'ok': False, 'error': '用户名需2-30个字符'}), 400
        # 用户名白名单：字母/数字/下划线/连字符/中文
        if not _re.fullmatch(r'[\w一-龥-]{2,30}', new_username):
            return jsonify({'ok': False, 'error': '用户名仅限中文、字母、数字、下划线、连字符'}), 400
        if _UserModel.query.filter_by(username=new_username).first():
            return jsonify({'ok': False, 'error': '用户名已存在'}), 409
        old_username = user.username
        user.username = new_username
        _flask_session['username'] = new_username
        # 同步更新历史帖子/投稿的作者名：否则改名后他人点赞/收藏/评论，
        # 通知按旧用户名找不到作者，发布者收不到消息
        try:
            with _community_lock:
                posts = _load_json_file(_COMMUNITY_FILE, [])
                p_changed = False
                for p in posts:
                    if p.get('author') == old_username:
                        p['author'] = new_username
                        p_changed = True
                if p_changed:
                    _save_json_file(_COMMUNITY_FILE, posts)
                items = _load_json_file(_WORKSHOP_FILE, [])
                w_changed = False
                for s in items:
                    if s.get('author') == old_username:
                        s['author'] = new_username
                        w_changed = True
                if w_changed:
                    _save_json_file(_WORKSHOP_FILE, items)
        except Exception:
            pass
    if new_email and new_email != user.email:
        if _UserModel.query.filter_by(email=new_email).first():
            return jsonify({'ok': False, 'error': '邮箱已被使用'}), 409
        user.email = new_email
    if phone:
        if len(phone) > 50:
            return jsonify({'ok': False, 'error': '手机号过长'}), 400
        user.phone = phone
    if bio:
        if len(bio) > 500:
            return jsonify({'ok': False, 'error': '简介过长（最多500字）'}), 400
        user.bio = bio

    db.session.commit()
    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/auth/change-password', methods=['POST'])
def api_auth_change_password():
    """修改密码"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    data = request.get_json(silent=True) or {}
    old_password = (data.get('old_password') or '').strip()
    new_password = (data.get('new_password') or '').strip()

    if not user.check_password(old_password):
        return jsonify({'ok': False, 'error': '原密码错误'}), 400
    if len(new_password) < 9:
        return jsonify({'ok': False, 'error': '新密码至少9位'}), 400
    if not (_re.search(r'[a-zA-Z]', new_password) and _re.search(r'\d', new_password)):
        return jsonify({'ok': False, 'error': '新密码必须包含英文和数字'}), 400

    user.set_password(new_password)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    """退出登录"""
    _flask_session.pop('user_id', None)
    _flask_session.pop('username', None)
    return jsonify({'ok': True})


# ━━━ 用户头像（上传 / 删除 / 批量查询）━━━
# 头像存 web/static/avatars/u{user_id}.{ext}，经 /static/ 直接提供
_AVATAR_DIR = _Path(__file__).parent / 'static' / 'avatars'
_AVATAR_EXTS = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')
_AVATAR_MAX = 2 * 1024 * 1024  # 2MB


@app.route('/api/auth/avatar', methods=['POST'])
def api_upload_avatar():
    """上传用户头像（需登录）。multipart/form-data 字段 'avatar'。

    自动删除同用户旧头像文件，avatar_url 指向 /static/avatars/ 下新文件。
    """
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    if 'avatar' not in request.files:
        return jsonify({'ok': False, 'error': '请选择图片文件'}), 400
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'ok': False, 'error': '未选择文件'}), 400

    ext = _os.path.splitext(file.filename)[1].lower()
    if ext not in _AVATAR_EXTS:
        return jsonify({'ok': False, 'error': '仅支持图片格式: JPG/PNG/GIF/WEBP/BMP'}), 400

    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > _AVATAR_MAX:
        return jsonify({'ok': False, 'error': '图片过大（最大2MB）'}), 400

    try:
        _AVATAR_DIR.mkdir(parents=True, exist_ok=True)
        # 删除同用户旧头像（可能是其他扩展名）
        for old in _AVATAR_DIR.glob(f'u{user.id}.*'):
            try:
                old.unlink()
            except Exception:
                pass
        fname = f'u{user.id}{ext}'
        file.save(str(_AVATAR_DIR / fname))
    except Exception as _e:
        return jsonify({'ok': False, 'error': f'头像保存失败: {_e}'}), 500

    user.avatar_url = f'/static/avatars/{fname}'
    db.session.commit()
    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/auth/avatar', methods=['DELETE'])
def api_delete_avatar():
    """删除用户头像（恢复默认）。"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401
    if user.avatar_url:
        try:
            old = _Path(__file__).parent / user.avatar_url.lstrip('/')
            if old.is_file():
                old.unlink()
        except Exception:
            pass
        user.avatar_url = ''
        db.session.commit()
    return jsonify({'ok': True, 'user': user.to_dict()})


@app.route('/api/users/avatars')
def api_users_avatars():
    """批量查询用户名→头像映射。?names=a,b,c（官网帖子/评论/在线列表展示用）"""
    names = [n.strip() for n in (request.args.get('names') or '').split(',') if n.strip()]
    if not names:
        return jsonify({'ok': True, 'avatars': {}})
    users = _UserModel.query.filter(_UserModel.username.in_(names)).all()
    avatars = {u.username: (u.avatar_url or '') for u in users}
    return jsonify({'ok': True, 'avatars': avatars})


# 注入到 Jinja2 模板全局变量
@app.context_processor
def _inject_user():
    return {'current_user': _get_current_user()}


# ━━━ 北境雪原存档 API ━━━

@app.route('/api/north/save', methods=['POST'])
def api_north_save():
    """保存北境雪原存档到数据库。

    请求体: { save_name, save_data }
      save_name: 存档名称（默认 'auto'）
      save_data: JSON 格式的完整游戏状态
    自动关联当前登录用户（如已登录）
    """
    data = request.get_json(silent=True) or {}
    save_name = (data.get('save_name') or 'auto').strip()
    save_data = data.get('save_data')
    if not save_data:
        return jsonify({'ok': False, 'error': '缺少存档数据'}), 400

    user = _get_current_user()
    user_id = user.id if user else None

    # 查找现有存档（同一用户+同一名称）
    existing = _NorthSave.query.filter_by(user_id=user_id, save_name=save_name).first()
    if existing:
        existing.save_data = _json.dumps(save_data, ensure_ascii=False)
        existing.updated_at = datetime.utcnow()
    else:
        ns = _NorthSave(
            user_id=user_id,
            save_name=save_name,
            save_data=_json.dumps(save_data, ensure_ascii=False),
        )
        db.session.add(ns)
    db.session.commit()

    return jsonify({'ok': True, 'save_name': save_name})


@app.route('/api/north/load', methods=['GET'])
def api_north_load():
    """加载北境雪原存档。

    URL参数: ?save_name=auto（默认 'auto'）
    已登录用户加载自己的存档，未登录返回空。
    """
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录官网账号'}), 401

    save_name = request.args.get('save_name', 'auto').strip()
    ns = _NorthSave.query.filter_by(user_id=user.id, save_name=save_name).first()
    if not ns:
        return jsonify({'ok': False, 'save_data': None})

    try:
        save_data = _json.loads(ns.save_data)
    except Exception:
        save_data = None

    return jsonify({
        'ok': True,
        'save_name': save_name,
        'save_data': save_data,
        'updated_at': ns.updated_at.isoformat() if ns.updated_at else '',
    })


@app.route('/api/north/saves', methods=['GET'])
def api_north_list_saves():
    """列出当前用户的所有北境存档"""
    user = _get_current_user()
    if not user:
        return jsonify({'ok': False, 'error': '请先登录'}), 401

    saves = _NorthSave.query.filter_by(user_id=user.id).order_by(_NorthSave.updated_at.desc()).all()
    return jsonify({
        'ok': True,
        'saves': [{
            'save_name': s.save_name,
            'updated_at': s.updated_at.isoformat() if s.updated_at else '',
            'size': len(s.save_data) if s.save_data else 0,
        } for s in saves],
    })


@app.route('/api/north/avatar', methods=['POST'])
def api_north_avatar_upload():
    """北境冒险者头像（徽章）上传。multipart/form-data 字段 'avatar'。

    登录用户：north_u<user_id>.<ext> 命名，上传时自动删除同用户旧头像；
    未登录：north_temp.<ext> 单文件覆盖（仅当前会话生效，与北境存档行为一致）。
    """
    if 'avatar' not in request.files:
        return jsonify({'ok': False, 'error': '请选择图片文件'}), 400
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'ok': False, 'error': '未选择文件'}), 400

    ext = _os.path.splitext(file.filename)[1].lower()
    if ext not in _AVATAR_EXTS:
        return jsonify({'ok': False, 'error': '仅支持图片格式: JPG/PNG/GIF/WEBP/BMP'}), 400

    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > _AVATAR_MAX:
        return jsonify({'ok': False, 'error': '图片过大（最大2MB）'}), 400

    user = _get_current_user()
    if user:
        fname = f'north_u{user.id}{ext}'
        # 删除同用户旧头像（可能是其他扩展名）
        for old in _AVATAR_DIR.glob(f'north_u{user.id}.*'):
            try:
                old.unlink()
            except Exception:
                pass
    else:
        fname = f'north_temp{ext}'

    try:
        _AVATAR_DIR.mkdir(parents=True, exist_ok=True)
        file.save(str(_AVATAR_DIR / fname))
    except Exception as _e:
        return jsonify({'ok': False, 'error': f'头像保存失败: {_e}'}), 500

    return jsonify({'ok': True, 'url': f'/static/avatars/{fname}'})


def run_server():
    """启动 Web 服务器（HTTP + WebSocket 共用 5000 端口）"""
    # 数据库表已在模块加载时通过 _init_orm_db(app) 创建
    # 安装请求日志中间件
    setup_request_logging(app, _app_logger)
    _app_logger.info('尘封之卷 Web 服务器启动中...')

    # Flask-Sock + simple-websocket 自动处理 /ws 路径的 WebSocket 升级
    # debug=False：生产使用；避免 Werkzeug 调试器暴露远程执行入口
    app.config['TEMPLATES_AUTO_RELOAD'] = True

    # IPv4 主监听（0.0.0.0：本机与局域网均可访问）
    server = make_server('0.0.0.0', 5000, app, threaded=True)
    # IPv6 ::1 副监听（同端口、不同地址族可共存）：
    # 浏览器访问 localhost 会先尝试 IPv6，若 ::1 无监听则被防火墙静默丢弃、
    # 超时后再回退 IPv4（每个连接约 300ms，页面多个资源叠加导致切换明显变慢）。
    # 增加 ::1 监听后浏览器直连成功，无回退等待。
    try:
        server_v6 = make_server('::1', 5000, app, threaded=True)
        _threading.Thread(target=server_v6.serve_forever, daemon=True).start()
        _app_logger.info('已监听 IPv6 ::1（加速 localhost 访问）')
    except Exception as _e:
        _app_logger.warning(f'IPv6 ::1 监听失败（不影响 IPv4 访问）: {_e}')
    server.serve_forever()


if __name__ == '__main__':
    run_server()

