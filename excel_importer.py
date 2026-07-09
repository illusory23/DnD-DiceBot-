"""DND 5E 人物卡 Excel 导入器

支持多种格式的 DND 5E 人物卡 Excel 文件:
  - NekoWorks ver2.1 (旧版)
  - New Edition v4.4.0 (新版, 丹恩·铁拳等角色使用的格式)
"""

import re
import os
from pathlib import Path
from typing import Any, Callable, Optional


def _safe_int(value: Any) -> int | None:
    """安全转换为整数"""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value: Any) -> str | None:
    """安全转换为字符串"""
    if value is None:
        return None
    return str(value).strip()


def _parse_attack_bonus(value: str) -> int | None:
    """解析攻击加值，例如 "D20+5" -> 5"""
    if value is None:
        return None
    value = str(value).strip()
    match = re.search(r'[+-]?\d+', value.replace('D20', ''))
    if match:
        return int(match.group())
    return None


def _or(val, default):
    """如果 val 是 None，返回 default"""
    return default if val is None else val


def _find_coins_in_cells(cells: dict, location_sets: list[dict]) -> dict:
    """在多个候选位置中查找钱币值，返回第一个非全零的结果

    每个 location_sets 项是一个 {币种: 单元格地址} 字典。
    """
    for loc_set in location_sets:
        coins = {}
        for coin_type in ('cp', 'sp', 'ep', 'gp', 'pp'):
            addr = loc_set.get(coin_type, '')
            val = _safe_int(cells.get(addr, 0)) or 0
            coins[coin_type] = val
        if sum(coins.values()) > 0:
            return coins
    return {'cp': 0, 'sp': 0, 'ep': 0, 'gp': 0, 'pp': 0}


# ━━━ Excel 图片提取 ━━━

def _extract_images_from_workbook(wb, dest_dir: Path) -> list[str]:
    """从 Excel 工作簿中提取嵌入的图片。

    遍历所有 sheet 的 _images 属性，将嵌入的图片保存到 dest_dir，
    返回已保存的文件路径列表。

    注意: 使用 openpyxl 内部 API (ws._images)，最佳努力提取。
    """
    saved_paths = []
    image_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, '_images'):
            continue
        for img in ws._images:
            try:
                # 尝试获取图片数据
                if hasattr(img, 'ref') and hasattr(img.ref, '_data'):
                    data = img.ref._data
                elif hasattr(img, '_data'):
                    data = img._data
                else:
                    continue

                # 确定扩展名
                ext = '.png'
                content_type = getattr(img.ref, 'content_type', '') if hasattr(img, 'ref') else ''
                if 'jpeg' in content_type or 'jpg' in content_type:
                    ext = '.jpg'
                elif 'gif' in content_type:
                    ext = '.gif'
                elif 'bmp' in content_type:
                    ext = '.bmp'

                filename = f'portrait_{image_count}{ext}'
                filepath = dest_dir / filename

                with open(filepath, 'wb') as f:
                    f.write(data)

                saved_paths.append(str(filepath))
                image_count += 1
            except Exception:
                continue  # 单张图片提取失败不影响其他

    return saved_paths


# ━━━ New Edition v4.4.0 单元格映射 ━━━

def _import_v440(ws, cells: dict) -> dict:
    """导入 New Edition v4.4.0 格式的角色卡"""
    result: dict = {}

    # ━ 基本信息 ━
    basic = {
        'name': cells.get('E3', ''),
        'player': cells.get('E4', ''),
        'class': cells.get('S3', ''),
        'level': _safe_int(cells.get('Y3', 1)) or 1,
        'xp': _safe_int(cells.get('S4', 0)) or 0,
        'race': cells.get('E6', ''),
        'gender': cells.get('M6', ''),
        'subrace': cells.get('E7', ''),
        'age': _safe_str(cells.get('M7', '')),
        'height': cells.get('E9', ''),
        'alignment': '' if cells.get('L9', '') in ('阵营', '陣營', None) else cells.get('L9', ''),
        'weight': cells.get('E10', ''),
        'faith': '' if cells.get('L10', '') in ('信仰', None) else cells.get('L10', ''),
        'proficiency_bonus': _safe_int(cells.get('U9', 2)) or 2,
    }
    result['basic'] = basic

    # ━ 属性值 (rows 14-24) ━
    ability_rows = {
        'str': 14, 'dex': 16, 'con': 18, 'int': 20, 'wis': 22, 'cha': 24,
    }
    ability_names = {
        'str': '力量', 'dex': '敏捷', 'con': '体质',
        'int': '智力', 'wis': '感知', 'cha': '魅力',
    }
    abilities = {}
    save_profs = {}
    for key, row_num in ability_rows.items():
        score = _safe_int(cells.get(f'F{row_num}', 10)) or 10
        save_val = _safe_int(cells.get(f'T{row_num}', 0)) or 0
        abilities[key] = score

        expected_mod = (score - 10) // 2
        name = ability_names[key]
        # 豁免加值与属性调整值差2以上则视为熟练
        prof_checkbox = cells.get(f'A{row_num}')  # 熟练复选框
        is_prof = bool(prof_checkbox) or (save_val != expected_mod)
        save_profs[name] = {
            'is_proficient': is_prof,
            'save_bonus': save_val,
        }
    result['abilities'] = abilities
    result['save_proficiencies'] = save_profs

    # ━ 战斗数据 ━
    hp_cur = _safe_int(cells.get('M29', 0)) or 0
    hp_max = _safe_int(cells.get('Q29', 0)) or hp_cur
    # 也可以从 AW4 获取总HP
    if hp_max == 0:
        hp_max = _safe_int(cells.get('AW4', 0)) or 10

    speed_str = _safe_str(cells.get('R38', '30尺')) or '30尺'
    speed_match = re.search(r'(\d+)', speed_str)
    speed = int(speed_match.group(1)) if speed_match else 30

    combat = {
        'hp_current': hp_cur,
        'hp_max': hp_max,
        'ac': _safe_int(cells.get('D32', 10)) or 10,
        'initiative_bonus': _safe_int(cells.get('D29', 0)) or 0,
        'speed': speed,
        'hd_count': _safe_int(cells.get('L32', 1)) or 1,
        'hd_type': '1d12',  # 默认野蛮人
        'passive_perception': _safe_int(cells.get('F40', 10)) or 10,
    }
    result['combat'] = combat

    # ━ 技能熟练 (rows 45-66) ━
    # 技能 → 对应属性 (用于计算加值)
    skill_to_ability_map = {
        '运动': 'str', '特技': 'dex', '巧手': 'dex', '隐匿': 'dex',
        '调查': 'int', '奥秘': 'int', '历史': 'int', '自然': 'int', '宗教': 'int',
        '察觉': 'wis', '洞悉': 'wis', '驯兽': 'wis', '医疗': 'wis', '生存': 'wis',
        '游说': 'cha', '欺诈': 'cha', '威吓': 'cha', '表演': 'cha',
    }
    skill_rows = {
        45: '运动', 47: '特技', 48: '巧手', 49: '隐匿',
        51: '调查', 52: '奥秘', 53: '历史', 54: '自然', 55: '宗教',
        57: '察觉', 58: '洞悉', 59: '驯兽', 60: '医疗', 61: '生存',
        63: '游说', 64: '欺诈', 65: '威吓', 66: '表演',
    }
    skill_profs = {}
    for row_num, skill_name in skill_rows.items():
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) if prof_checkbox is not None else False
        is_exp = False
        if isinstance(prof_checkbox, (int, float)) and prof_checkbox == 2:
            is_exp = True

        # 读取技能加值 (T列: 总加值 = 属性调整值 + 熟练加值)
        skill_bonus = _safe_int(cells.get(f'T{row_num}'))
        if skill_bonus is None:
            # 从属性调整值+熟练加值计算
            ability_key = skill_to_ability_map.get(skill_name, 'str')
            ability_score = abilities.get(ability_key, 10)
            ability_mod = (ability_score - 10) // 2
            prof_bonus = basic.get('proficiency_bonus', 2)
            if is_prof:
                skill_bonus = ability_mod + (prof_bonus * 2 if is_exp else prof_bonus)
            else:
                skill_bonus = ability_mod

        skill_profs[skill_name] = {
            'is_proficient': is_prof or is_exp,
            'is_expertise': is_exp,
            'bonus': skill_bonus or 0,
        }
    result['skill_proficiencies'] = skill_profs

    # ━ 武器 (rows 47-51) ━
    # 计算基础攻击加值: 先看武器列中是否有加值, 否则用力量/敏捷+熟练
    str_mod = (abilities.get('str', 10) - 10) // 2
    dex_mod = (abilities.get('dex', 10) - 10) // 2
    prof = basic.get('proficiency_bonus', 2)

    # 常见近战武器关键词
    melee_keywords = ['巨斧', '战锤', '长剑', '短剑', '匕首', '斧', '锤', '棍',
                      '矛', '戟', '鞭', '镰', '镐', '连枷', '巨剑', '弯刀',
                      '手斧', '轻锤', '钉头锤', '刺剑', '细剑', '军刀']
    # 常见远程武器关键词
    ranged_keywords = ['弓', '弩', '枪', '标枪', '飞镖', '投石', '矢']

    weapons = []
    for row_num in [47, 48, 49, 50, 51]:
        name = cells.get(f'P{row_num}', '')
        if name and str(name).strip():
            name_str = str(name).strip()

            # 跳过非武器条目
            non_weapon_kw = ['铜币', '银币', '金币', '白金币', 'GP', 'SP', 'CP', 'EP', 'PP']
            if any(kw in name_str for kw in non_weapon_kw) and len(name_str) < 8:
                continue

            # 尝试从T列读取加值 (可能存储为公式结果)
            atk_val = _safe_int(cells.get(f'T{row_num}'))
            if atk_val is None:
                # 计算攻击加值
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                if is_ranged and not is_melee:
                    atk_val = dex_mod + prof
                else:
                    atk_val = str_mod + prof

            # 伤害骰在AI列
            dmg_dice = _safe_str(cells.get(f'AI{row_num}', '')) or ''

            weapons.append({
                'name': name_str,
                'attack_bonus': atk_val or 0,
                'damage_dice': dmg_dice,
                'damage_type': _safe_str(cells.get(f'AN{row_num}', '')) or '',
                'ammo': _safe_str(cells.get(f'AQ{row_num}', '')) or '',
                'is_proficient': True,
                'weight': _safe_str(cells.get(f'AA{row_num}', '')) or '',
            })
    result['weapons'] = weapons

    # ━ 护甲 ━
    armor_info = {
        'armor_name': '',
        'armor_ac': _safe_int(cells.get('AD45', 0)) or 0,
        'armor_max_dex': _safe_str(cells.get('AF45', '')) or '',
        'shield_name': '',
        'shield_ac': 0,
        'shield_weight': '',
    }
    result['armor'] = armor_info

    # ━ 法术 ━
    spell_ability = _safe_str(cells.get('AD54', '')) or ''
    if spell_ability in ('施法属性', '施法屬性', None):
        spell_ability = ''
    spell_dc = _safe_int(cells.get('AI58', 10)) or 10
    spell_info = {
        'spellcasting_ability': spell_ability,
        'spell_attack_bonus': 0,
        'spell_ability': spell_ability,
        'spell_save_dc': spell_dc,
        'prepared_spell_count': 0,
    }
    result['spell_info'] = spell_info

    # ━ 法术位 (AM58-AM66 → 1环-9环) ━
    spell_slots = {}
    for level in range(1, 10):
        cell_row = 57 + level  # AM58=1环, AM59=2环, ...
        max_slots = _safe_int(cells.get(f'AM{cell_row}', 0)) or 0
        spell_slots[str(level)] = {'max': max_slots, 'used': 0}
    result['spell_slots'] = spell_slots

    # ━ 已准备法术 — 从法术书 sheet 或 prepared spells 区域读取 ━
    # v4.4.0 的法术在法术书I/II/III sheets
    result['prepared_spells'] = []

    # ━ 钱币 ━
    # 搜索多个可能的钱币位置 (AI列, AE/AG列)
    coins = _find_coins_in_cells(cells, [
        # v4.4.0 AI列 (旧位置)
        {'cp': 'AI43', 'sp': 'AI44', 'ep': 'AI45', 'gp': 'AI46', 'pp': 'AI47'},
        # v4.4.0 AE列 + AG列 (新位置, rows 62-66)
        {'cp': 'AG66', 'sp': 'AG65', 'ep': 'AG63', 'gp': 'AG64', 'pp': 'AG62'},
        {'cp': 'AE66', 'sp': 'AE65', 'ep': 'AE63', 'gp': 'AE64', 'pp': 'AE62'},
    ])
    result['coins'] = coins

    # ━ 负重 ━
    result['weight'] = {
        'light_load': 0, 'medium_load': 0, 'heavy_load': 0, 'current_load': 0,
    }

    # ━ 背景 (从背景 sheet 读取) ━
    result['background'] = {}

    return result


# ━━━ NekoWorks ver2.1 单元格映射 ━━━

def _import_nekox21(ws, cells: dict) -> dict:
    """导入 NekoWorks ver2.1 格式的角色卡

    NekoWorks v2.1 布局特点:
      - 武器在行28-32 (P列为名称, AE列为攻击骰)
      - 钱币在行43-48 (AE列为币种)
      - 法术位在 AM39-AM47
    """
    result: dict = {}

    # 基本信息
    name = cells.get('E3') or cells.get('B3', '')
    player = cells.get('E4') or cells.get('B4', '')
    cls = cells.get('S3') or cells.get('P3', '')
    race = cells.get('E6') or cells.get('B6', '')
    subrace = cells.get('E7') or cells.get('B7', '')

    # 阵营/信仰 — 过滤掉标签文本
    # NekoWorks v2.1: 标签在X列，值在AA列
    alignment_raw = (cells.get('AA6') or cells.get('L9') or cells.get('X6', ''))
    faith_raw = (cells.get('AA7') or cells.get('L10') or cells.get('X7', ''))
    if alignment_raw in ('阵营', '陣營', None):
        alignment_raw = ''
    if faith_raw in ('信仰', None):
        faith_raw = ''

    basic = {
        'name': name,
        'player': player,
        'class': cls,
        'level': _safe_int(cells.get('Y3', 1)) or 1,
        'xp': _safe_int(cells.get('S4', 0)) or _safe_int(cells.get('P4', 0)) or 0,
        'race': race,
        'gender': cells.get('M6') or cells.get('J6', ''),
        'subrace': subrace,
        'age': _safe_str(cells.get('M7') or cells.get('J7', '')) or '',
        'height': cells.get('E9') or cells.get('Q6', ''),
        'alignment': alignment_raw,
        'weight': cells.get('E10') or cells.get('Q7', ''),
        'faith': faith_raw,
        'proficiency_bonus': _safe_int(cells.get('U9', 2)) or _safe_int(cells.get('AD3', 2)) or 2,
    }
    result['basic'] = basic

    # 属性值
    use_new_rows = cells.get('F14') is not None
    if use_new_rows:
        ability_rows = {k: v for k, v in zip(
            ['str', 'dex', 'con', 'int', 'wis', 'cha'],
            [14, 16, 18, 20, 22, 24]
        )}
    else:
        ability_rows = {k: v for k, v in zip(
            ['str', 'dex', 'con', 'int', 'wis', 'cha'],
            [11, 13, 15, 17, 19, 21]
        )}
    ability_names = {k: v for k, v in zip(
        ['str', 'dex', 'con', 'int', 'wis', 'cha'],
        ['力量', '敏捷', '体质', '智力', '感知', '魅力']
    )}
    abilities = {}
    save_profs = {}
    for key, row_num in ability_rows.items():
        score = _safe_int(cells.get(f'F{row_num}', 10)) or 10
        save_val = _safe_int(cells.get(f'T{row_num}')) or _safe_int(cells.get(f'K{row_num}')) or 0
        abilities[key] = score

        expected_mod = (score - 10) // 2
        name = ability_names[key]
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) or (save_val != expected_mod)
        save_profs[name] = {'is_proficient': is_prof, 'save_bonus': save_val}
    result['abilities'] = abilities
    result['save_proficiencies'] = save_profs

    # 战斗数据
    hp_cur = (_safe_int(cells.get('M29')) or _safe_int(cells.get('Y11')) or 0)
    hp_max = (_safe_int(cells.get('Q29')) or _safe_int(cells.get('AC11')) or hp_cur or 10)
    ac_val = (_safe_int(cells.get('D32')) or _safe_int(cells.get('P13')) or 10)
    init_val = (_safe_int(cells.get('D29')) or _safe_int(cells.get('P11')) or 0)
    speed_str = _safe_str(cells.get('R38', '30尺')) or '30尺'
    speed_match = re.search(r'(\d+)', speed_str)
    speed = int(speed_match.group(1)) if speed_match else 30

    combat = {
        'hp_current': hp_cur,
        'hp_max': hp_max,
        'ac': ac_val,
        'initiative_bonus': init_val,
        'speed': speed,
        'hd_count': (_safe_int(cells.get('AA13')) or _safe_int(cells.get('X13'))
                     or _safe_int(cells.get('L32', 1)) or 1),
        'hd_type': '1d10',  # 游侠默认d10
        'passive_perception': (_safe_int(cells.get('R21')) or _safe_int(cells.get('F40')) or 10),
    }
    result['combat'] = combat

    # 技能熟练 — NekoWorks v2.1使用不同的行号
    # 先判断用哪组行: 检查行26的C列是否有技能名
    use_nekox_skill_rows = cells.get('C26') is not None and str(cells.get('C26', '')).strip() in ('运动', '運動')
    if use_nekox_skill_rows:
        neko_skill_rows = {
            26: '运动', 28: '特技', 29: '巧手', 30: '隐匿',
            33: '调查', 34: '历史', 35: '自然', 36: '宗教',
            38: '察觉', 39: '洞悉', 40: '驯兽', 41: '医疗', 42: '生存',
            44: '游说', 45: '欺诈', 46: '威吓', 47: '表演',
        }
        skill_rows = neko_skill_rows
    else:
        skill_rows = {
            45: '运动', 47: '特技', 48: '巧手', 49: '隐匿',
            51: '调查', 52: '奥秘', 53: '历史', 54: '自然', 55: '宗教',
            57: '察觉', 58: '洞悉', 59: '驯兽', 60: '医药', 61: '生存',
            63: '游说', 64: '欺诈', 65: '威吓', 66: '表演',
        }
    skill_profs = {}
    for row_num, skill_name in skill_rows.items():
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) if prof_checkbox is not None else False
        is_exp = isinstance(prof_checkbox, (int, float)) and prof_checkbox == 2

        # NekoWorks格式: I列为技能加值
        skill_bonus = _safe_int(cells.get(f'I{row_num}'))

        skill_profs[skill_name] = {
            'is_proficient': is_prof or is_exp,
            'is_expertise': is_exp,
            'bonus': skill_bonus or 0,
        }
    result['skill_proficiencies'] = skill_profs

    # 武器 — NekoWorks格式的行号(28-32)与v4.4.0(47-51)不同
    # NekoWorks v2.1 使用 O 列(武器名)，v4.4.0 使用 P 列
    # 判断用哪组行: 检查行28的O列或P列是否有武器名
    use_nekox_rows = (
        (cells.get('O28') is not None and str(cells.get('O28', '')).strip()) or
        (cells.get('P28') is not None and str(cells.get('P28', '')).strip())
    )
    weapon_rows = [28, 29, 30, 31, 32] if use_nekox_rows else [47, 48, 49, 50, 51]

    # NekoWorks v2.1: 武器名在O列, 攻击在AE列, 伤害在AI列
    if use_nekox_rows and cells.get('O28') is not None:
        name_col = 'O'
        atk_col = 'AE'
    else:
        name_col = 'P'
        atk_col = 'T'

    str_mod = (abilities.get('str', 10) - 10) // 2
    dex_mod = (abilities.get('dex', 10) - 10) // 2
    prof = basic.get('proficiency_bonus', 2)
    melee_keywords = ['巨斧', '战锤', '长剑', '短剑', '匕首', '斧', '锤', '棍',
                      '矛', '戟', '鞭', '镰', '镐', '连枷', '巨剑', '弯刀',
                      '手斧', '轻锤', '钉头锤', '刺剑', '细剑', '军刀',
                      '细刃', '短剑', '长矛', '长弓', '短弓']
    ranged_keywords = ['弓', '弩', '枪', '标枪', '飞镖', '投石', '矢']
    non_weapon_keywords = ['铜币', '银币', '金币', '白金币', 'CP', 'SP', 'GP', 'EP', 'PP',
                           'Lv', 'QuQ', '重量', '法术', '名称']

    weapons = []
    for row_num in weapon_rows:
        name = cells.get(f'{name_col}{row_num}', '')
        if not (name and str(name).strip()):
            continue
        name_str = str(name).strip()
        # 跳过非武器条目
        if any(kw in name_str for kw in non_weapon_keywords) and len(name_str) < 8:
            continue

        # 解析攻击加值
        if atk_col == 'AE':
            # NekoWorks v2.1: AE列可能是 "D20+" 文本，加值需从技能/属性推算
            atk_raw = _safe_str(cells.get(f'{atk_col}{row_num}', ''))
            atk_val = _parse_attack_bonus(atk_raw) if atk_raw else None
            if atk_val is None:
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                if is_ranged and not is_melee:
                    atk_val = dex_mod + prof
                else:
                    atk_val = str_mod + prof
        else:
            atk_val = _safe_int(cells.get(f'{atk_col}{row_num}'))
            if atk_val is None:
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                atk_val = (dex_mod if (is_ranged and not is_melee) else str_mod) + prof

        # NekoWorks v2.1: 伤害骰在AI列, 伤害类型在AN列
        dmg_dice = _safe_str(cells.get(f'AI{row_num}', '')) or ''
        dmg_type = _safe_str(cells.get(f'AN{row_num}', '')) or ''
        ammo = _safe_str(cells.get(f'AQ{row_num}', '')) or ''
        weight = _safe_str(cells.get(f'AA{row_num}', '')) or ''

        weapons.append({
            'name': name_str,
            'attack_bonus': atk_val or 0,
            'damage_dice': dmg_dice,
            'damage_type': dmg_type,
            'ammo': ammo,
            'is_proficient': True,
            'weight': weight,
            'description': '',
            'effect': '',
        })
    result['weapons'] = weapons

    # 护甲
    result['armor'] = {
        'armor_name': '',
        'armor_ac': (_safe_int(cells.get('AD45')) or _safe_int(cells.get('AD25')) or 0),
        'shield_name': '',
        'shield_ac': 0,
    }

    # 法术 — 过滤标签文本
    spell_ability = (_safe_str(cells.get('AD54')) or _safe_str(cells.get('AD35')) or '')
    if spell_ability in ('施法属性', '施法屬性'):
        spell_ability = ''
    spell_info = {
        'spellcasting_ability': spell_ability,
        'spell_ability': spell_ability,
        'spell_save_dc': (_safe_int(cells.get('AI58')) or _safe_int(cells.get('AD40')) or 10),
        'prepared_spell_count': 0,
    }
    result['spell_info'] = spell_info

    # 法术位 — NekoWorks格式用AM39-AM47, v4.4.0用AM58-AM66
    spell_slots = {}
    for level in range(1, 10):
        # 尝试两个位置
        max_slots = (_safe_int(cells.get(f'AM{57+level}'))  # v4.4.0
                     or _safe_int(cells.get(f'AM{38+level}'))  # NekoWorks
                     or 0)
        # 如果 max_slots 等于 level 号，可能是行号标签而非真实法术位
        if max_slots == level:
            max_slots = 0
        spell_slots[str(level)] = {'max': max_slots, 'used': 0}
    result['spell_slots'] = spell_slots

    # 已准备法术
    result['prepared_spells'] = []

    # 钱币 — 尝试多个位置
    # NekoWorks v2.1: AE(币种标签)/AG(数量) 列
    # 也可能在 AI 列
    coins = _find_coins_in_cells(cells, [
        {'cp': 'AG47', 'sp': 'AG46', 'ep': 'AG44', 'gp': 'AG45', 'pp': 'AG43'},
        {'cp': 'AI43', 'sp': 'AI44', 'ep': 'AI45', 'gp': 'AI46', 'pp': 'AI47'},
        {'cp': 'AI47', 'sp': 'AI46', 'ep': 'AI44', 'gp': 'AI45', 'pp': 'AI43'},
    ])
    result['coins'] = coins
    result['weight'] = {'light_load': 0, 'medium_load': 0, 'heavy_load': 0, 'current_load': 0}
    result['background'] = {}
    return result


# ━━━ 公共接口 ━━━

def _detect_version(ws) -> str:
    """检测Excel模板版本"""
    # 检查 O1 单元格
    o1 = str(ws.cell(1, 15).value or '')  # O1
    if 'v4.4' in o1 or 'New Edition' in o1:
        return 'v440'
    if 'v4.3' in o1:
        return 'v440'  # 兼容
    if 'NekoWorks' in o1 or 'ver2' in o1:
        return 'neko21'
    # 默认尝试 v4.4.0 (较新的格式)
    # 通过检查是否有 New Edition 特有单元格来判断
    e3 = str(ws.cell(3, 5).value or '')  # E3
    if e3 and e3 not in ('角色名', ''):
        # 如果E3有非标签内容，可能是v4.4.0格式
        b3 = str(ws.cell(3, 2).value or '')  # B3
        if '角色名' in b3:
            return 'v440'
    return 'neko21'


def _build_cell_dict(ws, max_rows: int = 200) -> dict:
    """构建单元格值字典"""
    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                            max_col=min(60, ws.max_column)):
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value
    return cells


def _read_multi_row_cells(cells: dict, col: str, start_row: int, max_empty: int = 2) -> list[str]:
    """读取某列连续行的多个值，遇到连续空行或B列有新标签时停止"""
    values = []
    empty_count = 0
    label_keywords = {'角色名', '玩家', '年龄', '身高', '体重', '肤色', '发色', '瞳色',
                      '故乡', '出身', '语言', '熟练工具', '个性', '理念', '羁绊', '缺陷',
                      '外貌', '背景', '人物形象'}
    for row_num in range(start_row, start_row + 20):
        # 检查B列是否有新标签（表示进入了新section）
        b_val = str(cells.get(f'B{row_num}', '')).strip().replace('\n', '')
        if b_val:
            # 检查是否是已知标签（非数据行标题）
            is_label = any(kw in b_val for kw in label_keywords)
            if is_label and row_num > start_row:
                break

        val = cells.get(f'{col}{row_num}', '')
        if val:
            val_str = str(val).strip()
            if val_str:
                values.append(val_str)
                empty_count = 0
            else:
                empty_count += 1
        else:
            empty_count += 1

        if empty_count >= max_empty:
            break
    return values


def _import_background_sheet(wb, result: dict) -> None:
    """从背景sheet读取背景信息，支持v4.4.0和NekoWorks v2.1两种格式"""
    bg_sheet = None
    for name in wb.sheetnames:
        if '背景' in name:
            bg_sheet = name
            break

    if not bg_sheet:
        return

    ws = wb[bg_sheet]
    bg = {}
    cells = _build_cell_dict(ws, max_rows=50)

    # 检测格式: v4.4.0在R列有"个性"标签, NekoWorks在E列
    is_v440 = '个性' in str(cells.get('R3', ''))

    # 标签关键词（用于过滤掉纯标签文本）
    label_texts = {'个性', '理念', '羁绊', '缺陷', '背景特性', '背景故事',
                   '外貌描述', '外貌描写', '人物衣装', '人物背景',
                   '角色名', '玩家', '年龄', '身高', '体重',
                   '故乡', '出身', '语言', '熟练工具'}

    if is_v440:
        # ━ v4.4.0 格式: R列=标签, U列=值 ━
        # 个性/理念/羁绊/缺陷/外貌/背景故事在U列
        v440_fields = {
            'U3': 'personality_traits',
            'U5': 'ideals',
            'U6': 'bonds',
            'U7': 'flaws',
            'U9': 'appearance',
            'U18': 'backstory',
        }
        for cell_addr, field in v440_fields.items():
            val = cells.get(cell_addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg[field] = val_str

        # background_feature: R12=标签, 值可能在U12或R13/U13区域
        for addr in ('U12', 'R13', 'U13', 'R14', 'U14'):
            val = cells.get(addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg['background_feature'] = val_str
                    break

        # origin (出身/故乡): B列标签→E列值
        for row in range(3, 15):
            b_label = str(cells.get(f'B{row}', '')).strip()
            if '故乡' in b_label or '出身' in b_label:
                e_val = cells.get(f'E{row}', '')
                if e_val:
                    val_str = str(e_val).strip()
                    if val_str and val_str not in label_texts:
                        bg['origin'] = val_str
                        break

        # languages (语言): B列"语言"标签 → C列连续值
        for row in range(28, 38):
            b_label = str(cells.get(f'B{row}', '')).strip().replace('\n', '')
            if '语' in b_label and '言' in b_label:
                lang_values = _read_multi_row_cells(cells, 'C', row)
                if lang_values:
                    bg['languages'] = '\n'.join(lang_values)
                break

        # tool_proficiencies (熟练工具): B列"工具"标签 → C列连续值
        for row in range(33, 43):
            b_label = str(cells.get(f'B{row}', '')).strip().replace('\n', '')
            if '工具' in b_label and '熟练' in b_label:
                tool_values = _read_multi_row_cells(cells, 'C', row)
                if tool_values:
                    bg['tool_proficiencies'] = '\n'.join(tool_values)
                break

    else:
        # ━ NekoWorks v2.1 格式: E列=标签+值 ━
        neko_bg_fields = {
            'E3': 'personality_traits',
            'E4': 'personality_traits_ext',
            'E5': 'ideals',
            'E6': 'bonds',
            'E7': 'flaws',
            'E9': 'background_feature',
            'E13': 'appearance',
            'E17': 'backstory',
            'E18': 'origin',
            'E20': 'languages',
            'E27': 'tool_proficiencies',
        }
        for cell_addr, field in neko_bg_fields.items():
            val = cells.get(cell_addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg[field] = val_str

    result['background'] = bg


def _import_inventory_sheet(wb) -> list[dict]:
    """从物品/背包sheet读取物品列表

    支持 NekoWorks v2.1 的「物品」sheet 和 v4.4.0 的「背包」sheet。
    物品以 B 列为名称，I 列为描述，逐行列出。
    """
    items = []

    # 查找物品或背包sheet（优先"背包"，其次"物品"，避免匹配到"装备物品"）
    inv_sheet_name = None
    for name in wb.sheetnames:
        if '背包' in name:
            inv_sheet_name = name
            break
    if not inv_sheet_name:
        for name in wb.sheetnames:
            if '物品' in name and '装备' not in name:
                inv_sheet_name = name
                break

    if not inv_sheet_name:
        return items

    ws = wb[inv_sheet_name]
    cells = _build_cell_dict(ws, max_rows=250)

    # 跳过关键词
    skip_exact = {'名称', '背包', '描述', 'Inventory'}

    for row_num in range(4, min(ws.max_row + 1, 250)):
        name = cells.get(f'B{row_num}', '')
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()

        # 跳过标题标签行
        if name_str in skip_exact:
            continue
        # 跳过"背包N"或"背包N（...）"等标签
        if (name_str.startswith('背包') and
            (len(name_str) <= 6 or name_str[2:3].isdigit())):
            # 但保留如"背包（次元袋）"这样的实际物品
            if '（' not in name_str and '(' not in name_str:
                continue
        # 跳过纯数字
        if name_str.isdigit():
            continue
        # 跳过"次元袋负重"等元数据
        if '负重' in name_str:
            continue

        desc = str(cells.get(f'I{row_num}', '')) if cells.get(f'I{row_num}') else ''
        items.append({
            'item_name': name_str,
            'quantity': 1,
            'weight': 0,
            'location': '背包',
            'description': desc.strip(),
            'effect': '',
        })

    return items


def _import_equipment_sheet(wb) -> dict:
    """从装备物品sheet读取护甲、武器和冒险用品

    装备物品sheet包含三张表：
      - 护甲表 (A-G列): 护甲名称/AC/属性
      - 武器表 (I-N列): 武器名称/伤害/属性(含描述)
      - 冒险用品表 (Q-T列): 物品/价格/重量
    这些是通用参考表，不是角色实际装备的物品。
    角色的实际装备在主要sheet的武器区域和物品sheet中。
    此函数用于补充读取可能遗漏的装备描述信息。
    """
    eq_sheet_name = None
    for name in wb.sheetnames:
        if '装备物品' in name:
            eq_sheet_name = name
            break

    if not eq_sheet_name:
        return {}

    ws = wb[eq_sheet_name]
    cells = _build_cell_dict(ws, max_rows=60)

    # 读取武器属性描述（N列），用于补充主要sheet中武器的描述
    weapon_descriptions = {}
    for row_num in range(3, 40):
        name = cells.get(f'J{row_num}', '')
        prop = cells.get(f'N{row_num}', '')
        if name and str(name).strip():
            name_str = str(name).strip()
            prop_str = str(prop).strip() if prop else ''
            weapon_descriptions[name_str] = prop_str

    return {'weapon_descriptions': weapon_descriptions}


def import_character_from_excel(filepath: str | Path) -> dict:
    """从 DND 5E 人物卡 Excel 文件导入角色数据

    自动检测模板版本并适配对应的单元格映射。

    Args:
        filepath: Excel 文件路径

    Returns:
        包含完整角色数据的字典

    Raises:
        FileNotFoundError: 文件不存在
        ImportError: 缺少 openpyxl 库
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "需要安装 openpyxl 库来读取 Excel 文件。\n"
            "请运行: pip install openpyxl"
        )

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    # 尝试提取嵌入图片（最佳努力）
    extracted_images = []
    try:
        portrait_dir = filepath.parent
        extracted_images = _extract_images_from_workbook(wb, portrait_dir)
    except Exception:
        pass  # 图片提取失败不影响导入

    # 查找"主要"sheet
    main_sheet = None
    for name in wb.sheetnames:
        if '主要' in name:
            main_sheet = name
            break

    if main_sheet is None:
        if len(wb.sheetnames) > 1:
            main_sheet = wb.sheetnames[1]
        else:
            main_sheet = wb.sheetnames[0]

    ws = wb[main_sheet]

    # 检测版本
    version = _detect_version(ws)

    # 构建单元格字典
    cells = _build_cell_dict(ws)

    # 按版本解析
    if version == 'v440':
        result = _import_v440(ws, cells)
    else:
        result = _import_nekox21(ws, cells)

    # 读取背景
    _import_background_sheet(wb, result)

    # 读取物品/背包
    inventory_items = _import_inventory_sheet(wb)
    if inventory_items:
        result['inventory'] = inventory_items

    # 读取装备物品表（补充武器描述）
    eq_data = _import_equipment_sheet(wb)
    if eq_data.get('weapon_descriptions'):
        # 用装备物品表的属性描述补充武器信息
        wp_descs = eq_data['weapon_descriptions']
        for w in result.get('weapons', []):
            wp_name = w.get('name', '')
            if wp_name in wp_descs and not w.get('description'):
                w['description'] = wp_descs[wp_name]

    # 确保 inventory 键存在
    if 'inventory' not in result:
        result['inventory'] = []

    # 附加提取的图片路径（第一张作为头像）
    if extracted_images:
        result['portrait_path'] = extracted_images[0]

    wb.close()
    return result


def import_and_print_summary(filepath: str | Path) -> str:
    """导入角色并返回摘要文本，用于预览"""
    data = import_character_from_excel(filepath)
    basic = data.get('basic', {})
    abilities = data.get('abilities', {})
    combat = data.get('combat', {})
    spell_info = data.get('spell_info', {})

    name = _or(basic.get('name'), '未知')
    level = _or(basic.get('level'), 1)
    cls = _or(basic.get('class'), '未知')
    race = _or(basic.get('race'), '未知')
    alignment = _or(basic.get('alignment'), '未知')
    faith = _or(basic.get('faith'), '—')
    hp_cur = _or(combat.get('hp_current'), 0)
    hp_max = _or(combat.get('hp_max'), 0)
    ac = _or(combat.get('ac'), 10)
    init_bonus = _or(combat.get('initiative_bonus'), 0)
    speed = _or(combat.get('speed'), 30)

    lines = [
        f"📜 角色卡预览: {name}",
        f"   等级: {level} | 职业: {cls} | 种族: {race}",
        f"   阵营: {alignment} | 信仰: {faith}",
        f"   ❤️ HP: {hp_cur}/{hp_max} | 🛡️ AC: {ac}",
        f"   ⚡ 先攻: {init_bonus:+d} | 🏃 速度: {speed}ft",
    ]

    # 属性
    abbr = {'str': '力量', 'dex': '敏捷', 'con': '体质',
            'int': '智力', 'wis': '感知', 'cha': '魅力'}
    ab_parts = []
    for key, label in abbr.items():
        score = _or(abilities.get(key), 10)
        mod_val = (score - 10) // 2 if score else 0
        ab_parts.append(f"{label}:{score}({mod_val:+d})")
    lines.append(f"   📊 {' | '.join(ab_parts)}")

    # 法术
    spell_ability = _or(spell_info.get('spellcasting_ability'), '')
    if spell_ability:
        spell_dc = _or(spell_info.get('spell_save_dc'), '—')
        lines.append(f"   ✨ 施法: {spell_ability} | DC: {spell_dc}")

    # 武器
    weapons = data.get('weapons', [])
    if weapons:
        wp_lines = []
        for w in weapons:
            wp_lines.append(f"{w['name']} ({w.get('damage_dice', '?')})")
        lines.append(f"   ⚔️ 武器: {', '.join(wp_lines)}")

    # 已准备法术
    prepared = data.get('prepared_spells', [])
    if prepared:
        spell_names = [s['name'] for s in prepared[:8]]
        more = f"... 等{len(prepared)}个" if len(prepared) > 8 else ""
        lines.append(f"   📖 已准备法术: {', '.join(spell_names)}{more}")

    # 物品
    inventory = data.get('inventory', [])
    if inventory:
        item_names = [it['item_name'] for it in inventory[:10]]
        more = f"... 等{len(inventory)}件" if len(inventory) > 10 else ""
        lines.append(f"   🎒 物品: {', '.join(item_names)}{more}")

    return "\n".join(lines)
