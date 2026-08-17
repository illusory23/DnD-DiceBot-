"""DND 5E 人物卡 Excel 导入器

支持多种格式的 DND 5E 人物卡 Excel 文件:
  - NekoWorks ver2.1 (旧版)
  - New Edition v4.4.0 (新版, 丹恩·铁拳等角色使用的格式)
"""

import re
import os
from pathlib import Path
from typing import Any, Callable, Optional


def _safe_int(value: Any) -> int | None:
    """安全转换为整数"""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value: Any) -> str | None:
    """安全转换为字符串"""
    if value is None:
        return None
    return str(value).strip()


def _parse_attack_bonus(value: str) -> int | None:
    """解析攻击加值,例如 "D20+5" -> 5"""
    if value is None:
        return None
    value = str(value).strip()
    match = re.search(r'[+-]?\d+', value.replace('D20', ''))
    if match:
        return int(match.group())
    return None


def _or(val, default):
    """如果 val 是 None,返回 default"""
    return default if val is None else val


def _find_coins_in_cells(cells: dict, location_sets: list[dict]) -> dict:
    """在多个候选位置中查找钱币值,返回第一个非全零的结果

    每个 location_sets 项是一个 {币种: 单元格地址} 字典。
    """
    for loc_set in location_sets:
        coins = {}
        for coin_type in ('cp', 'sp', 'ep', 'gp', 'pp'):
            addr = loc_set.get(coin_type, '')
            val = _safe_int(cells.get(addr, 0)) or 0
            coins[coin_type] = val
        if sum(coins.values()) > 0:
            return coins
    return {'cp': 0, 'sp': 0, 'ep': 0, 'gp': 0, 'pp': 0}


# ━━━ Excel 图片提取 ━━━

def _extract_images_from_workbook(wb, dest_dir: Path) -> list[str]:
    """从 Excel 工作簿中提取嵌入的图片。

    遍历所有 sheet 的 _images 属性,将嵌入的图片保存到 dest_dir,
    返回已保存的文件路径列表。

    注意: 使用 openpyxl 内部 API (ws._images),最佳努力提取。
    """
    saved_paths = []
    image_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, '_images'):
            continue
        for img in ws._images:
            try:
                # 尝试获取图片数据
                if hasattr(img, 'ref') and hasattr(img.ref, '_data'):
                    data = img.ref._data
                elif hasattr(img, '_data'):
                    data = img._data
                else:
                    continue

                # 确定扩展名
                ext = '.png'
                content_type = getattr(img.ref, 'content_type', '') if hasattr(img, 'ref') else ''
                if 'jpeg' in content_type or 'jpg' in content_type:
                    ext = '.jpg'
                elif 'gif' in content_type:
                    ext = '.gif'
                elif 'bmp' in content_type:
                    ext = '.bmp'

                filename = f'portrait_{image_count}{ext}'
                filepath = dest_dir / filename

                with open(filepath, 'wb') as f:
                    f.write(data)

                saved_paths.append(str(filepath))
                image_count += 1
            except Exception:
                continue  # 单张图片提取失败不影响其他

    return saved_paths


# ━━━ New Edition v4.4.0 单元格映射 ━━━

def _import_v440(ws, cells: dict) -> dict:
    """导入 New Edition v4.4.0 格式的角色卡"""
    print('[excel-import] _import_v440 开始解析')
    result: dict = {}

    # ━ 基本信息 ━
    basic = {
        'name': cells.get('E3', ''),
        'player': cells.get('E4', ''),
        'class': cells.get('S3', ''),
        'level': _safe_int(cells.get('Y3', 1)) or 1,
        'xp': _safe_int(cells.get('S4', 0)) or 0,
        'race': cells.get('E6', ''),
        'gender': cells.get('M6', ''),
        'subrace': cells.get('E7', ''),
        'age': _safe_str(cells.get('M7', '')),
        'height': cells.get('E9', ''),
        'alignment': '' if cells.get('L9', '') in ('阵营', '陣營', None) else cells.get('L9', ''),
        'weight': cells.get('E10', ''),
        'faith': '' if cells.get('L10', '') in ('信仰', None) else cells.get('L10', ''),
        'proficiency_bonus': _safe_int(cells.get('U9', 2)) or 2,
    }
    result['basic'] = basic

    # ━ 属性值 (rows 14-24) ━
    ability_rows = {
        'str': 14, 'dex': 16, 'con': 18, 'int': 20, 'wis': 22, 'cha': 24,
    }
    ability_names = {
        'str': '力量', 'dex': '敏捷', 'con': '体质',
        'int': '智力', 'wis': '感知', 'cha': '魅力',
    }
    abilities = {}
    save_profs = {}
    for key, row_num in ability_rows.items():
        score = _safe_int(cells.get(f'F{row_num}', 10)) or 10
        save_val = _safe_int(cells.get(f'T{row_num}', 0)) or 0
        abilities[key] = score

        expected_mod = (score - 10) // 2
        name = ability_names[key]
        # 豁免加值与属性调整值差2以上则视为熟练
        prof_checkbox = cells.get(f'A{row_num}')  # 熟练复选框
        is_prof = bool(prof_checkbox) or (save_val != expected_mod)
        save_profs[name] = {
            'is_proficient': is_prof,
            'save_bonus': save_val,
        }
    result['abilities'] = abilities
    result['save_proficiencies'] = save_profs

    # ━ 战斗数据 ━
    hp_cur = _safe_int(cells.get('M29', 0)) or 0
    hp_max = _safe_int(cells.get('Q29', 0)) or hp_cur
    # 也可以从 AW4 获取总HP
    if hp_max == 0:
        hp_max = _safe_int(cells.get('AW4', 0)) or 10

    speed_str = _safe_str(cells.get('R38', '30尺')) or '30尺'
    speed_match = re.search(r'(\d+)', speed_str)
    speed = int(speed_match.group(1)) if speed_match else 30

    combat = {
        'hp_current': hp_cur,
        'hp_max': hp_max,
        'ac': _safe_int(cells.get('D32', 10)) or 10,
        'initiative_bonus': _safe_int(cells.get('D29', 0)) or 0,
        'speed': speed,
        'hd_count': _safe_int(cells.get('L32', 1)) or 1,
        'hd_type': '1d12',  # 默认野蛮人
        'passive_perception': _safe_int(cells.get('F40', 10)) or 10,
    }
    result['combat'] = combat

    # ━ 技能熟练 (rows 45-66) ━
    # 技能 → 对应属性 (用于计算加值)
    skill_to_ability_map = {
        '运动': 'str', '特技': 'dex', '巧手': 'dex', '隐匿': 'dex',
        '调查': 'int', '奥秘': 'int', '历史': 'int', '自然': 'int', '宗教': 'int',
        '察觉': 'wis', '洞悉': 'wis', '驯兽': 'wis', '医疗': 'wis', '生存': 'wis',
        '游说': 'cha', '欺诈': 'cha', '威吓': 'cha', '表演': 'cha',
    }
    skill_rows = {
        45: '运动', 47: '特技', 48: '巧手', 49: '隐匿',
        51: '调查', 52: '奥秘', 53: '历史', 54: '自然', 55: '宗教',
        57: '察觉', 58: '洞悉', 59: '驯兽', 60: '医疗', 61: '生存',
        63: '游说', 64: '欺诈', 65: '威吓', 66: '表演',
    }
    skill_profs = {}
    for row_num, skill_name in skill_rows.items():
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) if prof_checkbox is not None else False
        is_exp = False
        if isinstance(prof_checkbox, (int, float)) and prof_checkbox == 2:
            is_exp = True

        # 读取技能加值 (T列: 总加值 = 属性调整值 + 熟练加值)
        skill_bonus = _safe_int(cells.get(f'T{row_num}'))
        if skill_bonus is None:
            # 从属性调整值+熟练加值计算
            ability_key = skill_to_ability_map.get(skill_name, 'str')
            ability_score = abilities.get(ability_key, 10)
            ability_mod = (ability_score - 10) // 2
            prof_bonus = basic.get('proficiency_bonus', 2)
            if is_prof:
                skill_bonus = ability_mod + (prof_bonus * 2 if is_exp else prof_bonus)
            else:
                skill_bonus = ability_mod

        skill_profs[skill_name] = {
            'is_proficient': is_prof or is_exp,
            'is_expertise': is_exp,
            'bonus': skill_bonus or 0,
        }
    result['skill_proficiencies'] = skill_profs

    # ━ 武器 (rows 47-51) ━
    # 计算基础攻击加值: 先看武器列中是否有加值, 否则用力量/敏捷+熟练
    str_mod = (abilities.get('str', 10) - 10) // 2
    dex_mod = (abilities.get('dex', 10) - 10) // 2
    prof = basic.get('proficiency_bonus', 2)

    # 常见近战武器关键词
    melee_keywords = ['巨斧', '战锤', '长剑', '短剑', '匕首', '斧', '锤', '棍',
                      '矛', '戟', '鞭', '镰', '镐', '连枷', '巨剑', '弯刀',
                      '手斧', '轻锤', '钉头锤', '刺剑', '细剑', '军刀']
    # 常见远程武器关键词
    ranged_keywords = ['弓', '弩', '枪', '标枪', '飞镖', '投石', '矢']

    weapons = []
    for row_num in [47, 48, 49, 50, 51]:
        name = cells.get(f'P{row_num}', '')
        if name and str(name).strip():
            name_str = str(name).strip()

            # 跳过非武器条目
            non_weapon_kw = ['铜币', '银币', '金币', '白金币', 'GP', 'SP', 'CP', 'EP', 'PP']
            if any(kw in name_str for kw in non_weapon_kw) and len(name_str) < 8:
                continue

            # 尝试从T列读取加值 (可能存储为公式结果)
            atk_val = _safe_int(cells.get(f'T{row_num}'))
            if atk_val is None:
                # 计算攻击加值
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                if is_ranged and not is_melee:
                    atk_val = dex_mod + prof
                else:
                    atk_val = str_mod + prof

            # 伤害骰在AI列
            dmg_dice = _safe_str(cells.get(f'AI{row_num}', '')) or ''

            weapons.append({
                'name': name_str,
                'attack_bonus': atk_val or 0,
                'damage_dice': dmg_dice,
                'damage_type': _safe_str(cells.get(f'AN{row_num}', '')) or '',
                'ammo': _safe_str(cells.get(f'AQ{row_num}', '')) or '',
                'is_proficient': True,
                'weight': _safe_str(cells.get(f'AA{row_num}', '')) or '',
            })
    result['weapons'] = weapons

    # ━ 护甲 ━
    armor_info = {
        'armor_name': '',
        'armor_ac': _safe_int(cells.get('AD45', 0)) or 0,
        'armor_max_dex': _safe_str(cells.get('AF45', '')) or '',
        'shield_name': '',
        'shield_ac': 0,
        'shield_weight': '',
    }
    result['armor'] = armor_info

    # ━ 法术 ━
    spell_ability = _safe_str(cells.get('AD54', '')) or ''
    if spell_ability in ('施法属性', '施法屬性', None):
        spell_ability = ''
    spell_dc = _safe_int(cells.get('AI58', 10)) or 10
    spell_info = {
        'spellcasting_ability': spell_ability,
        'spell_attack_bonus': 0,
        'spell_ability': spell_ability,
        'spell_save_dc': spell_dc,
        'prepared_spell_count': 0,
    }
    result['spell_info'] = spell_info

    # ━ 法术位 (AM58-AM66 → 1环-9环) ━
    spell_slots = {}
    for level in range(1, 10):
        cell_row = 57 + level  # AM58=1环, AM59=2环, ...
        max_slots = _safe_int(cells.get(f'AM{cell_row}', 0)) or 0
        spell_slots[str(level)] = {'max': max_slots, 'used': 0}
    result['spell_slots'] = spell_slots

    # ━ 已准备法术 — 从法术书 sheet 或 prepared spells 区域读取 ━
    # v4.4.0 的法术在法术书I/II/III sheets
    result['prepared_spells'] = []

    # ━ 钱币 ━
    # 搜索多个可能的钱币位置 (AI列, AE/AG列)
    coins = _find_coins_in_cells(cells, [
        # v4.4.0 AI列 (旧位置)
        {'cp': 'AI43', 'sp': 'AI44', 'ep': 'AI45', 'gp': 'AI46', 'pp': 'AI47'},
        # v4.4.0 AE列 + AG列 (新位置, rows 62-66)
        {'cp': 'AG66', 'sp': 'AG65', 'ep': 'AG63', 'gp': 'AG64', 'pp': 'AG62'},
        {'cp': 'AE66', 'sp': 'AE65', 'ep': 'AE63', 'gp': 'AE64', 'pp': 'AE62'},
    ])
    result['coins'] = coins

    # ━ 负重 ━
    result['weight'] = {
        'light_load': 0, 'medium_load': 0, 'heavy_load': 0, 'current_load': 0,
    }

    # ━ 关键属性（全区域搜索「关键属性」标签）━━
    key_abilities = ''
    import re as _re_ka
    for row_num in range(1, 60):
        for col_idx in range(1, 40):
            col_letter = ''
            n = col_idx
            while n > 0:
                n, r = divmod(n - 1, 26)
                col_letter = chr(65 + r) + col_letter
            val = str(cells.get(f'{col_letter}{row_num}', ''))
            if val.strip() == '关键属性' or val.startswith('关键属性'):
                print(f'[excel-import] 找到关键属性标签: {col_letter}{row_num}')
                # 值在同行右侧列
                for offset in range(1, 10):
                    next_n = col_idx + offset
                    next_col = ''
                    nn = next_n
                    while nn > 0:
                        nn, r = divmod(nn - 1, 26)
                        next_col = chr(65 + r) + next_col
                    next_val = str(cells.get(f'{next_col}{row_num}', '')).strip()
                    if next_val and next_val != '关键属性' and len(next_val) < 30:
                        key_abilities = next_val
                        print(f'[excel-import] 提取值(相邻列): {next_col}{row_num} = "{key_abilities}"')
                        break
                if not key_abilities:
                    m = _re_ka.search(r'关键属性[；;:\s]*(\S+)', val)
                    if m:
                        key_abilities = m.group(1)
                break
        if key_abilities:
            break
    if not key_abilities:
        print(f'[excel-import] 未找到关键属性标签')
    result['key_abilities'] = key_abilities

    # ━ 背景 (从背景 sheet 读取) ━
    result['background'] = {}

    # ━ 特性/专长/职业能力 ━
    # 全表扫描（种族特性可能在前面行,职业能力/专长在后面）
    result['features'] = _import_features(cells, start_row=3, max_row=130)

    return result


def _import_features(cells: dict, start_row: int = 50, max_row: int = 130) -> list[dict]:
    """从主 sheet 提取特性/专长/职业能力/种族特性等词条。

    先扫描 A/B 列找到特性区域的起始行,再从该行开始提取词条。
    支持通过 B 列标签自动推断分类。
    """
    import re as _re

    # ━ 第一步：找到特性区域的起始行 ━
    feature_section_start = 999
    feature_section_labels = [
        '职业能力', '职业特性', '专长', '种族特性', '特殊能力',
        '特性与专长', '职业特性与专长', '职业能力与专长',
    ]
    for row_num in range(3, min(max_row + 1, 200)):
        for col in ('B', 'A'):
            val = str(cells.get(f'{col}{row_num}', '')).strip()
            if val and len(val) < 15:
                for label in feature_section_labels:
                    if val == label or val.startswith(label):
                        if row_num < feature_section_start:
                            print(f'[excel-import] 特性标签匹配: {col}{row_num}="{val}" → {label}, 更新起始行={row_num}')
                            feature_section_start = row_num
                        break
    if feature_section_start == 999:
        feature_section_start = start_row
        print(f'[excel-import] 未找到特性标签,回退起始行={start_row}')
    else:
        print(f'[excel-import] 特性标签匹配完成,起始行={feature_section_start}')

    features = []
    # 分类标签关键词 → 对应分类
    b_label_map = {
        '职业': '职业能力', '职业特性': '职业能力', '职业能力': '职业能力',
        '种族': '种族特性', '种族特性': '种族特性',
        '专长': '专长', '特殊': '特殊能力', '其他': '其他', '特性': '其他',
    }
    current_category = '职业能力'
    # 非数据标签（章节标题/分隔符等）
    skip_exact = {'职业特性', '种族特性', '专长', '特性', '职业能力',
                  '特性与专长', '职业特性与专长', '特性 & 专长',
                  '职业特性 & 专长', '职业能力与专长',
                  '职业能力/特性', '职业特性/专长', '特性/专长',
                  '其他特性', '特殊能力'}
    # 表头关键词（跳过）
    header_keywords = ('名称', '描述', 'Lv', '等级', '来源', '备注')
    # 已知的非特性区域关键词（遇到则停止提取）
    stop_labels = {'法术', '钱币', '负重', '财产', '背包', '装备物品',
                   '已准备', '已学会', '铜币', '银币', '金币',
                   '其他特性', '笔记', '你的角色', '技能分布'}

    b_labels_seen = []  # 诊断
    stop_extraction = False

    for row_num in range(feature_section_start, min(max_row + 1, 200)):
        if stop_extraction:
            break

        # ━ 检测 A/B 列标签 ━
        for label_col in ('B', 'A'):
            label_val = str(cells.get(f'{label_col}{row_num}', '')).strip()
            if not label_val or len(label_val) >= 15:
                continue
            if label_val not in b_labels_seen:
                b_labels_seen.append(label_val)
            # 停止条件：遇到非特性区域的标签
            if any(kw in label_val for kw in stop_labels):
                stop_extraction = True
                break
            # 切换分类
            for kw, cat in b_label_map.items():
                if kw in label_val:
                    current_category = cat
                    break
            if label_val in skip_exact:
                continue  # 跳过纯标签行本身

        if stop_extraction:
            break

        # ━ 尝试从 C/P/B/E/D 列读取词条名 ━
        name = ''
        for col in ('C', 'P', 'B', 'E', 'D'):
            val = cells.get(f'{col}{row_num}', '')
            if val:
                s = str(val).strip()
                if s and s not in skip_exact and len(s) >= 2 and len(s) <= 80:
                    # 过滤掉数值行、分隔符、属性名
                    if (s.replace('.', '').replace('>', '').replace('<', '').replace('+', '')
                          .replace('-', '').replace(' ', '').replace('尺', '').isdigit()):
                        continue
                    name = s
                    break

        if not name:
            continue
        if any(hk in name for hk in header_keywords) and len(name) <= 4:
            continue
        # 跳过元数据行：这些是标签名,不是特性词条
        if name in ('职业', '经验值', '等级', '职业等级', '背景', '种族',
                     '属性', '熟练', '技能', '力量', '敏捷', '体质', '智力', '感知', '魅力'):
            continue
        # 跳过长分隔符行
        if name.startswith('——') or name.startswith('---'):
            continue

        # ━ 描述：从名称列右侧的宽列读取 ━
        desc = ''
        for col in ('U', 'AI', 'T', 'S', 'R', 'Q', 'V', 'W', 'X', 'Y', 'Z',
                     'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH'):
            val = cells.get(f'{col}{row_num}', '')
            if val:
                s = str(val).strip()
                if len(s) > 5:
                    desc = s
                    break

        # 如果宽列没有描述,拼接该行后面列
        if not desc:
            row_texts = []
            for col_idx in range(1, 40):
                col_letter = ''
                n = col_idx
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    col_letter = chr(65 + r) + col_letter
                val = cells.get(f'{col_letter}{row_num}', '')
                if val:
                    row_texts.append(str(val).strip())
            desc = '；'.join(row_texts[1:]) if len(row_texts) > 1 else ''

        # ━ 推断分类 ━
        inferred_cat = current_category
        cat_keywords = [
            (['狂暴', '偷袭', '战斗风格', '回气', '动作如潮', '至圣斩', '引导神力',
              '野蛮', '战技', '先攻', '偷袭攻击', '疾风连击', '气合', '魔力',
              '子职业', '领域', '道途', '范型', '宗派', '宿敌', '熟练探险',
              '武器精通'], '职业能力'),
            (['黑暗视觉', '精灵血统', '矮人', '半身人', '龙族', '侏儒', '半兽',
              '提夫林', '天生', '血脉', '种族武器', '精灵'], '种族特性'),
            (['专长', '警觉', '幸运', '健壮', '巨武器', '神射手', '盾牌大师',
              '双持', '法术射手', '战地施法', '地城探险', '元素导师',
              '起源专长', '战斗风格专长'], '专长'),
        ]
        for keywords, cat in cat_keywords:
            if any(kw in name for kw in keywords):
                inferred_cat = cat
                break

        features.append({
            'name': name,
            'description': desc if desc else '',
            'category': inferred_cat,
            'sort_order': len(features),
        })

    print(f'[excel-import] 特性区域起始行={feature_section_start}, 提取到 {len(features)} 条')
    print(f'[excel-import] B列标签: {b_labels_seen}')
    cats_found = {}
    for f in features:
        cats_found[f['category']] = (cats_found.get(f['category'], 0) + 1)
    print(f'[excel-import] 分类统计: {cats_found}')
    for f in features[:10]:
        d = f.get('description','') or ''
        print(f'  [{f["category"]}] {f["name"]}: {d[:100]}')
    if len(features) > 10:
        print(f'  ... 还有 {len(features)-10} 条')

    return features


def _import_features_from_sheet(wb, result: dict) -> None:
    """从专用的「特性」或「专长」sheet 提取词条。

    如果工作簿中存在名为「特性」「专长」「职业能力」「Features」的 sheet,
    从中提取词条列表并存入 result['features']。
    """
    features = []
    target_sheets = []
    for name in wb.sheetnames:
        lower = name.lower()
        if any(kw in lower for kw in ('特性', '专长', '职业能力', 'feature', 'feat', 'trait')):
            target_sheets.append(name)

    if not target_sheets:
        return

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        cells = _build_cell_dict(ws, max_rows=80)
        features.extend(_import_features(cells, start_row=1, max_row=80))

    if features:
        result['features'] = features
        print(f'[excel-import] 从特性sheet提取到 {len(features)} 条词条')


# ━━━ NekoWorks ver2.1 单元格映射 ━━━

def _import_origin_sheet(wb, result: dict) -> None:
    """DND5E24模板的「起源」sheet：补充年龄/性别/身高/体重/人物形象/
    个性/理念/羁绊/缺陷/背景故事/故乡/背景。

    起源表布局：标签在B:D合并区，值在E列（E3角色名/E4玩家/E5故乡/
    E6背景/E8年龄/K8身高/E9性别/K9体重）；
    人物形象值在B12起；个性/理念/羁绊/缺陷在S列（O列标签定位）；
    背景故事在S17起。
    """
    origin_sheet = None
    for name in wb.sheetnames:
        if '起源' in name:
            origin_sheet = name
            break
    if not origin_sheet:
        return
    ws = wb[origin_sheet]
    cells = _build_cell_dict(ws, max_rows=40)

    basic = result.get('basic', {})
    if not isinstance(basic, dict):
        basic = {}
    bg = result.get('background', {})
    if not isinstance(bg, dict):
        bg = {}

    # 基本信息
    basic['age'] = cells.get('E8', basic.get('age', ''))
    basic['gender'] = cells.get('E9', basic.get('gender', ''))
    basic['height'] = cells.get('K8', basic.get('height', ''))
    basic['weight'] = cells.get('K9', basic.get('weight', ''))
    basic['hometown'] = cells.get('E5', '')
    origin_bg = cells.get('E6', '')
    if origin_bg and str(origin_bg).strip() not in ('背景', '自定义背景'):
        bg['background_feature'] = f'背景：{str(origin_bg).strip()}'

    # 人物形象（B11标签，值B12起）
    appearance = []
    for r in range(12, 20):
        v = cells.get(f'B{r}')
        if v and str(v).strip() and str(v).strip() != '人物形象':
            appearance.append(str(v).strip())
        elif appearance and not v:
            break
    if appearance:
        bg['appearance'] = '\n'.join(appearance)

    # 个性/理念/羁绊/缺陷（O列标签定位，S列取值）
    label_rows = {}
    for r in range(11, 22):
        lbl = cells.get(f'O{r}')
        if lbl and str(lbl).strip() in ('个性', '理念', '羁绊', '缺陷'):
            label_rows[str(lbl).strip()] = r
    sections = [('个性', 'personality_traits'), ('理念', 'ideals'),
                ('羁绊', 'bonds'), ('缺陷', 'flaws')]
    for i, (lbl, key) in enumerate(sections):
        start = label_rows.get(lbl)
        if not start:
            continue
        end = label_rows.get(sections[i + 1][0]) if i + 1 < len(sections) else start + 3
        vals = []
        for rr in range(start, end):
            v = cells.get(f'S{rr}')
            if v and str(v).strip() and str(v).strip() != lbl:
                vals.append(str(v).strip())
        if vals:
            bg[key] = '\n'.join(vals)

    # 背景故事（O17标签，S17起；故乡并入开头）
    story = []
    hometown = basic.get('hometown', '')
    if hometown and str(hometown).strip():
        story.append(f'故乡：{str(hometown).strip()}')
    for r in range(17, 25):
        v = cells.get(f'S{r}')
        if v and str(v).strip() and str(v).strip() != '背景故事':
            story.append(str(v).strip())
    if story:
        bg['backstory'] = '\n'.join(story)

    result['basic'] = basic
    result['background'] = bg


def _import_dnd5e24(ws, cells: dict) -> dict:
    """导入 DND 5E2024 人物卡（悲灵 v1.0.1 模板）

    主要sheet布局（B列=标签，值在标签右侧）：
      - B3=角色名标签,D3=角色名; E4=玩家; Q3=熟练标签,S3=熟练加值
      - E5=主职标签,F5=主职; I5=子职标签,J5=子职; O6=等级
      - B6=主职业标签,D6=主职业; B7/B8=兼职1/2
      - R6-R9=种族/亚种/阵营/信仰标签,S6-S9=值; B9=经验值标签,D9=经验
      - C13-C18=属性名,F13-F18=属性值
      - 战斗: D22=先攻, M22=HP当前, Q22=HP最大, D23=AC, M23=HD骰数,
        P23=HD骰面, N25=体型, T25=速度, D25=法术DC, F26=施法属性, F27=被动察觉
      - 技能: B31-B53=熟练标记(X), C31-C53=技能名, I31-I53=总值
      - 法术: E63-E71=1-9环法术位, K57=法术攻击骰, Q57+=无消耗法术(戏法),
        Y57+=准备法术
    """
    result: dict = {}
    basic: dict = {}
    combat: dict = {}
    abilities: dict = {}

    # ━━ 基本信息（标签在B:D合并区，值在右侧合并区锚点）━━
    basic['name'] = cells.get('E3') or cells.get('D3') or cells.get('C3') or '未命名'
    basic['player'] = cells.get('E4', '')
    basic['class'] = cells.get('E6', '')  # 主职业（B6:D6标签，E6:H6值）
    basic['subclass'] = cells.get('I6', '')  # 子职（I5:N5标签，I6:N6值）
    basic['level'] = _safe_int(cells.get('O6')) or 1  # O5=Lv标签，O6:P6=值
    basic['race'] = cells.get('T6', '')  # R6:S6标签，T6:X6值
    basic['subrace'] = cells.get('T7', '')  # R7:S7标签，T7:X7值
    basic['alignment'] = cells.get('T8', '')  # R8:S8标签
    basic['faith'] = cells.get('T9', '')  # R9:S9标签，T9:X9值
    basic['xp'] = _safe_int(cells.get('E9')) or _safe_int(cells.get('D9')) or 0  # B9:D9标签，E9:N9值
    basic['proficiency_bonus'] = _safe_int(cells.get('S3')) or 2  # Q3:R4标签，S3=值
    # 兼职（B7:D7/B8:D8标签，E7:H7/E8:H8值）
    multiclass = []
    for mc in ('E7', 'E8'):
        v = cells.get(mc, '')
        if v and str(v).strip():
            multiclass.append(str(v).strip())
    basic['multiclass'] = '、'.join(multiclass)
    result['basic'] = basic

    # ━━ 属性 ━━
    attr_rows = {13: 'str', 14: 'dex', 15: 'con', 16: 'int', 17: 'wis', 18: 'cha'}
    for row_num, key in attr_rows.items():
        score = _safe_int(cells.get(f'F{row_num}'))
        abilities[key] = score if score is not None else 10
    result['abilities'] = abilities

    # ━━ 豁免：职业能力表（AX=名称列, BC=描述列）中"豁免熟练"行的
    # 描述（如"智力与感知"）标记熟练属性；加值为属性修正+熟练（最终值）━━
    prof_bonus = basic.get('proficiency_bonus', 2)
    save_profs = {}
    for key in ('str', 'dex', 'con', 'int', 'wis', 'cha'):
        mod = (abilities.get(key, 10) - 10) // 2
        save_profs[key] = {'is_proficient': False, 'save_bonus': mod}
    _attr_cn_map = {'力量': 'str', '敏捷': 'dex', '体质': 'con',
                    '智力': 'int', '感知': 'wis', '魅力': 'cha'}
    for row_num in range(11, 32):
        ax_val = cells.get(f'AX{row_num}')
        if not ax_val or str(ax_val).strip() != '豁免熟练':
            continue
        bc_val = cells.get(f'BC{row_num}')
        if not bc_val:
            break
        save_desc = str(bc_val).strip()
        for cn_name, key in _attr_cn_map.items():
            if cn_name in save_desc:
                save_profs[key]['is_proficient'] = True
                save_profs[key]['save_bonus'] = (abilities.get(key, 10) - 10) // 2 + prof_bonus
        break
    result['save_proficiencies'] = save_profs

    # ━━ 技能 ━━
    skill_profs = {}
    skill_rows = [
        (32, '运动'), (34, '特技'), (35, '巧手'), (36, '隐匿'),
        (38, '调查'), (39, '奥秘'), (40, '历史'), (41, '自然'), (42, '宗教'),
        (44, '察觉'), (45, '洞悉'), (46, '驯兽'), (47, '医药'), (48, '求生'),
        (50, '游说'), (51, '欺瞒'), (52, '威吓'), (53, '表演'),
    ]
    # 标记含义（DND5E24模板）：X=无熟练，O=有熟练，🅞/2/E=双倍熟练（专精）
    for row_num, skill_name in skill_rows:
        prof_mark = cells.get(f'B{row_num}', '')
        mark = str(prof_mark).strip() if prof_mark is not None else ''
        is_exp = mark in ('🅞', '⒪', '㊀', '2', 'E', 'e', 'X2', 'x2')
        is_prof = mark in ('O', 'o', '1', '🅞', '⒪', '㊀', '2', 'E', 'e', 'X2', 'x2')
        if is_prof:
            # 仅输出真正有熟练标记的技能
            skill_profs[skill_name] = {
                'is_proficient': True,
                'is_expertise': is_exp,
                'bonus': _safe_int(cells.get(f'I{row_num}')) or 0,
            }
    result['skill_proficiencies'] = skill_profs

    # ━━ 战斗 ━━
    combat['hp_max'] = _safe_int(cells.get('Q22')) or 10
    combat['hp_current'] = _safe_int(cells.get('M22'))
    if not combat['hp_current']:
        combat['hp_current'] = combat['hp_max']
    combat['ac'] = _safe_int(cells.get('D23')) or 10
    combat['initiative_bonus'] = _safe_int(cells.get('D22')) or 0
    combat['passive_perception'] = _safe_int(cells.get('F27')) or 10
    # 速度：S25='速度'标签，W25:AA25=速度值
    speed_raw = cells.get('W25') or cells.get('T25') or cells.get('S25') or ''
    import re as _re
    speed_match = _re.search(r'(\d+)', str(speed_raw))
    combat['speed'] = int(speed_match.group(1)) if speed_match else 30
    # HD：M23=骰数, P23=骰面
    hd_count = _safe_int(cells.get('M23'))
    combat['hd_count'] = hd_count if hd_count else 1
    hd_die_raw = str(cells.get('P23') or '').strip()
    if hd_die_raw and hd_die_raw.isdigit() and int(hd_die_raw) > 1:
        combat['hit_dice'] = f'1d{hd_die_raw}'
    else:
        combat['hit_dice'] = '1d8'
    result['combat'] = combat

    # ━━ 法术信息 ━━
    spell_ability = cells.get('F26', '') or ''
    if spell_ability in ('属性', '特殊能力'):
        spell_ability = ''
    spell_attack = _parse_attack_bonus(str(cells.get('K57') or '')) if cells.get('K57') else None
    spell_info = {
        'spellcasting_ability': spell_ability,
        'spell_ability': spell_ability,
        'spell_save_dc': _safe_int(cells.get('D25')) or 10,
        'spell_attack_bonus': spell_attack if spell_attack is not None else 0,
        'prepared_spell_count': 0,
    }
    result['spell_info'] = spell_info

    # ━━ 法术位 ━
    # 悲灵 v1.0.1 模板：B62='环阶'标签，B63-B71=1-9环；
    #   G列='已用/最大'输入格（模板默认'/'，填后如'1/2'或纯数字）；
    #   J62='法术值'标签，J63-J71=每环法术位数量（如1级法师J63=2）。
    # 兼容旧布局：E63-E71（早期模板把法术位数值放在E列）。
    spell_slots = {}
    for level in range(1, 10):
        row = 62 + level
        used, max_slots = 0, 0
        # G列优先：'已用/最大'格式
        g_val = str(cells.get(f'G{row}', '') or '').strip()
        g_match = re.match(r'^(\d+)\s*/\s*(\d+)$', g_val)
        if g_match:
            used = int(g_match.group(1))
            max_slots = int(g_match.group(2))
        else:
            if g_val.isdigit():
                max_slots = int(g_val)
            else:
                # J列（法术值）为主，E列为旧布局兜底
                max_slots = (_safe_int(cells.get(f'J{row}'))
                             or _safe_int(cells.get(f'E{row}')) or 0)
        spell_slots[str(level)] = {'max': max_slots, 'used': used}
    result['spell_slots'] = spell_slots

    # ━━ 法术列表：Q57-Q66=无消耗法术(戏法), Y57-Y66=准备法术(X列=等级) ━━
    prepared_spells = []
    for row_num in range(57, 67):
        # Q列 = 无消耗法术（戏法，0环）
        q = cells.get(f'Q{row_num}')
        if q and str(q).strip() and str(q).strip() not in ('法术名称', 'Lv', '无消耗法术'):
            prepared_spells.append({'name': str(q).strip(), 'level': 0})
        # Y列 = 准备法术，X列为环阶
        y = cells.get(f'Y{row_num}')
        if y and str(y).strip() and str(y).strip() not in ('法术名称', 'Lv', '准备法术'):
            lv = _safe_int(cells.get(f'X{row_num}'))
            prepared_spells.append({'name': str(y).strip(), 'level': lv or 0})
    result['prepared_spells'] = prepared_spells

    # ━━ 武器/装备（L列=物品名+类别标签, R列=部位/名称兜底）━━
    weapons = []
    inventory = []
    armor_name = ''
    _BODY_PARTS = {'颈部', '身体', '头部', '手部', '脚部', '腰部', '背部', '手指',
                   '面部', '手臂', '腿部', '戒指', '护腕', '披风', '项链', '头盔',
                   '手套', '靴子', '腰带', '斗篷', '指环'}
    _SKIP_LABELS = {'装备', '名称', '部位', '同调', '加值', '武器名称', '武器',
                    '护甲', '奇物', '背包1', '背包2', '——负重&财产——', '属性',
                    '物品', '价格', '重量', '描述'}
    melee_kw = ['剑', '斧', '锤', '棍', '矛', '戟', '匕首', '镰', '镐', '连枷',
                '弯刀', '军刀', '细剑', '刺剑', '钉头锤', '拳套', '鞭']
    ranged_kw = ['弓', '弩', '枪', '标枪', '飞镖', '投石', '矢', '弹']
    current_cat = '装备'
    for row_num in range(31, 48):
        name_raw = cells.get(f'L{row_num}')
        if name_raw and str(name_raw).strip() in ('护甲', '武器', '奇物'):
            current_cat = str(name_raw).strip()  # 类别标签，切换当前类别
            continue
        if not name_raw or not str(name_raw).strip():
            # L列空则尝试R列（部分模板用R列存名称）
            name_raw = cells.get(f'R{row_num}')
            if not name_raw or not str(name_raw).strip():
                continue
        name = str(name_raw).strip()
        if name in _SKIP_LABELS or name in _BODY_PARTS:
            continue
        if len(name) > 40:
            continue  # 疑似描述文本
        is_weapon = (current_cat == '武器' or any(k in name for k in melee_kw + ranged_kw))
        if is_weapon:
            # 武器数据列：AF=攻击加值, AH=伤害骰, AL=伤害类型, V=特性
            atk_raw = str(cells.get(f'AF{row_num}') or '')
            atk_val = _parse_attack_bonus(atk_raw) if atk_raw else 0
            dmg_dice = str(cells.get(f'AH{row_num}') or '').strip()
            dmg_type = str(cells.get(f'AL{row_num}') or '').strip()
            weapon_effect = str(cells.get(f'V{row_num}') or '').strip()
            weapons.append({
                'name': name,
                'attack_bonus': atk_val or 0,
                'damage_dice': dmg_dice,
                'damage_type': dmg_type,
                'ammo': '',
                'is_proficient': True,
                'weight': '',
                'description': '',
                'effect': weapon_effect,
            })
        else:
            if current_cat == '护甲' and not armor_name:
                armor_name = name
            # 描述：T列=特性（奇物/装备的描述）
            t_desc = str(cells.get(f'T{row_num}') or '').strip()
            inventory.append({
                'item_name': name,
                'quantity': 1,
                'description': t_desc,
                'weight': 0,
                'location': '背包',
            })
    result['weapons'] = weapons
    result['inventory'] = inventory
    if armor_name:
        result['armor'] = {'armor_name': armor_name, 'armor_ac': 0, 'shield_name': '', 'shield_ac': 0}
    else:
        # ━━ 护甲（默认空）━━
        result['armor'] = {'armor_name': '', 'armor_ac': 0, 'shield_name': '', 'shield_ac': 0}

    # ━━ 钱币（AC50=钱包标签, AC51=数值, 视为金币；其余默认0）━━
    result['coins'] = {'cp': 0, 'sp': 0, 'ep': 0, 'gp': _safe_int(cells.get('AC51')) or 0, 'pp': 0}

    # ━━ 背景（模板背景sheet是参考表；背景名暂存backstory，详情留空）━━
    bg_name = cells.get('E6') or cells.get('D6') or ''
    if bg_name and str(bg_name).strip() not in ('背景', '自定义背景', '起源'):
        bg_name = str(bg_name).strip()
    else:
        bg_name = ''
    result['background'] = {
        'personality_traits': '',
        'ideals': '',
        'bonds': '',
        'flaws': '',
        'background_feature': f'背景：{bg_name}' if bg_name else '',
        'appearance': '',
        'backstory': '',
    }

    result['key_abilities'] = spell_ability or ''

    # ━━ 特性/种族特性/专长/职业能力（主要表右侧区域）━━
    features = []
    _feat_re = __import__('re')

    # 种族特性：BT25='种族特性'标签，BT27起（名称/描述），跳过生物种类/体型/速度基础行
    for row_num in range(27, 41):
        name_cell = cells.get(f'BT{row_num}')
        desc_cell = cells.get(f'BZ{row_num}')
        if name_cell and str(name_cell).strip() in ('专长', '战斗风格专长', '特殊能力'):
            break  # 下一区域开始
        if not name_cell or not str(name_cell).strip():
            continue
        name_str = str(name_cell).strip()
        desc_str = str(desc_cell).strip() if desc_cell else ''
        if name_str in ('名称', '种族特性') or not desc_str:
            continue
        if name_str in ('生物种类', '体型', '速度'):
            continue  # 基础信息，非特性
        features.append({'name': name_str, 'category': '种族特性', 'description': desc_str})

    # 专长：BT41='专长'标签，BT列=等级数字，BU列=名称，BZ列=描述
    for row_num in range(43, 55):
        desc_cell = cells.get(f'BZ{row_num}')
        if not desc_cell or not str(desc_cell).strip():
            continue
        desc_str = str(desc_cell).strip()
        name_cell = cells.get(f'BU{row_num}')
        feat_name = ''
        if name_cell and str(name_cell).strip() not in ('名称',):
            feat_name = str(name_cell).strip()
        else:
            # 名称列缺失时从描述第二行中文提取
            for line in desc_str.split('\n')[1:]:
                m = _feat_re.match(r'^([一-鿿]{2,12})', line.strip())
                if m:
                    feat_name = m.group(1)
                    break
        features.append({'name': feat_name or '专长', 'category': '专长', 'description': desc_str})

    # 战斗风格专长：BT55='战斗风格专长'标签，BT56=表头，数据57起
    for row_num in range(57, 62):
        name_cell = cells.get(f'BT{row_num}')
        desc_cell = cells.get(f'BZ{row_num}')
        if not name_cell or not str(name_cell).strip():
            continue
        name_str = str(name_cell).strip()
        desc_str = str(desc_cell).strip() if desc_cell else ''
        if name_str in ('名称', '战斗风格专长') or not desc_str:
            continue
        features.append({'name': name_str, 'category': '战斗风格专长', 'description': desc_str})

    # 特殊能力：BT62='特殊能力'标签，数据63起
    for row_num in range(63, 70):
        name_cell = cells.get(f'BT{row_num}')
        desc_cell = cells.get(f'BZ{row_num}')
        if not name_cell or not str(name_cell).strip():
            continue
        name_str = str(name_cell).strip()
        desc_str = str(desc_cell).strip() if desc_cell else ''
        if name_str in ('名称', '特殊能力') or not desc_str:
            continue
        features.append({'name': name_str, 'category': '特殊能力', 'description': desc_str})

    # 职业能力：AX列=名称列表（AX10='名称'表头, AX11起名称），
    # BC列=对应描述（BC10='描述'表头, BC11起描述）
    for row_num in range(11, 32):
        ax_val = cells.get(f'AX{row_num}')
        bc_val = cells.get(f'BC{row_num}')
        if not bc_val or not str(bc_val).strip():
            continue
        desc_str = str(bc_val).strip()
        if desc_str in ('描述',):
            continue
        name = str(ax_val).strip() if ax_val and str(ax_val).strip() not in ('名称',) else ''
        if not name:
            # 名称列缺失时从描述开头中文提取
            m = _feat_re.match(r'^([一-鿿]{2,10})', desc_str)
            name = m.group(1) if m else '职业能力'
        features.append({'name': name, 'category': '职业能力', 'description': desc_str})

    result['features'] = features
    return result


def _import_nekox21(ws, cells: dict) -> dict:
    """导入 NekoWorks ver2.1 格式的角色卡

    NekoWorks v2.1 布局特点:
      - 武器在行28-32 (P列为名称, AE列为攻击骰)
      - 钱币在行43-48 (AE列为币种)
      - 法术位在 AM39-AM47
    """
    print('[excel-import] _import_nekox21 开始解析')
    result: dict = {}

    # 基本信息
    name = cells.get('E3') or cells.get('B3', '')
    player = cells.get('E4') or cells.get('B4', '')
    cls = cells.get('S3') or cells.get('P3', '')
    race = cells.get('E6') or cells.get('B6', '')
    subrace = cells.get('E7') or cells.get('B7', '')

    # 阵营/信仰 — 过滤掉标签文本
    # NekoWorks v2.1: 标签在X列,值在AA列
    alignment_raw = (cells.get('AA6') or cells.get('L9') or cells.get('X6', ''))
    faith_raw = (cells.get('AA7') or cells.get('L10') or cells.get('X7', ''))
    if alignment_raw in ('阵营', '陣營', None):
        alignment_raw = ''
    if faith_raw in ('信仰', None):
        faith_raw = ''

    basic = {
        'name': name,
        'player': player,
        'class': cls,
        'level': _safe_int(cells.get('Y3', 1)) or 1,
        'xp': _safe_int(cells.get('S4', 0)) or _safe_int(cells.get('P4', 0)) or 0,
        'race': race,
        'gender': cells.get('M6') or cells.get('J6', ''),
        'subrace': subrace,
        'age': _safe_str(cells.get('M7') or cells.get('J7', '')) or '',
        'height': cells.get('E9') or cells.get('Q6', ''),
        'alignment': alignment_raw,
        'weight': cells.get('E10') or cells.get('Q7', ''),
        'faith': faith_raw,
        'proficiency_bonus': _safe_int(cells.get('U9', 2)) or _safe_int(cells.get('AD3', 2)) or 2,
    }
    result['basic'] = basic

    # 属性值
    use_new_rows = cells.get('F14') is not None
    if use_new_rows:
        ability_rows = {k: v for k, v in zip(
            ['str', 'dex', 'con', 'int', 'wis', 'cha'],
            [14, 16, 18, 20, 22, 24]
        )}
    else:
        ability_rows = {k: v for k, v in zip(
            ['str', 'dex', 'con', 'int', 'wis', 'cha'],
            [11, 13, 15, 17, 19, 21]
        )}
    ability_names = {k: v for k, v in zip(
        ['str', 'dex', 'con', 'int', 'wis', 'cha'],
        ['力量', '敏捷', '体质', '智力', '感知', '魅力']
    )}
    abilities = {}
    save_profs = {}
    for key, row_num in ability_rows.items():
        score = _safe_int(cells.get(f'F{row_num}', 10)) or 10
        save_val = _safe_int(cells.get(f'T{row_num}')) or _safe_int(cells.get(f'K{row_num}')) or 0
        abilities[key] = score

        expected_mod = (score - 10) // 2
        name = ability_names[key]
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) or (save_val != expected_mod)
        save_profs[name] = {'is_proficient': is_prof, 'save_bonus': save_val}
    result['abilities'] = abilities
    result['save_proficiencies'] = save_profs

    # 战斗数据
    hp_cur = (_safe_int(cells.get('M29')) or _safe_int(cells.get('Y11')) or 0)
    hp_max = (_safe_int(cells.get('Q29')) or _safe_int(cells.get('AC11')) or hp_cur or 10)
    ac_val = (_safe_int(cells.get('D32')) or _safe_int(cells.get('P13')) or 10)
    init_val = (_safe_int(cells.get('D29')) or _safe_int(cells.get('P11')) or 0)
    speed_str = _safe_str(cells.get('R38', '30尺')) or '30尺'
    speed_match = re.search(r'(\d+)', speed_str)
    speed = int(speed_match.group(1)) if speed_match else 30

    combat = {
        'hp_current': hp_cur,
        'hp_max': hp_max,
        'ac': ac_val,
        'initiative_bonus': init_val,
        'speed': speed,
        'hd_count': (_safe_int(cells.get('AA13')) or _safe_int(cells.get('X13'))
                     or _safe_int(cells.get('L32', 1)) or 1),
        'hd_type': '1d10',  # 游侠默认d10
        'passive_perception': (_safe_int(cells.get('R21')) or _safe_int(cells.get('F40')) or 10),
    }
    result['combat'] = combat

    # 技能熟练 — NekoWorks v2.1使用不同的行号
    # 先判断用哪组行: 检查行26的C列是否有技能名
    use_nekox_skill_rows = cells.get('C26') is not None and str(cells.get('C26', '')).strip() in ('运动', '運動')
    if use_nekox_skill_rows:
        neko_skill_rows = {
            26: '运动', 28: '特技', 29: '巧手', 30: '隐匿',
            32: '调查', 33: '奥秘', 34: '历史', 35: '自然', 36: '宗教',
            38: '察觉', 39: '洞悉', 40: '驯兽', 41: '医疗', 42: '生存',
            44: '游说', 45: '欺诈', 46: '威吓', 47: '表演',
        }
        skill_rows = neko_skill_rows
    else:
        skill_rows = {
            45: '运动', 47: '特技', 48: '巧手', 49: '隐匿',
            51: '调查', 52: '奥秘', 53: '历史', 54: '自然', 55: '宗教',
            57: '察觉', 58: '洞悉', 59: '驯兽', 60: '医药', 61: '生存',
            63: '游说', 64: '欺诈', 65: '威吓', 66: '表演',
        }
    skill_profs = {}
    for row_num, skill_name in skill_rows.items():
        prof_checkbox = cells.get(f'A{row_num}')
        is_prof = bool(prof_checkbox) if prof_checkbox is not None else False
        is_exp = isinstance(prof_checkbox, (int, float)) and prof_checkbox == 2

        # NekoWorks格式: I列为技能加值
        skill_bonus = _safe_int(cells.get(f'I{row_num}'))

        skill_profs[skill_name] = {
            'is_proficient': is_prof or is_exp,
            'is_expertise': is_exp,
            'bonus': skill_bonus or 0,
        }
    result['skill_proficiencies'] = skill_profs

    # 武器 — NekoWorks格式的行号(28-32)与v4.4.0(47-51)不同
    # NekoWorks v2.1 使用 O 列(武器名),v4.4.0 使用 P 列
    # 判断用哪组行: 检查行28的O列或P列是否有武器名
    use_nekox_rows = (
        (cells.get('O28') is not None and str(cells.get('O28', '')).strip()) or
        (cells.get('P28') is not None and str(cells.get('P28', '')).strip())
    )
    weapon_rows = [28, 29, 30, 31, 32] if use_nekox_rows else [47, 48, 49, 50, 51]

    # NekoWorks v2.1: 武器名在O列, 攻击在AE列, 伤害在AI列
    if use_nekox_rows and cells.get('O28') is not None:
        name_col = 'O'
        atk_col = 'AE'
    else:
        name_col = 'P'
        atk_col = 'T'

    str_mod = (abilities.get('str', 10) - 10) // 2
    dex_mod = (abilities.get('dex', 10) - 10) // 2
    prof = basic.get('proficiency_bonus', 2)
    melee_keywords = ['巨斧', '战锤', '长剑', '短剑', '匕首', '斧', '锤', '棍',
                      '矛', '戟', '鞭', '镰', '镐', '连枷', '巨剑', '弯刀',
                      '手斧', '轻锤', '钉头锤', '刺剑', '细剑', '军刀',
                      '细刃', '短剑', '长矛', '长弓', '短弓']
    ranged_keywords = ['弓', '弩', '枪', '标枪', '飞镖', '投石', '矢']
    non_weapon_keywords = ['铜币', '银币', '金币', '白金币', 'CP', 'SP', 'GP', 'EP', 'PP',
                           'Lv', 'QuQ', '重量', '法术', '名称']

    weapons = []
    for row_num in weapon_rows:
        name = cells.get(f'{name_col}{row_num}', '')
        if not (name and str(name).strip()):
            continue
        name_str = str(name).strip()
        # 跳过非武器条目
        if any(kw in name_str for kw in non_weapon_keywords) and len(name_str) < 8:
            continue

        # 解析攻击加值
        if atk_col == 'AE':
            # NekoWorks v2.1: AE列可能是 "D20+" 文本,加值需从技能/属性推算
            atk_raw = _safe_str(cells.get(f'{atk_col}{row_num}', ''))
            atk_val = _parse_attack_bonus(atk_raw) if atk_raw else None
            if atk_val is None:
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                if is_ranged and not is_melee:
                    atk_val = dex_mod + prof
                else:
                    atk_val = str_mod + prof
        else:
            atk_val = _safe_int(cells.get(f'{atk_col}{row_num}'))
            if atk_val is None:
                is_ranged = any(kw in name_str for kw in ranged_keywords)
                is_melee = any(kw in name_str for kw in melee_keywords)
                atk_val = (dex_mod if (is_ranged and not is_melee) else str_mod) + prof

        # NekoWorks v2.1: 伤害骰在AI列, 伤害类型在AN列
        dmg_dice = _safe_str(cells.get(f'AI{row_num}', '')) or ''
        dmg_type = _safe_str(cells.get(f'AN{row_num}', '')) or ''
        ammo = _safe_str(cells.get(f'AQ{row_num}', '')) or ''
        weight = _safe_str(cells.get(f'AA{row_num}', '')) or ''

        weapons.append({
            'name': name_str,
            'attack_bonus': atk_val or 0,
            'damage_dice': dmg_dice,
            'damage_type': dmg_type,
            'ammo': ammo,
            'is_proficient': True,
            'weight': weight,
            'description': '',
            'effect': '',
        })
    result['weapons'] = weapons

    # 护甲
    result['armor'] = {
        'armor_name': '',
        'armor_ac': (_safe_int(cells.get('AD45')) or _safe_int(cells.get('AD25')) or 0),
        'shield_name': '',
        'shield_ac': 0,
    }

    # 法术 — 过滤标签文本
    spell_ability = (_safe_str(cells.get('AD54')) or _safe_str(cells.get('AD35')) or '')
    if spell_ability in ('施法属性', '施法屬性'):
        spell_ability = ''
    spell_info = {
        'spellcasting_ability': spell_ability,
        'spell_ability': spell_ability,
        'spell_save_dc': (_safe_int(cells.get('AI58')) or _safe_int(cells.get('AD40')) or 10),
        'prepared_spell_count': 0,
    }
    result['spell_info'] = spell_info

    # 法术位 — NekoWorks格式用AM39-AM47, v4.4.0用AM58-AM66
    spell_slots = {}
    for level in range(1, 10):
        # 尝试两个位置
        max_slots = (_safe_int(cells.get(f'AM{57+level}'))  # v4.4.0
                     or _safe_int(cells.get(f'AM{38+level}'))  # NekoWorks
                     or 0)
        # 如果 max_slots 等于 level 号,可能是行号标签而非真实法术位
        if max_slots == level:
            max_slots = 0
        spell_slots[str(level)] = {'max': max_slots, 'used': 0}
    result['spell_slots'] = spell_slots

    # 已准备法术
    result['prepared_spells'] = []

    # 钱币 — 尝试多个位置
    # NekoWorks v2.1: AE(币种标签)/AG(数量) 列
    # 也可能在 AI 列
    coins = _find_coins_in_cells(cells, [
        {'cp': 'AG47', 'sp': 'AG46', 'ep': 'AG44', 'gp': 'AG45', 'pp': 'AG43'},
        {'cp': 'AI43', 'sp': 'AI44', 'ep': 'AI45', 'gp': 'AI46', 'pp': 'AI47'},
        {'cp': 'AI47', 'sp': 'AI46', 'ep': 'AI44', 'gp': 'AI45', 'pp': 'AI43'},
    ])
    result['coins'] = coins
    result['weight'] = {'light_load': 0, 'medium_load': 0, 'heavy_load': 0, 'current_load': 0}
    # 关键属性
    key_abilities = ''
    for row_num in range(1, 60):
        for col_idx in range(1, 40):
            col_letter = ''
            n = col_idx
            while n > 0:
                n, r = divmod(n - 1, 26)
                col_letter = chr(65 + r) + col_letter
            val = str(cells.get(f'{col_letter}{row_num}', ''))
            if val.strip() == '关键属性' or val.startswith('关键属性'):
                print(f'[excel-import] 找到关键属性标签: {col_letter}{row_num}')
                for offset in range(1, 10):
                    next_n = col_idx + offset
                    next_col = ''
                    nn = next_n
                    while nn > 0:
                        nn, r = divmod(nn - 1, 26)
                        next_col = chr(65 + r) + next_col
                    next_val = str(cells.get(f'{next_col}{row_num}', '')).strip()
                    if next_val and next_val != '关键属性' and len(next_val) < 30:
                        key_abilities = next_val
                        print(f'[excel-import] 提取值(相邻列): {next_col}{row_num} = "{key_abilities}"')
                        break
                if not key_abilities:
                    m = re.search(r'关键属性[；;:\s]*(\S+)', val)
                    if m:
                        key_abilities = m.group(1)
                break
        if key_abilities:
            break
    if not key_abilities:
        print(f'[excel-import] 未找到关键属性标签')
    result['key_abilities'] = key_abilities
    result['background'] = {}
    result['features'] = _import_features(cells, start_row=3, max_row=130)
    return result


# ━━━ 公共接口 ━━━

def _detect_version(ws) -> str:
    """检测Excel模板版本"""
    # DND 5E2024 人物卡（悲灵版）：A1 含 5E2024 / 5E 2024 标识
    a1 = str(ws.cell(1, 1).value or '')  # A1
    if '5E2024' in a1.upper() or '5E 2024' in a1.upper() or 'DND5E24' in a1.upper():
        print(f'[excel-import] 版本检测结果: dnd5e24 (A1="{a1[:30]}")')
        return 'dnd5e24'
    # 检查 O1 单元格
    o1 = str(ws.cell(1, 15).value or '')  # O1
    if 'v4.4' in o1 or 'New Edition' in o1:
        print(f'[excel-import] 版本检测结果: v440 (O1匹配)')
        return 'v440'
    if 'v4.3' in o1:
        return 'v440'  # 兼容
    if 'NekoWorks' in o1 or 'ver2' in o1:
        return 'neko21'
    # 默认尝试 v4.4.0 (较新的格式)
    # 通过检查是否有 New Edition 特有单元格来判断
    e3 = str(ws.cell(3, 5).value or '')  # E3
    if e3 and e3 not in ('角色名', ''):
        # 如果E3有非标签内容,可能是v4.4.0格式
        b3 = str(ws.cell(3, 2).value or '')  # B3
        if '角色名' in b3:
            return 'v440'
    print(f'[excel-import] 版本检测结果: neko21 (O1="{o1}", E3="{e3}")')
    return 'neko21'


def _build_cell_dict(ws, max_rows: int = 200) -> dict:
    """构建单元格值字典"""
    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                            max_col=min(95, ws.max_column)):
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value
    return cells


def _read_multi_row_cells(cells: dict, col: str, start_row: int, max_empty: int = 2) -> list[str]:
    """读取某列连续行的多个值, 遇到连续空行或B列有新标签时停止"""
    values = []
    empty_count = 0
    label_keywords = {'角色名', '玩家', '年龄', '身高', '体重', '肤色', '发色', '瞳色',
                      '故乡', '出身', '语言', '熟练工具', '个性', '理念', '羁绊', '缺陷',
                      '外貌', '背景', '人物形象'}
    for row_num in range(start_row, start_row + 20):
        # 检查B列是否有新标签（表示进入了新section）
        b_val = str(cells.get(f'B{row_num}', '')).strip().replace('\n', '')
        if b_val:
            # 检查是否是已知标签（非数据行标题）
            is_label = any(kw in b_val for kw in label_keywords)
            if is_label and row_num > start_row:
                break

        val = cells.get(f'{col}{row_num}', '')
        if val:
            val_str = str(val).strip()
            if val_str:
                values.append(val_str)
                empty_count = 0
            else:
                empty_count += 1
        else:
            empty_count += 1

        if empty_count >= max_empty:
            break
    return values


def _import_background_sheet(wb, result: dict) -> None:
    """从背景sheet读取背景信息,支持v4.4.0和NekoWorks v2.1两种格式"""
    bg_sheet = None
    for name in wb.sheetnames:
        if '背景' in name:
            bg_sheet = name
            break

    if not bg_sheet:
        return

    ws = wb[bg_sheet]
    bg = {}
    cells = _build_cell_dict(ws, max_rows=50)

    # 检测格式: v4.4.0在R列有"个性"标签, NekoWorks在E列
    is_v440 = '个性' in str(cells.get('R3', ''))

    # 标签关键词（用于过滤掉纯标签文本）
    label_texts = {'个性', '理念', '羁绊', '缺陷', '背景特性', '背景故事',
                   '外貌描述', '外貌描写', '人物衣装', '人物背景',
                   '角色名', '玩家', '年龄', '身高', '体重',
                   '故乡', '出身', '语言', '熟练工具'}

    if is_v440:
        # ━ v4.4.0 格式: R列=标签, U列=值 ━
        # 个性/理念/羁绊/缺陷/外貌/背景故事在U列
        v440_fields = {
            'U3': 'personality_traits',
            'U5': 'ideals',
            'U6': 'bonds',
            'U7': 'flaws',
            'U9': 'appearance',
            'U18': 'backstory',
        }
        for cell_addr, field in v440_fields.items():
            val = cells.get(cell_addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg[field] = val_str

        # background_feature: R12=标签, 值可能在U12或R13/U13区域
        for addr in ('U12', 'R13', 'U13', 'R14', 'U14'):
            val = cells.get(addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg['background_feature'] = val_str
                    break

        # origin (出身/故乡): B列标签→E列值
        for row in range(3, 15):
            b_label = str(cells.get(f'B{row}', '')).strip()
            if '故乡' in b_label or '出身' in b_label:
                e_val = cells.get(f'E{row}', '')
                if e_val:
                    val_str = str(e_val).strip()
                    if val_str and val_str not in label_texts:
                        bg['origin'] = val_str
                        break

        # languages (语言): B列"语言"标签 → C列连续值
        for row in range(28, 38):
            b_label = str(cells.get(f'B{row}', '')).strip().replace('\n', '')
            if '语' in b_label and '言' in b_label:
                lang_values = _read_multi_row_cells(cells, 'C', row)
                if lang_values:
                    bg['languages'] = '\n'.join(lang_values)
                break

        # tool_proficiencies (熟练工具): B列"工具"标签 → C列连续值
        for row in range(33, 43):
            b_label = str(cells.get(f'B{row}', '')).strip().replace('\n', '')
            if '工具' in b_label and '熟练' in b_label:
                tool_values = _read_multi_row_cells(cells, 'C', row)
                if tool_values:
                    bg['tool_proficiencies'] = '\n'.join(tool_values)
                break

    else:
        # ━ NekoWorks v2.1 格式: E列=标签+值 ━
        neko_bg_fields = {
            'E3': 'personality_traits',
            'E4': 'personality_traits_ext',
            'E5': 'ideals',
            'E6': 'bonds',
            'E7': 'flaws',
            'E9': 'background_feature',
            'E13': 'appearance',
            'E17': 'backstory',
            'E18': 'origin',
            'E20': 'languages',
            'E27': 'tool_proficiencies',
        }
        for cell_addr, field in neko_bg_fields.items():
            val = cells.get(cell_addr, '')
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in label_texts:
                    bg[field] = val_str

    result['background'] = bg


def _import_inventory_sheet(wb) -> list[dict]:
    """从物品/背包sheet读取物品列表

    支持 NekoWorks v2.1 的「物品」sheet 和 v4.4.0 的「背包」sheet。
    物品以 B 列为名称,I 列为描述,逐行列出。
    """
    items = []

    # 查找物品或背包sheet（优先"背包",其次"物品",避免匹配到"装备物品"）
    inv_sheet_name = None
    for name in wb.sheetnames:
        if '背包' in name:
            inv_sheet_name = name
            break
    if not inv_sheet_name:
        for name in wb.sheetnames:
            if '物品' in name and '装备' not in name:
                inv_sheet_name = name
                break

    if not inv_sheet_name:
        return items

    ws = wb[inv_sheet_name]
    cells = _build_cell_dict(ws, max_rows=250)

    # 检测 DND5E24 布局：背包sheet 用 Y列=物品名 / AC列=描述 / AP列=重量 / AS列=数量
    use_dnd5e24_layout = (
        str(cells.get('Y4', '')).strip() == '名称' or
        str(cells.get('Y3', '')).strip() in ('背包1', '背包2') or
        str(cells.get('Y5', '')).strip() not in ('', None)
    )
    if use_dnd5e24_layout:
        items = []
        for row_num in range(5, min(ws.max_row + 1, 80)):
            name = cells.get(f'Y{row_num}')
            if not name or not str(name).strip():
                continue
            name_str = str(name).strip()
            # 跳过表头行（名称/描述/lb/数量）与纯数字
            if name_str in ('名称', '描述', 'lb', '数量', '价格', '项目', '明细', '合计结余'):
                continue
            if name_str.isdigit():
                continue
            desc = str(cells.get(f'AC{row_num}', '')).strip() if cells.get(f'AC{row_num}') else ''
            qty = _safe_int(cells.get(f'AS{row_num}')) or 1
            wt = _safe_int(cells.get(f'AP{row_num}')) or 0
            items.append({
                'item_name': name_str,
                'quantity': qty,
                'weight': wt,
                'location': '背包',
                'description': desc,
                'effect': '',
            })
        return items

    # 跳过关键词
    skip_exact = {'名称', '背包', '描述', 'Inventory', '冒险装备', '工具', '速查'}

    for row_num in range(4, min(ws.max_row + 1, 250)):
        name = cells.get(f'B{row_num}', '')
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()

        # 跳过标题标签行
        if name_str in skip_exact:
            continue
        # 跳过"背包N"或"背包N（...）"等标签
        if (name_str.startswith('背包') and
            (len(name_str) <= 6 or name_str[2:3].isdigit())):
            # 但保留如"背包（次元袋）"这样的实际物品
            if '（' not in name_str and '(' not in name_str:
                continue
        # 跳过纯数字
        if name_str.isdigit():
            continue
        # 跳过"次元袋负重"等元数据
        if '负重' in name_str:
            continue

        # 描述列：NekoWorks用I列，DND5E24模板用F列
        desc = ''
        for dcol in ('I', 'F'):
            dv = cells.get(f'{dcol}{row_num}')
            if dv and str(dv).strip():
                desc = str(dv).strip()
                break
        items.append({
            'item_name': name_str,
            'quantity': 1,
            'weight': 0,
            'location': '背包',
            'description': desc,
            'effect': '',
        })

    return items


def _import_equipment_sheet(wb) -> dict:
    """从装备物品sheet读取护甲、武器和冒险用品

    装备物品sheet包含三张表：
      - 护甲表 (A-G列): 护甲名称/AC/属性
      - 武器表 (I-N列): 武器名称/伤害/属性(含描述)
      - 冒险用品表 (Q-T列): 物品/价格/重量
    这些是通用参考表,不是角色实际装备的物品。
    角色的实际装备在主要sheet的武器区域和物品sheet中。
    此函数用于补充读取可能遗漏的装备描述信息。
    """
    eq_sheet_name = None
    for name in wb.sheetnames:
        if '装备物品' in name:
            eq_sheet_name = name
            break

    if not eq_sheet_name:
        return {}

    ws = wb[eq_sheet_name]
    cells = _build_cell_dict(ws, max_rows=60)

    # 读取武器属性描述（N列）,用于补充主要sheet中武器的描述
    weapon_descriptions = {}
    for row_num in range(3, 40):
        name = cells.get(f'J{row_num}', '')
        prop = cells.get(f'N{row_num}', '')
        if name and str(name).strip():
            name_str = str(name).strip()
            prop_str = str(prop).strip() if prop else ''
            weapon_descriptions[name_str] = prop_str

    return {'weapon_descriptions': weapon_descriptions}


def _sanitize_wps_xlsx(filepath: Path):
    """修复 WPS 表格保存的 xlsx 兼容性问题,返回内存中的文件对象

    WPS 在数据有效性(下拉框)的 sqref 属性中用分号分隔多个区域
    (如 sqref="R19;T19"),而 OOXML 标准要求空格分隔,
    openpyxl 严格校验会抛出 TypeError: expected MultiCellRange。
    此函数将工作表 XML 中 sqref 属性内的分号替换为空格。
    """
    import io
    import zipfile

    def _fix_sqref(match: re.Match) -> bytes:
        return match.group(0).replace(b';', b' ')

    buf = io.BytesIO()
    with zipfile.ZipFile(str(filepath), 'r') as src, \
            zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename.startswith('xl/worksheets/') and info.filename.endswith('.xml'):
                data = re.sub(rb'sqref="[^"]*"', _fix_sqref, data)
            dst.writestr(info, data)
    buf.seek(0)
    return buf


def _load_workbook_compat(filepath: Path, openpyxl):
    """加载工作簿；遇到 WPS 分号 sqref 导致的解析错误时自动修复后重试"""
    try:
        return openpyxl.load_workbook(str(filepath), data_only=True)
    except TypeError:
        # WPS 兼容性问题：修复 sqref 后重试
        try:
            fixed = _sanitize_wps_xlsx(filepath)
            return openpyxl.load_workbook(fixed, data_only=True)
        except Exception:
            raise ValueError(
                f"无法解析该 Excel 文件（可能由 WPS 保存,存在兼容性问题）。\n"
                f"请尝试用 Microsoft Excel 或 LibreOffice 打开并另存为 .xlsx 后重新导入。\n"
                f"文件: {filepath}"
            )


def import_character_from_excel(filepath: str | Path) -> dict:
    """从 DND 5E 人物卡 Excel 文件导入角色数据

    自动检测模板版本并适配对应的单元格映射。

    Args:
        filepath: Excel 文件路径

    Returns:
        包含完整角色数据的字典

    Raises:
        FileNotFoundError: 文件不存在
        ImportError: 缺少 openpyxl 库
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "需要安装 openpyxl 库来读取 Excel 文件。\n"
            "请运行: pip install openpyxl"
        )

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = _load_workbook_compat(filepath, openpyxl)

    # 尝试提取嵌入图片（最佳努力）
    extracted_images = []
    try:
        portrait_dir = filepath.parent
        extracted_images = _extract_images_from_workbook(wb, portrait_dir)
    except Exception:
        pass  # 图片提取失败不影响导入

    # 查找"主要"sheet
    main_sheet = None
    for name in wb.sheetnames:
        if '主要' in name:
            main_sheet = name
            break

    if main_sheet is None:
        if len(wb.sheetnames) > 1:
            main_sheet = wb.sheetnames[1]
        else:
            main_sheet = wb.sheetnames[0]

    ws = wb[main_sheet]

    # 检测版本
    version = _detect_version(ws)

    # 构建单元格字典
    cells = _build_cell_dict(ws)

    # 按版本解析
    if version == 'v440':
        result = _import_v440(ws, cells)
    elif version == 'dnd5e24':
        result = _import_dnd5e24(ws, cells)
        # DND5E24模板补充：起源sheet（年龄/性别/身高/体重/形象/个性等）
        _import_origin_sheet(wb, result)
    else:
        result = _import_nekox21(ws, cells)

    # 读取背景（DND5E24模板的背景sheet是参考表，主sheet已解析）
    if version != 'dnd5e24':
        _import_background_sheet(wb, result)

    # 读取特性/专长（优先专用 sheet,主 sheet 已在版本解析器中提取）
    # DND5E24模板的「专长与据点」sheet是专长参考表，跳过避免误抓
    if version != 'dnd5e24':
        _import_features_from_sheet(wb, result)

    # 读取物品/背包（与主要表装备区的物品合并，去重）
    inventory_items = _import_inventory_sheet(wb)
    if inventory_items:
        existing = result.get('inventory', [])
        seen = {str(i.get('item_name', '')) for i in existing}
        for it in inventory_items:
            if str(it.get('item_name', '')) not in seen:
                existing.append(it)
        result['inventory'] = existing

    # 读取装备物品表（补充武器描述；DND5E24的「装备」表是物品价格参考表，跳过）
    eq_data = _import_equipment_sheet(wb) if version != 'dnd5e24' else {}
    if eq_data.get('weapon_descriptions'):
        # 用装备物品表的属性描述补充武器信息
        wp_descs = eq_data['weapon_descriptions']
        for w in result.get('weapons', []):
            wp_name = w.get('name', '')
            if wp_name in wp_descs and not w.get('description'):
                w['description'] = wp_descs[wp_name]

    # 确保 inventory 键存在
    if 'inventory' not in result:
        result['inventory'] = []

    # 附加提取的图片路径（第一张作为头像）
    if extracted_images:
        result['portrait_path'] = extracted_images[0]

    wb.close()
    return result


def import_and_print_summary(filepath: str | Path) -> str:
    """导入角色并返回摘要文本, 用于预览"""
    data = import_character_from_excel(filepath)
    basic = data.get('basic', {})
    abilities = data.get('abilities', {})
    combat = data.get('combat', {})
    spell_info = data.get('spell_info', {})

    name = _or(basic.get('name'), '未知')
    level = _or(basic.get('level'), 1)
    cls = _or(basic.get('class'), '未知')
    race = _or(basic.get('race'), '未知')
    alignment = _or(basic.get('alignment'), '未知')
    faith = _or(basic.get('faith'), '—')
    hp_cur = _or(combat.get('hp_current'), 0)
    hp_max = _or(combat.get('hp_max'), 0)
    ac = _or(combat.get('ac'), 10)
    init_bonus = _or(combat.get('initiative_bonus'), 0)
    speed = _or(combat.get('speed'), 30)

    lines = [
        f"📜 角色卡预览: {name}",
        f"   等级: {level} | 职业: {cls} | 种族: {race}",
        f"   阵营: {alignment} | 信仰: {faith}",
        f"   ❤️ HP: {hp_cur}/{hp_max} | 🛡️ AC: {ac}",
        f"   ⚡ 先攻: {init_bonus:+d} | 🏃 速度: {speed}ft",
    ]

    # 属性
    abbr = {'str': '力量', 'dex': '敏捷', 'con': '体质',
            'int': '智力', 'wis': '感知', 'cha': '魅力'}
    ab_parts = []
    for key, label in abbr.items():
        score = _or(abilities.get(key), 10)
        mod_val = (score - 10) // 2 if score else 0
        ab_parts.append(f"{label}:{score}({mod_val:+d})")
    lines.append(f"   📊 {' | '.join(ab_parts)}")

    # 法术
    spell_ability = _or(spell_info.get('spellcasting_ability'), '')
    if spell_ability:
        spell_dc = _or(spell_info.get('spell_save_dc'), '—')
        lines.append(f"   ✨ 施法: {spell_ability} | DC: {spell_dc}")

    # 武器
    weapons = data.get('weapons', [])
    if weapons:
        wp_lines = []
        for w in weapons:
            wp_lines.append(f"{w['name']} ({w.get('damage_dice', '?')})")
        lines.append(f"   ⚔️ 武器: {', '.join(wp_lines)}")

    # 已准备法术
    prepared = data.get('prepared_spells', [])
    if prepared:
        spell_names = [s['name'] for s in prepared[:8]]
        more = f"... 等{len(prepared)}个" if len(prepared) > 8 else ""
        lines.append(f"   📖 已准备法术: {', '.join(spell_names)}{more}")

    # 物品
    inventory = data.get('inventory', [])
    if inventory:
        item_names = [it['item_name'] for it in inventory[:10]]
        more = f"... 等{len(inventory)}件" if len(inventory) > 10 else ""
        lines.append(f"   🎒 物品: {', '.join(item_names)}{more}")

    return "\n".join(lines)
