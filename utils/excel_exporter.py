# -*- coding: utf-8 -*-
"""角色 → 悲灵模板（DND5E24人物卡）Excel 导出

与 utils/excel_importer.py 的 DND5E24 导入互为逆向：
导入读哪些单元格（_import_dnd5e24 / _import_origin_sheet），导出就写哪些单元格。
仅写入非空字段，模板默认值/样式/数据验证保留；每次从原模板复制后写入。
"""
import re as _re
import warnings
from pathlib import Path

# ━━ 技能行映射（模板行号 → 平台技能名；模板标签为"求生/欺瞒"，平台为"生存/欺诈"）━━
_SKILL_ROWS = {
    '运动': 32, '特技': 34, '巧手': 35, '隐匿': 36,
    '调查': 38, '奥秘': 39, '历史': 40, '自然': 41, '宗教': 42,
    '察觉': 44, '洞悉': 45, '驯兽': 46, '医药': 47,
    '生存': 48,  # 模板标签"求生"
    '游说': 50, '欺诈': 51,  # 模板标签"欺瞒"
    '威吓': 52, '表演': 53,
}

_AB_CN = {'str': '力量', 'dex': '敏捷', 'con': '体质', 'int': '智力', 'wis': '感知', 'cha': '魅力'}

# ━━ 特性分类 → 模板区域（列=名称列，描述列 BZ；起始行）━━
_FEATURE_AREAS = (
    ('racial_trait', '种族特性', 'BT', 27),   # 跳过模板"生物种类/体型/速度"占位行
    ('feat', '专长', 'BU', 43),
    ('special_ability', '特殊能力', 'BT', 63),
)
_FEATURE_SKIP = ('名称', '生物种类', '体型', '速度', '种族特性', '专长', '特殊能力', '等级', '战斗风格专长')


def _write(ws, cell: str, value):
    """写单元格；合并单元格（MergedCell 只读）或空值跳过"""
    if value is None or value == '':
        return
    from openpyxl.cell.cell import MergedCell
    target = ws[cell]
    if isinstance(target, MergedCell):
        return
    target.value = value


def _find_writable_row(ws, cols, start: int, end: int) -> int:
    """在 [start, end] 区间找可写空行：指定列均非合并单元格且首列为空
    （跳过模板占位标签行）；找不到返回 -1"""
    from openpyxl.cell.cell import MergedCell
    for r in range(start, end + 1):
        if any(isinstance(ws[f'{c}{r}'], MergedCell) for c in cols):
            continue
        v = ws[f'{cols[0]}{r}'].value
        if v is None or str(v).strip() == '':
            return r
        if str(v).strip() in _FEATURE_SKIP:
            continue
    return -1


def _fmt_bonus(value):
    """数字 → '+5' / '-1' 格式"""
    try:
        n = int(value)
        return f'+{n}' if n >= 0 else str(n)
    except (TypeError, ValueError):
        return value


def _hd_die_face(hit_dice):
    """'d8' / '2d10' → 骰面 8 / 10"""
    if not hit_dice:
        return None
    m = _re.search(r'd(\d+)', str(hit_dice), _re.I)
    return int(m.group(1)) if m else None


def export_to_belling(data: dict, template_path, out_path) -> Path:
    """把标准化角色字典（core.character._char_to_dict 输出同构）写入悲灵模板。

    data 键参考：name/player/class/subclass/level/xp/race/subrace/alignment/faith、
    abilities{str..cha}/hp_max/hp_current/ac/initiative_bonus/speed/hit_dice/hd_count/
    passive_perception/proficiency_bonus/spell_save_dc/spellcasting_ability/
    spell_attack_bonus/skill_proficiencies/save_proficiencies/weapons/armor_name/
    spell_slots/prepared_spells/features{racial_trait,feat,special_ability,class_feature}/
    age/gender/height/weight_field/background{personality_traits,ideals,bonds,flaws,backstory,appearance}
    """
    warnings.filterwarnings('ignore')  # 静默 openpyxl Data Validation 扩展警告
    import openpyxl

    wb = openpyxl.load_workbook(str(template_path))
    main_ws = wb['主要'] if '主要' in wb.sheetnames else wb[wb.sheetnames[1]]
    origin_ws = None
    for name in wb.sheetnames:
        if '起源' in name:
            origin_ws = wb[name]
            break

    # ━━ 基本信息 ━━
    _write(main_ws, 'E3', data.get('name'))
    _write(main_ws, 'E4', data.get('player'))
    _write(main_ws, 'S3', data.get('proficiency_bonus'))
    _write(main_ws, 'E6', data.get('class'))
    _write(main_ws, 'I6', data.get('subclass'))
    _write(main_ws, 'O6', data.get('level'))
    _write(main_ws, 'E9', data.get('xp'))
    _write(main_ws, 'T6', data.get('race'))
    _write(main_ws, 'T7', data.get('subrace'))
    _write(main_ws, 'T8', data.get('alignment'))
    _write(main_ws, 'T9', data.get('faith'))

    # ━━ 属性（F13-F18）━━
    abilities = data.get('abilities') or {}
    for i, key in enumerate(('str', 'dex', 'con', 'int', 'wis', 'cha')):
        v = abilities.get(key)
        if v not in (None, ''):
            _write(main_ws, f'F{13 + i}', int(v))

    # ━━ 战斗 ━━
    _write(main_ws, 'Q22', data.get('hp_max'))
    _write(main_ws, 'M22', data.get('hp_current'))
    # 注：临时 HP（U22:V23 标签合并块）、抗性/免疫/优势（AC25:AF27 标签合并块）、
    # 背景（K12:L12 标签合并块）、钱币（AC51:AE51 为公式自算区，输入区结构不明）——
    # 模板无独立值格，不写入
    _write(main_ws, 'D23', data.get('ac'))
    _write(main_ws, 'D22', data.get('initiative_bonus'))
    _write(main_ws, 'M23', data.get('hd_count'))
    _write(main_ws, 'P23', _hd_die_face(data.get('hit_dice')))
    _write(main_ws, 'N25', data.get('size'))
    _write(main_ws, 'W25', data.get('speed'))
    _write(main_ws, 'F27', data.get('passive_perception'))
    _write(main_ws, 'D25', data.get('spell_save_dc'))
    _write(main_ws, 'F26', data.get('spellcasting_ability'))
    _write(main_ws, 'K57', _fmt_bonus(data.get('spell_attack_bonus')) if data.get('spell_attack_bonus') not in (None, '') else None)

    # ━━ 豁免加值（T12='豁免'标签，T13-T18 = 力/敏/体/智/感/魅；
    # 有记录用存储值，无记录用属性修正——与角色卡显示逻辑一致）━━
    saves = data.get('save_proficiencies') or {}
    ab_mods = data.get('ability_mods') or {}
    for i, key in enumerate(('str', 'dex', 'con', 'int', 'wis', 'cha')):
        prof = saves.get(key) or {}
        bonus = prof.get('save_bonus') if prof.get('is_proficient') else (ab_mods.get(key) or 0)
        if bonus not in (None, ''):
            _write(main_ws, f'T{13 + i}', int(bonus))

    # ━━ 工具熟练（AX14='工具熟练'标签 → BC14 描述）━━
    _write(main_ws, 'BC14', (data.get('background') or {}).get('tool_proficiencies'))

    # ━━ 钱币（AI50 附近填写区：AI50=GP / AI51=EP / AI52=SP / AI53=CP / AP50=PP，
    # 由 AN52 总财富公式 =GP*1+EP/2+SP/10+CP/100+PP*10 反推）━━
    _write(main_ws, 'AI50', data.get('gp'))
    _write(main_ws, 'AI51', data.get('ep'))
    _write(main_ws, 'AI52', data.get('sp'))
    _write(main_ws, 'AI53', data.get('cp'))
    _write(main_ws, 'AP50', data.get('pp'))

    # ━━ 技能熟练（B 列标记 O + I 列总值）━━
    skills = data.get('skill_proficiencies') or {}
    for name, prof in skills.items():
        row = _SKILL_ROWS.get(name)
        if not row or not prof or not prof.get('is_proficient'):
            continue
        _write(main_ws, f'B{row}', 'O')
        bonus = prof.get('bonus')
        if bonus not in (None, ''):
            _write(main_ws, f'I{row}', int(bonus))

    # ━━ 豁免熟练（模板预置 AX11="豁免熟练"行，直接写其 BC 描述"力量与敏捷"；
    # 模板无该行时在 AX/BC 区找空行）━━
    saves = data.get('save_proficiencies') or {}
    save_names = [_AB_CN[k] for k in _AB_CN if (saves.get(k) or {}).get('is_proficient')]
    if save_names:
        row = next((r for r in range(11, 41)
                    if str(main_ws[f'AX{r}'].value or '').strip() == '豁免熟练'), None)
        if row is None:
            row = _find_writable_row(main_ws, ('AX', 'BC'), 11, 40)
        if row and row > 0:
            _write(main_ws, f'BC{row}', '与'.join(save_names))

    # ━━ 武器（L 列名称 + AF 命中 + AH 伤害骰 + AL 类型 + V 特效）━━
    # 模板 L31-L48 区 AF/AH/AL/V 列存在合并单元格（仅 31-37 行部分可写），跳行处理
    weapons = data.get('weapons') or []
    if weapons:
        start = _find_writable_row(main_ws, ('L', 'AF', 'AH', 'AL', 'V'), 31, 48)
        if start < 0:
            start = 31
        for i, w in enumerate(weapons):
            row = start + i
            if row > 48:
                break
            _write(main_ws, f'L{row}', w.get('name'))
            _write(main_ws, f'AF{row}', _fmt_bonus(w.get('attack_bonus')))
            _write(main_ws, f'AH{row}', w.get('damage_dice'))
            _write(main_ws, f'AL{row}', w.get('damage_type'))
            _write(main_ws, f'V{row}', w.get('effect'))
    # 护甲（L 列）
    armor_name = data.get('armor_name') or ''
    if armor_name:
        row = _find_writable_row(main_ws, ('L',), 31, 48)
        if row < 0:
            row = 31
        _write(main_ws, f'L{row}', armor_name)

    # ━━ 背包物品（「背包」sheet：Y 名称 / AC 描述 / AP 重量磅 / AS 数量，第 5 行起）━━
    inventory = data.get('inventory') or []
    if inventory:
        inv_ws = None
        for name in wb.sheetnames:
            if '背包' in name:
                inv_ws = wb[name]
                break
        if inv_ws is not None:
            row = 5
            for it in inventory:
                item_name = it.get('item_name') or it.get('name')
                if not item_name or row > 79:
                    break
                _write(inv_ws, f'Y{row}', item_name)
                _write(inv_ws, f'AC{row}', it.get('description'))
                wt = it.get('weight')
                if wt not in (None, ''):
                    _write(inv_ws, f'AP{row}', float(wt))
                qty = it.get('quantity')
                if qty not in (None, ''):
                    _write(inv_ws, f'AS{row}', int(qty))
                row += 1

    # ━━ 法术位（E 列最大数 + G 列"已用/最大"）━━
    slots = data.get('spell_slots') or {}
    for level in range(1, 10):
        s = slots.get(str(level)) or {}
        max_slots = s.get('max') or 0
        if max_slots > 0:
            row = 62 + level
            _write(main_ws, f'E{row}', max_slots)
            _write(main_ws, f'G{row}', f"{s.get('used') or 0}/{max_slots}")

    # ━━ 法术（Q 列戏法 0 环 / Y 列准备法术 + X 列环阶；跳过 57 表头行）━━
    spells = data.get('prepared_spells') or []
    q_row, y_row = 58, 58
    for sp in spells:
        name = sp.get('spell_name')
        if not name:
            continue
        level = int(sp.get('spell_level') or 0)
        if level == 0:
            if q_row <= 66:
                _write(main_ws, f'Q{q_row}', name)
                q_row += 1
        else:
            if y_row <= 66:
                _write(main_ws, f'Y{y_row}', name)
                _write(main_ws, f'X{y_row}', level)
                y_row += 1

    # ━━ 特性（种族特性 BT / 专长 BU / 特殊能力 BT，描述列 BZ）━━
    features = data.get('features') or {}
    for cat_key, _label, col, start_row in _FEATURE_AREAS:
        items = features.get(cat_key) or []
        if not items:
            continue
        row = _find_writable_row(main_ws, (col, 'BZ'), start_row, start_row + 12)
        if row < 0:
            row = start_row
        for i, f in enumerate(items):
            r = row + i
            if r > start_row + 12:
                break
            _write(main_ws, f'{col}{r}', f.get('name'))
            _write(main_ws, f'BZ{r}', f.get('description'))
    # 职业能力（AX/BC 区）
    class_feats = features.get('class_feature') or []
    if class_feats:
        row = _find_writable_row(main_ws, ('AX', 'BC'), 11, 40)
        if row > 0:
            for i, f in enumerate(class_feats):
                r = row + i
                if r > 40:
                    break
                _write(main_ws, f'AX{r}', f.get('name'))
                _write(main_ws, f'BC{r}', f.get('description'))

    # ━━ 起源 sheet（年龄/性别/身高/体重/形象/个性/理念/羁绊/缺陷/背景故事/
    # 故乡/语言）━━
    if origin_ws is not None:
        _write(origin_ws, 'E8', data.get('age'))
        _write(origin_ws, 'E9', data.get('gender'))
        _write(origin_ws, 'K8', data.get('height'))
        _write(origin_ws, 'K9', data.get('weight_field'))
        _write(origin_ws, 'E5', (data.get('background') or {}).get('origin'))  # 故乡
        _write(origin_ws, 'B24', (data.get('background') or {}).get('languages'))  # 语言（B23 标签下）
        bg = data.get('background') or {}
        _write(origin_ws, 'S11', bg.get('personality_traits'))
        _write(origin_ws, 'S13', bg.get('ideals'))
        _write(origin_ws, 'S14', bg.get('bonds'))
        _write(origin_ws, 'S15', bg.get('flaws'))
        _write(origin_ws, 'S17', bg.get('backstory'))
        appearance = bg.get('appearance') or ''
        if appearance:
            lines = [ln.strip() for ln in str(appearance).split('\n') if ln.strip()]
            for i, ln in enumerate(lines[:8]):
                _write(origin_ws, f'B{12 + i}', ln)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
